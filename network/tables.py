import json
import os
import shutil
from typing import Any

from django.conf import settings
from django.template.loader import render_to_string

from network.community import labelgroup_display_labels, strategy_display_label
from network.community_stats import PARTITION_COMPARISON_METRICS, network_summary_rows
from network.measures._registry import role_companions
from network.utils import CommunityTableData, GraphData

import openpyxl
from openpyxl.styles import Font, PatternFill


def _pulpit_ctx() -> dict:
    # Display names for the manual LABELGROUP<id> partitions, injected into every table page as
    # window.STRATEGY_LABELS so labels.js → strategy_label() shows the analyst's group name ("Area")
    # instead of a title-cased key ("Labelgroup1"). Escape "</" so a group name containing "</script>"
    # can't break out of the inline <script> (mirrors network.exporter._patch_html_file).
    strategy_labels_json = json.dumps(labelgroup_display_labels()).replace("</", "<\\/")
    return {
        "repository_url": getattr(settings, "REPOSITORY_URL", ""),
        "app_version": getattr(settings, "APP_VERSION", ""),
        "strategy_labels_json": strategy_labels_json,
    }


def _write_page(
    template: str,
    output_filename: str,
    *,
    seo: bool,
    project_title: str,
    title_part: str,
    seo_title_part: str | None = None,
    description: str = "",
) -> None:
    """Render a standard pulpit HTML page.

    ``title_part`` is the human-facing page name used in the rendered title.
    ``seo_title_part`` overrides ``title_part`` when SEO mode is on (used by
    pages where the public/search-engine title differs from the in-app one).
    ``project_title``, when set, prefixes the title as ``"<project> | <part>"``.
    """
    label = seo_title_part if (seo and seo_title_part) else title_part
    title = f"{project_title} | {label}" if project_title else label
    context = {
        "title": title,
        "robots_meta": "index, follow" if seo else "noindex, nofollow",
        "description": description,
        **_pulpit_ctx(),
    }
    content = render_to_string(template, context)
    with open(output_filename, "w") as f:
        f.write(content)


_BASE_MEASURE_KEYS: frozenset[str] = frozenset({"in_deg", "out_deg", "fans", "messages_count"})


def write_table_html(
    graph_data: GraphData,
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    n = len(graph_data["nodes"])
    _write_page(
        "network/channel_table.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Channels",
        seo_title_part="Channel network data",
        description=(
            f"Network data for {n} Telegram channels, "
            "including activity metrics, inbound and outbound tie strength, and community assignments."
        ),
    )


def write_table_xlsx(
    graph_data: GraphData,
    measures_labels: list[tuple[str, str]],
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
    year_data: "list[tuple[int, GraphData]] | None" = None,
) -> None:
    extra = [(k, lbl) for k, lbl in measures_labels if k not in _BASE_MEASURE_KEYS]
    pagerank_col = next(((k, lbl) for k, lbl in extra if k == "pagerank"), None)
    other_extra = [(k, lbl) for k, lbl in extra if k != "pagerank"]
    # Categorical role label (Module role) rides alongside its numeric measure (within_module_z*)
    # — one column per role-measure instance, carrying the instance's parameter annotation.
    role_label_cols: list[tuple[str, str]] = []
    for key, label in measures_labels:
        comp = role_companions(key)
        if comp:
            annot = label[label.find(" (") :] if " (" in label else ""
            role_label_cols.append((comp["role_key"], comp["role_label"] + annot))

    headers = ["Channel", "URL", "Label", "Users", "Messages", "Inbound", "Outbound"]
    if pagerank_col:
        headers.append(pagerank_col[1])
    headers += [lbl for _, lbl in other_extra]
    headers += [hdr for _, hdr in role_label_cols]
    headers += [strategy_display_label(s) for s in strategies]
    headers += ["Activity start", "Activity end"]

    def _fill(ws: Any, gd: GraphData) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for node in sorted(gd["nodes"], key=lambda n: n.get("in_deg") or 0, reverse=True):
            communities = node.get("communities") or {}
            row: list[Any] = [
                node.get("label") or node["id"],
                node.get("url") or "",
                node.get("organization") or "",
                node.get("fans"),
                node.get("messages_count"),
                node.get("in_deg"),
                node.get("out_deg"),
            ]
            if pagerank_col:
                row.append(node.get(pagerank_col[0]))
            for key, _ in other_extra:
                row.append(node.get(key))
            for role_key, _ in role_label_cols:
                row.append(node.get(role_key) or "")
            for s in strategies:
                row.append(communities.get(s, ""))
            row.append(node.get("activity_start") or "")
            row.append(node.get("activity_end") or "")
            ws.append(row)

    wb = openpyxl.Workbook()
    wb.properties.creator = "Pulpit"
    if project_title:
        wb.properties.title = project_title
    ws = wb.active
    ws.title = "All" if year_data else "Channels"
    _fill(ws, graph_data)

    if year_data:
        for yr, yr_gd in year_data:
            _fill(wb.create_sheet(title=str(yr)), yr_gd)

    wb.save(output_filename)


def write_network_metrics_json(
    community_table_data: CommunityTableData,
    strategies: list[str],
    graph_dir: str,
) -> None:
    def _fmt(val: float | None, decimals: int = 4) -> str:
        return "—" if val is None else f"{val:.{decimals}f}"

    data_dir = os.path.join(graph_dir, "data")
    summary = community_table_data["network_summary"]
    summary_rows = []
    for label, value, group in network_summary_rows(summary):
        if isinstance(value, float):
            display = _fmt(value)
        elif value is None:
            display = "—"
        else:
            display = str(value)
        summary_rows.append({"label": label, "value": display, "group": group})

    modularity_rows = []
    for strategy_key in strategies:
        entry = community_table_data["strategies"].get(strategy_key)
        mod = entry["modularity"] if entry else None
        icr = entry.get("inter_community_edge_ratio") if entry else None
        mei = entry.get("mean_ei_index") if entry else None
        modularity_rows.append(
            {
                "strategy": strategy_key,
                "modularity": _fmt(mod) if mod is not None else "—",
                "inter_community_ratio": _fmt(icr) if icr is not None else "—",
                "mean_ei": _fmt(mei) if mei is not None else "—",
            }
        )

    payload = {
        "wcc_note_visible": not summary["path_on_full"],
        "scc_note_visible": (
            summary.get("avg_path_length_directed") is not None and not summary.get("scc_path_on_full", True)
        ),
        "summary_rows": summary_rows,
        "modularity_rows": modularity_rows,
        "partition_comparison": community_table_data.get("partition_comparison"),
    }
    with open(os.path.join(data_dir, "network_metrics.json"), "w") as f:
        f.write(json.dumps(payload))


def write_network_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/network_table.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Network",
        seo_title_part="Network statistics",
    )


def write_network_table_xlsx(
    community_table_data: CommunityTableData,
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
    year_data: "list[tuple[int, CommunityTableData]] | None" = None,
) -> None:
    def _fill(ws: Any, ctd: CommunityTableData) -> None:
        summary = ctd["network_summary"]
        ws.append(["Metric", "Value"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for label, value, _group in network_summary_rows(summary):
            ws.append([label, value])
        if not summary["path_on_full"]:
            ws.append([])
            ws.append(["† Computed on the largest weakly connected component (undirected)"])
        if summary.get("avg_path_length_directed") is not None and not summary.get("scc_path_on_full", True):
            ws.append([])
            ws.append(["‡ Computed on the largest strongly connected component (directed)"])
        ws.append([])
        ws.append(["Strategy", "Modularity"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for strategy_key in strategies:
            entry = ctd["strategies"].get(strategy_key)
            ws.append([strategy_display_label(strategy_key), entry["modularity"] if entry else None])

        comparison = ctd.get("partition_comparison")
        if comparison and len(comparison.get("strategies", [])) >= 2:
            cmp_strats = comparison["strategies"]
            cmp_labels = [strategy_display_label(s) for s in cmp_strats]
            for metric_key, abbr, name, _is_distance in PARTITION_COMPARISON_METRICS:
                cells = comparison["metrics"].get(metric_key)
                if not cells:
                    continue
                ws.append([])
                ws.append([f"{name} ({abbr})"])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                ws.append([""] + cmp_labels)
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                for i, lbl in enumerate(cmp_labels):
                    ws.append([lbl] + [cells[i][j] for j in range(len(cmp_strats))])

    wb = openpyxl.Workbook()
    wb.properties.creator = "Pulpit"
    if project_title:
        wb.properties.title = project_title
    ws = wb.active
    ws.title = "All" if year_data else "Network"
    _fill(ws, community_table_data)

    if year_data:
        for yr, yr_ctd in year_data:
            _fill(wb.create_sheet(title=str(yr)), yr_ctd)

    wb.save(output_filename)


_COMPARE_RENAMES: dict[str, str] = {
    "graph.html": "graph_2.html",
    "graph3d.html": "graph3d_2.html",
    "channel_table.html": "channel_table_2.html",
    "network_table.html": "network_table_2.html",
    "community_table.html": "community_table_2.html",
    "consensus_matrix.html": "consensus_matrix_2.html",
    "structural_similarity.html": "structural_similarity_2.html",
    "behavioural_equivalence.html": "behavioural_equivalence_2.html",
    "channel_table.xlsx": "channel_table_2.xlsx",
    "network_table.xlsx": "network_table_2.xlsx",
    "community_table.xlsx": "community_table_2.xlsx",
}

# Ordered so that graph3d.html is replaced before graph.html (avoid partial match)
_HTML_LINK_RENAMES: list[tuple[str, str]] = [
    ("graph3d.html", "graph3d_2.html"),
    ("graph.html", "graph_2.html"),
    ("channel_table.html", "channel_table_2.html"),
    ("network_table.html", "network_table_2.html"),
    ("community_table.html", "community_table_2.html"),
    ("consensus_matrix.html", "consensus_matrix_2.html"),
    ("structural_similarity.html", "structural_similarity_2.html"),
    ("behavioural_equivalence.html", "behavioural_equivalence_2.html"),
    ("channel_table.xlsx", "channel_table_2.xlsx"),
    ("network_table.xlsx", "network_table_2.xlsx"),
    ("community_table.xlsx", "community_table_2.xlsx"),
]


def _patch_compare_html(content: str) -> str:
    """Patch an HTML file from the compare project: rewrite internal links and inject DATA_DIR."""
    for old, new in _HTML_LINK_RENAMES:
        content = content.replace(old, new)
    injection = '<script>window.DATA_DIR = "data_2/";</script>\n'
    for marker in ('<script src="js/', '<script type="module" src="js/'):
        idx = content.find(marker)
        if idx != -1:
            content = content[:idx] + injection + content[idx:]
            break
    return content


def copy_compare_project(compare_dir: str, graph_dir: str) -> set[str]:
    """Copy files from a compare graph/ directory into graph/, renaming with _2 suffix.

    Returns the set of destination filenames that were actually written.
    """
    copied: set[str] = set()

    # data/ → data_2/
    src_data = os.path.join(compare_dir, "data")
    dst_data = os.path.join(graph_dir, "data_2")
    if os.path.exists(dst_data):
        shutil.rmtree(dst_data)
    if os.path.isdir(src_data):
        shutil.copytree(src_data, dst_data)
        copied.add("data_2")

    for src_name, dst_name in _COMPARE_RENAMES.items():
        src = os.path.join(compare_dir, src_name)
        if not os.path.isfile(src):
            continue
        dst = os.path.join(graph_dir, dst_name)
        if src_name.endswith(".html"):
            with open(src) as f:
                content = f.read()
            with open(dst, "w") as f:
                f.write(_patch_compare_html(content))
        else:
            shutil.copy2(src, dst)
        copied.add(dst_name)

    return copied


def write_timeline_json(timeline_entries: list[dict], graph_dir: str) -> None:
    data_dir = os.path.join(graph_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    # Strip private implementation keys (e.g. _xlsx_graph_data, _xlsx_community_data)
    # that are passed through timeline_entries for XLSX assembly but must not appear in JSON output.
    clean = [{k: v for k, v in e.items() if not k.startswith("_")} for e in timeline_entries]
    with open(os.path.join(data_dir, "timeline.json"), "w") as f:
        f.write(json.dumps({"years": clean}))


def write_network_compare_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/network_compare_table.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Network comparison",
    )


def write_community_table_xlsx(
    community_table_data: CommunityTableData,
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
    year_data: "list[tuple[int, CommunityTableData]] | None" = None,
) -> None:
    headers = [
        "Community",
        "Color",
        "Nodes",
        "Internal Edges",
        "External Edges",
        "E-I Index",
        "Density",
        "Reciprocity",
        "Avg Clustering",
        "Avg Path Length",
        "Diameter",
        "Channels",
    ]

    def _fill_strategy(ws: Any, strategy_key: str, ctd: CommunityTableData, first: bool = True) -> None:
        strategy_entry = ctd["strategies"].get(strategy_key)
        if not strategy_entry:
            return
        if not first:
            ws.append([])  # blank separator between strategies
        # Bold after appending: openpyxl reports max_row == 1 for an *empty* sheet,
        # so pre-computing max_row + 1 would bold the wrong row for the first
        # strategy on each sheet.
        ws.append([strategy_display_label(strategy_key)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for entry in strategy_entry["rows"]:
            _community_id, _count, label, hex_color = entry["group"]
            hex_color = str(hex_color).lstrip("#")
            metrics = entry["metrics"]
            channels_str = ", ".join(c["label"] for c in entry.get("channels", []))
            ws.append(
                [
                    str(label),
                    f"#{hex_color}",
                    entry["node_count"],
                    metrics["internal_edges"],
                    metrics["external_edges"],
                    metrics.get("ei_index"),
                    metrics["density"],
                    metrics["reciprocity"],
                    metrics["avg_clustering"],
                    metrics["avg_path_length"],
                    metrics["diameter"],
                    channels_str,
                ]
            )
            try:
                fill = PatternFill(start_color=hex_color.upper(), end_color=hex_color.upper(), fill_type="solid")
                ws.cell(row=ws.max_row, column=1).fill = fill
            except (ValueError, TypeError):
                pass

    wb = openpyxl.Workbook()
    wb.properties.creator = "Pulpit"
    if project_title:
        wb.properties.title = project_title
    wb.remove(wb.active)

    if year_data:
        # One sheet per year; each sheet lists all strategies sequentially
        ws_all = wb.create_sheet(title="All")
        for i, sk in enumerate(strategies):
            _fill_strategy(ws_all, sk, community_table_data, first=(i == 0))
        for yr, yr_ctd in year_data:
            ws_yr = wb.create_sheet(title=str(yr))
            for i, sk in enumerate(strategies):
                _fill_strategy(ws_yr, sk, yr_ctd, first=(i == 0))
    else:
        # One sheet per strategy (original format)
        for strategy_key in strategies:
            ws = wb.create_sheet(title=strategy_display_label(strategy_key)[:31])
            _fill_strategy(ws, strategy_key, community_table_data)

    wb.save(output_filename)


def write_community_metrics_json(
    community_table_data: CommunityTableData,
    strategies: list[str],
    graph_dir: str,
) -> None:
    data_dir = os.path.join(graph_dir, "data")
    communities_path = os.path.join(data_dir, "communities.json")
    with open(communities_path) as f:
        communities_file = json.load(f)

    for strategy_key in strategies:
        entry = community_table_data["strategies"].get(strategy_key)
        if not entry:
            continue
        rows_out = []
        for row in entry["rows"]:
            _community_id, _count, label, hex_color = row["group"]
            hex_color = str(hex_color)
            if not hex_color.startswith("#"):
                hex_color = f"#{hex_color}"
            rows_out.append(
                {
                    "label": str(label),
                    "hex_color": hex_color,
                    "node_count": row["node_count"],
                    "metrics": row["metrics"],
                    "channels": row.get("channels", []),
                }
            )
        strategy_entry = communities_file["strategies"].get(strategy_key)
        if strategy_entry is not None:
            strategy_entry["rows"] = rows_out
            mod = entry.get("modularity")
            strategy_entry["modularity"] = round(mod, 6) if mod is not None else None
            icr = entry.get("inter_community_edge_ratio")
            strategy_entry["inter_community_edge_ratio"] = round(icr, 6) if icr is not None else None
            mei = entry.get("mean_ei_index")
            strategy_entry["mean_ei_index"] = round(mei, 6) if mei is not None else None
            cross_tabs = entry.get("cross_tabs")
            if cross_tabs is not None:
                strategy_entry["cross_tabs"] = cross_tabs

    with open(communities_path, "w") as f:
        f.write(json.dumps(communities_file))


def write_community_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/community_table.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Communities",
        seo_title_part="Community statistics",
    )


def write_consensus_matrix_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/consensus_matrix.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Consensus matrix",
    )


def write_structural_similarity_json(sim_data: dict, graph_dir: str) -> None:
    # Filename kept as structural_similarity.json for saved-config / URL compatibility;
    # the page now shows true structural equivalence (Lorrain & White 1971).
    data_dir = os.path.join(graph_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    with open(os.path.join(data_dir, "structural_similarity.json"), "w") as f:
        f.write(json.dumps(sim_data))


def write_structural_similarity_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/structural_similarity.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Structural equivalence",
    )


def write_behavioural_equivalence_json(sim_data: dict, graph_dir: str) -> None:
    data_dir = os.path.join(graph_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    with open(os.path.join(data_dir, "behavioural_equivalence.json"), "w") as f:
        f.write(json.dumps(sim_data))


def write_behavioural_equivalence_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/behavioural_equivalence.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Behavioural equivalence",
    )


def write_vacancy_analysis_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/vacancy_analysis.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Vacancy Analysis",
    )


def write_robustness_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/robustness_table.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Robustness",
        seo_title_part="Network robustness analysis",
    )


def write_interest_structural_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    _write_page(
        "network/interest_structural.html",
        output_filename,
        seo=seo,
        project_title=project_title,
        title_part="Interesting messages",
        seo_title_part="Structurally interesting messages",
    )


_ROBUSTNESS_METRICS: tuple[str, ...] = ("wcc", "scc", "reach")


def _robustness_sheet_name(prefix: str, suffix: str) -> str:
    """Excel sheet names are capped at 31 chars.  With a non-empty *suffix*
    (e.g. ``"All"``, ``"2019"``), the prefix is truncated to make room for
    ``"<prefix> <suffix>"`` while keeping the suffix intact."""
    if not suffix:
        return prefix[:31]
    full = f"{prefix} {suffix}"
    if len(full) <= 31:
        return full
    return f"{prefix[: 30 - len(suffix)]} {suffix}"


def _fill_robustness_summary(wb: Any, payload: dict, suffix: str) -> None:
    ws = wb.create_sheet(title=_robustness_sheet_name("Summary", suffix))
    headers = ["Strategy", "Metric", "R", "R_null_mean", "R_null_std", "z", "f_c"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for s in payload.get("strategies", {}):
        strat = payload["strategies"][s]
        null = strat.get("null") or {}
        for m in _ROBUSTNESS_METRICS:
            null_m = null.get(f"r_{m}") or {}
            ws.append(
                [
                    s,
                    m.upper(),
                    strat.get(f"r_{m}"),
                    null_m.get("mean"),
                    null_m.get("std"),
                    null_m.get("z"),
                    strat.get(f"fc_{m}"),
                ]
            )


def _fill_robustness_curves(wb: Any, payload: dict, suffix: str) -> None:
    for s in payload.get("strategies", {}):
        strat = payload["strategies"][s]
        null = strat.get("null") or {}
        ws = wb.create_sheet(title=_robustness_sheet_name(f"Curve {s}", suffix))
        cols: list[tuple[str, list]] = [
            ("q", list(range(len(strat["curve_wcc"])))),
            ("S_wcc", strat["curve_wcc"]),
            ("S_scc", strat["curve_scc"]),
            ("S_reach", strat["curve_reach"]),
        ]
        if null:
            for m in _ROBUSTNESS_METRICS:
                cols.append((f"null_{m}_mean", null.get(f"curve_{m}_mean", [])))
                cols.append((f"null_{m}_std", null.get(f"curve_{m}_std", [])))
        ws.append([c[0] for c in cols])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for q in range(len(cols[0][1])):
            ws.append([col[1][q] if q < len(col[1]) else None for col in cols])


def _fill_robustness_modular(wb: Any, payload: dict, suffix: str) -> None:
    strategies = list(payload.get("strategies", {}).keys())
    modular = payload.get("modular") or {}
    for partition_name, per_strategy in modular.items():
        ws = wb.create_sheet(title=_robustness_sheet_name(f"Modular {partition_name}", suffix))
        any_strategy = next(iter(per_strategy.values()))
        n_points = len(any_strategy["intra"])
        header = ["q"]
        rows_cols: list[list] = [list(range(n_points))]
        for s in strategies:
            curves = per_strategy.get(s)
            if curves is None:
                continue
            for key in ("intra", "inter", "ratio"):
                header.append(f"{key}_{s}")
                rows_cols.append(curves[key])
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for q in range(n_points):
            ws.append([col[q] if q < len(col) else None for col in rows_cols])


def write_robustness_table_xlsx(
    rob_payload: dict,
    output_filename: str,
    project_title: str = "",
    year_data: "list[tuple[int, dict]] | None" = None,
) -> None:
    """Write the robustness payload(s) as an Excel workbook.

    Without *year_data* sheets are named ``"Summary"``, ``"Curve <strategy>"``,
    and ``"Modular <partition>"``.

    With *year_data* (a list of ``(year, year_payload)`` from the timeline
    loop), one contiguous block of sheets is emitted per scope — ``"All"``
    for the global payload first, then ``"<year>"`` for each per-year
    payload — so sheet names become ``"Summary All"``, ``"Curve pagerank
    All"``, …, ``"Summary 2019"``, ``"Curve pagerank 2019"``, etc.  Sheet
    titles are truncated to fit Excel's 31-char cap while preserving the
    year suffix.
    """
    wb = openpyxl.Workbook()
    wb.properties.creator = "Pulpit"
    if project_title:
        wb.properties.title = project_title
    wb.remove(wb.active)

    if year_data:
        scopes: list[tuple[str, dict]] = [("All", rob_payload)]
        for yr, yr_payload in year_data:
            scopes.append((str(yr), yr_payload))
    else:
        scopes = [("", rob_payload)]

    for suffix, payload in scopes:
        _fill_robustness_summary(wb, payload, suffix)
        _fill_robustness_curves(wb, payload, suffix)
        _fill_robustness_modular(wb, payload, suffix)

    wb.save(output_filename)


def write_index_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
    include_graph: bool = False,
    include_3d_graph: bool = False,
    include_channel_html: bool = False,
    include_channel_xlsx: bool = False,
    include_network_html: bool = False,
    include_network_xlsx: bool = False,
    include_community_html: bool = False,
    include_community_xlsx: bool = False,
    include_consensus_matrix_html: bool = False,
    include_structural_similarity: bool = False,
    include_behavioural_equivalence: bool = False,
    include_compare_html: bool = False,
    compare_files: set[str] | None = None,
    strategies: list[str] | None = None,
    timeline_entries: list[dict] | None = None,
    include_vacancy_analysis: bool = False,
    include_robustness_html: bool = False,
    include_robustness_xlsx: bool = False,
    include_interest_structural: bool = False,
) -> None:
    if seo:
        title = project_title or "Network Analysis"
        robots_meta = "index, follow"
    else:
        title = project_title or "Network Analysis"
        robots_meta = "noindex, nofollow"

    context = {
        "title": title,
        "robots_meta": robots_meta,
        "project_title": project_title,
        "include_graph": include_graph,
        "include_3d_graph": include_3d_graph,
        "include_channel_html": include_channel_html,
        "include_channel_xlsx": include_channel_xlsx,
        "include_network_html": include_network_html,
        "include_network_xlsx": include_network_xlsx,
        "include_community_html": include_community_html,
        "include_community_xlsx": include_community_xlsx,
        "include_consensus_matrix_html": include_consensus_matrix_html,
        "include_structural_similarity": include_structural_similarity,
        "include_behavioural_equivalence": include_behavioural_equivalence,
        "include_compare_html": include_compare_html,
        "compare_files": compare_files or set(),
        "strategies": [strategy_display_label(s) for s in (strategies or [])],
        "timeline_entries": timeline_entries or [],
        "include_vacancy_analysis": include_vacancy_analysis,
        "include_robustness_html": include_robustness_html,
        "include_robustness_xlsx": include_robustness_xlsx,
        "include_interest_structural": include_interest_structural,
        **_pulpit_ctx(),
    }
    content = render_to_string("network/index.html", context)
    with open(output_filename, "w") as f:
        f.write(content)

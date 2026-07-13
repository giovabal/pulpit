import datetime
import io
import json
import os
import tempfile
from typing import Any
from unittest.mock import MagicMock, patch

from django.test import TestCase, override_settings

from network.community import (
    COMMUNITY_ALGORITHMS,
    VALID_STRATEGIES,
    apply_edge_colors,
    apply_to_graph,
    build_communities_payload,
    build_community_label,
    build_community_palette,
    canonical_strategy_key,
    detect_kcore,
    detect_labelgroup,
    detect_leiden,
    detect_louvain,
    normalize_community_map,
    parse_strategies,
    strategy_display_label,
)
from network.community_stats import (
    PARTITION_COMPARISON_METRICS,
    _compare_partitions,
    _freeman_centralization,
    _network_summary,
    compute_community_metrics,
    network_summary_rows,
)
from network.coordination import build_nx_graph, compute_coordination
from network.exporter import (
    build_coordination_graph_data,
    build_graph_data,
    ensure_graph_root,
    find_main_component,
    reposition_isolated_nodes,
    write_coordination_files,
    write_coordination_pages,
    write_coordination_timeline_json,
    write_graph_files,
)
from network.graph_builder import build_graph, resolve_window_label
from network.measures import (
    apply_amplification_factor,
    apply_base_node_measures,
    apply_burt_constraint,
    apply_content_originality,
    apply_hits,
    apply_in_degree_centrality,
    apply_local_clustering,
    apply_module_role,
    apply_out_degree_centrality,
    apply_pagerank,
    apply_reciprocity,
    canonical_measure_key,
    parse_measures,
    role_companions,
)
from network.tokens import split_tokens
from network.utils import channel_cutoff_q
from webapp.models import Channel, Message
from webapp.test_helpers import attribute, label_group, make_channel, make_label
from webapp.utils.colors import parse_color

import networkx as nx
import numpy as np

# ---------------------------------------------------------------------------
# community.py — normalize_community_map
# ---------------------------------------------------------------------------


class NormalizeCommunityMapTests(TestCase):
    def test_empty_map_returns_empty(self) -> None:
        self.assertEqual(normalize_community_map({}), {})

    def test_single_node_maps_to_community_1(self) -> None:
        result = normalize_community_map({"a": 5})
        self.assertEqual(result, {"a": 1})

    def test_largest_community_gets_id_1(self) -> None:
        # community 99 has 2 nodes, community 1 has 1 node
        result = normalize_community_map({"x": 99, "y": 99, "z": 1})
        self.assertEqual(result["x"], 1)
        self.assertEqual(result["y"], 1)
        self.assertEqual(result["z"], 2)

    def test_tie_broken_by_original_community_id_ascending(self) -> None:
        # Two communities of equal size (1 each); lower original id gets remapped first
        result = normalize_community_map({"a": 10, "b": 20})
        self.assertEqual(result["a"], 1)
        self.assertEqual(result["b"], 2)

    def test_all_original_nodes_preserved(self) -> None:
        original = {"n1": 3, "n2": 3, "n3": 7, "n4": 7, "n5": 9}
        result = normalize_community_map(original)
        self.assertEqual(set(result.keys()), set(original.keys()))

    def test_community_ids_start_at_1(self) -> None:
        result = normalize_community_map({"a": 10, "b": 20, "c": 30})
        self.assertGreaterEqual(min(result.values()), 1)


# ---------------------------------------------------------------------------
# community.py — build_community_label
# ---------------------------------------------------------------------------


class BuildCommunityLabelTests(TestCase):
    def test_integer_id_with_algorithm_strategy(self) -> None:
        self.assertEqual(build_community_label(1, "LEIDEN"), "1-leiden")

    def test_string_id_with_metadata_strategy(self) -> None:
        self.assertEqual(build_community_label("my-org", "LABELGROUP1"), "my-org-labelgroup1")

    def test_spaces_become_hyphens_via_slugify(self) -> None:
        self.assertEqual(build_community_label("Foo Bar", "TEST"), "foo-bar-test")

    def test_uppercase_strategy_lowercased(self) -> None:
        label = build_community_label(5, "KCORE")
        self.assertIn("kcore", label)


# ---------------------------------------------------------------------------
# community.py — build_community_palette
# ---------------------------------------------------------------------------


class BuildCommunityPaletteTests(TestCase):
    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_empty_community_map_returns_empty(self, _mock: MagicMock) -> None:
        result = build_community_palette({}, "SomePalette")
        self.assertEqual(result, {})

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00"])
    def test_returns_one_entry_per_community(self, _mock: MagicMock) -> None:
        community_map = {"a": 1, "b": 2}
        result = build_community_palette(community_map, "SomePalette")
        self.assertIn(1, result)
        self.assertIn(2, result)
        self.assertEqual(len(result), 2)

    @patch("network.community.palette_colors", return_value=["#ff0000"])
    def test_palette_values_are_three_int_tuples(self, _mock: MagicMock) -> None:
        community_map = {"a": 1}
        result = build_community_palette(community_map, "SomePalette")
        color = result[1]
        self.assertIsInstance(color, tuple)
        self.assertEqual(len(color), 3)
        self.assertTrue(all(isinstance(c, int) for c in color))

    @patch("network.community.palette_colors", return_value=["#ff0000"])
    def test_expands_colors_when_fewer_palette_entries_than_communities(self, _mock: MagicMock) -> None:
        community_map = {"a": 1, "b": 2, "c": 3}
        result = build_community_palette(community_map, "SomePalette")
        self.assertEqual(set(result.keys()), {1, 2, 3})

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_canonical_order_when_reverse_false(self, _mock: MagicMock) -> None:
        community_map = {"a": 1, "b": 2, "c": 3}
        result = build_community_palette(community_map, "vaporwave")
        self.assertEqual(result[1], parse_color("#ff0000"))
        self.assertEqual(result[2], parse_color("#00ff00"))
        self.assertEqual(result[3], parse_color("#0000ff"))

    def test_reverse_kwarg_forwards_to_palette_colors(self) -> None:
        # palette_colors honours its own ``reverse`` kwarg (it forwards it to
        # pypalettes.load_palette); build_community_palette only has to pass it through.
        community_map = {"a": 1, "b": 2, "c": 3}
        with patch("network.community.palette_colors") as mock:
            mock.return_value = ["#0000ff", "#00ff00", "#ff0000"]
            result = build_community_palette(community_map, "vaporwave", reverse=True)
        mock.assert_called_once_with("vaporwave", reverse=True)
        self.assertEqual(result[1], parse_color("#0000ff"))
        self.assertEqual(result[3], parse_color("#ff0000"))


# ---------------------------------------------------------------------------
# community.py — detect_labelgroup
# ---------------------------------------------------------------------------


class DetectLabelGroupTests(TestCase):
    def setUp(self) -> None:
        self.group = label_group()
        self.label = make_label("Test Org", color="#FF0000")
        self.ch1 = make_channel(telegram_id=1, label=self.label)
        self.ch2 = make_channel(telegram_id=2, label=self.label)

    def _node(self, channel, label):
        # The window resolution graph_builder writes: group_partitions[gid] = (label_id, label_color).
        return {"channel": channel, "data": {"group_partitions": {self.group.pk: (label.pk, label.color)}}}

    def test_maps_channels_to_label_id(self) -> None:
        channel_dict = {
            str(self.ch1.pk): self._node(self.ch1, self.label),
            str(self.ch2.pk): self._node(self.ch2, self.label),
        }
        community_map, _ = detect_labelgroup(self.group.pk, channel_dict)
        self.assertEqual(community_map[str(self.ch1.pk)], self.label.pk)
        self.assertEqual(community_map[str(self.ch2.pk)], self.label.pk)

    def test_channel_without_label_excluded_from_map(self) -> None:
        ch3 = make_channel(telegram_id=3, label=None)
        channel_dict = {str(ch3.pk): {"channel": ch3, "data": {"group_partitions": {}}}}
        community_map, _ = detect_labelgroup(self.group.pk, channel_dict)
        self.assertNotIn(str(ch3.pk), community_map)

    def test_palette_uses_label_color(self) -> None:
        channel_dict = {str(self.ch1.pk): self._node(self.ch1, self.label)}
        _, palette = detect_labelgroup(self.group.pk, channel_dict)
        expected = parse_color(self.label.color)
        self.assertEqual(palette[self.label.pk], expected)

    def test_palette_has_one_entry_per_unique_label(self) -> None:
        label2 = make_label("Org2", color="#0000FF")
        ch4 = make_channel(telegram_id=4, label=label2)
        channel_dict = {
            str(self.ch1.pk): self._node(self.ch1, self.label),
            str(ch4.pk): self._node(ch4, label2),
        }
        _, palette = detect_labelgroup(self.group.pk, channel_dict)
        self.assertEqual(len(palette), 2)


# ---------------------------------------------------------------------------
# community.py — detect_consensus
# ---------------------------------------------------------------------------


class DetectConsensusTests(TestCase):
    """Lancichinetti & Fortunato 2012 consensus over the other strategies' partitions."""

    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["a", "b", "c", "d"])
        self.graph.add_edges_from([("a", "b"), ("c", "d"), ("b", "c")])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_majority_agreement_groups_survive(self, _mock: MagicMock) -> None:
        # Two of three partitions agree on {a,b} + {c,d}; the third scatters everything.
        # At τ=0.5 the 2/3 co-assignments survive and the consensus recovers the agreed blocks.
        from network.community import detect_consensus

        inputs = {
            "leiden": {"a": 1, "b": 1, "c": 2, "d": 2},
            "louvain": {"a": 7, "b": 7, "c": 9, "d": 9},
            "sbm_mode_nested": {"a": 1, "b": 2, "c": 3, "d": 4},
        }
        community_map, palette = detect_consensus(self.graph, "SomePalette", inputs, 0.5)
        self.assertEqual(community_map["a"], community_map["b"])
        self.assertEqual(community_map["c"], community_map["d"])
        self.assertNotEqual(community_map["a"], community_map["c"])
        for community_id in community_map.values():
            self.assertIn(community_id, palette)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_threshold_prunes_minority_pairs(self, _mock: MagicMock) -> None:
        # a-b co-assigned by 1 of 3 partitions (agreement 1/3): kept at τ=0.3, dropped at τ=0.5.
        from network.community import detect_consensus

        inputs = {
            "p1": {"a": 1, "b": 1, "c": 2, "d": 3},
            "p2": {"a": 1, "b": 2, "c": 3, "d": 4},
            "p3": {"a": 1, "b": 2, "c": 3, "d": 4},
        }
        loose, _ = detect_consensus(self.graph, "SomePalette", inputs, 0.3)
        self.assertEqual(loose["a"], loose["b"])
        strict, _ = detect_consensus(self.graph, "SomePalette", inputs, 0.5)
        self.assertNotEqual(strict["a"], strict["b"])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_unanimous_inputs_recover_the_partition(self, _mock: MagicMock) -> None:
        from network.community import detect_consensus

        shared = {"a": 1, "b": 1, "c": 2, "d": 2}
        community_map, _ = detect_consensus(self.graph, "SomePalette", {"x": shared, "y": dict(shared)}, 0.5)
        self.assertEqual(community_map["a"], community_map["b"])
        self.assertEqual(community_map["c"], community_map["d"])
        self.assertNotEqual(community_map["a"], community_map["c"])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_all_nodes_assigned_even_without_consensus(self, _mock: MagicMock) -> None:
        # Fully disagreeing inputs → no pair survives; every node still gets a community id.
        from network.community import detect_consensus

        inputs = {
            "p1": {"a": 1, "b": 2, "c": 3, "d": 4},
            "p2": {"a": 4, "b": 3, "c": 2, "d": 1},
        }
        community_map, _ = detect_consensus(self.graph, "SomePalette", inputs, 0.9)
        self.assertEqual(set(community_map.keys()), set(self.graph.nodes()))

    def test_requires_two_input_partitions(self) -> None:
        from network.community import detect_consensus

        with self.assertRaisesRegex(ValueError, "at least two"):
            detect_consensus(self.graph, "SomePalette", {"only": {"a": 1}}, 0.5)

    def test_detect_rejects_consensus_token(self) -> None:
        # CONSENSUS is dispatched by the pipeline (detect_consensus), never by detect().
        from network.community import detect

        with self.assertRaisesRegex(ValueError, "dispatched separately"):
            detect(parse_strategies(["CONSENSUS"])[0], "vaporwave", self.graph, {})

    def test_consensus_eligibility_rule(self) -> None:
        from network.community import consensus_eligible

        self.assertTrue(consensus_eligible("LEIDEN"))
        self.assertTrue(consensus_eligible("SBM"))
        self.assertFalse(consensus_eligible("KCORE"))
        self.assertFalse(consensus_eligible("CONSENSUS"))
        self.assertFalse(consensus_eligible("LABELGROUP3"))


# ---------------------------------------------------------------------------
# community.py — detect_leiden_temporal
# ---------------------------------------------------------------------------


class DetectLeidenTemporalTests(TestCase):
    """Interslice-coupled temporal detection over synthetic year slices (real leidenalg)."""

    def _year_graph(self, edges: list[tuple[str, str]], isolated: tuple[str, ...] = ()) -> "Any":
        graph = nx.DiGraph()
        for s, t in edges:
            graph.add_edge(s, t, weight=1.0)
        graph.add_nodes_from(isolated)
        return graph

    def _two_blocks(self) -> list[tuple[str, str]]:
        return [("a", "b"), ("b", "c"), ("c", "a"), ("d", "e"), ("e", "f"), ("f", "d"), ("c", "d")]

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_stable_ids_across_years(self, _mock: MagicMock) -> None:
        from network.community import detect_leiden_temporal

        year_graphs = {2020: self._year_graph(self._two_blocks()), 2021: self._year_graph(self._two_blocks())}
        per_year, plurality, palette = detect_leiden_temporal(year_graphs, "SomePalette", 0.3, 1.0)
        self.assertEqual(set(per_year), {2020, 2021})
        for node in "abcdef":
            # Same structure in both years + identity coupling → same community id per node.
            self.assertEqual(per_year[2020][node], per_year[2021][node])
            self.assertEqual(plurality[node], per_year[2020][node])
        self.assertEqual(per_year[2020]["a"], per_year[2020]["b"])
        self.assertEqual(per_year[2020]["d"], per_year[2020]["e"])
        self.assertNotEqual(per_year[2020]["a"], per_year[2020]["d"])
        for cid in plurality.values():
            self.assertIn(cid, palette)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_coupling_carries_an_isolated_year_through(self, _mock: MagicMock) -> None:
        # 'a' has no edges in 2021 but exists; with ω=1 the identity link keeps it in its
        # 2020 community — the persistence the method is for. With ω=0 the years decouple
        # and isolated 'a' falls out into its own community.
        from network.community import detect_leiden_temporal

        year_graphs = {
            2020: self._year_graph(self._two_blocks()),
            2021: self._year_graph(
                [("b", "c"), ("c", "b"), ("d", "e"), ("e", "f"), ("f", "d"), ("c", "d")], isolated=("a",)
            ),
        }
        per_year, _, _ = detect_leiden_temporal(year_graphs, "SomePalette", 0.3, 1.0)
        self.assertEqual(per_year[2021]["a"], per_year[2021]["b"])
        per_year_uncoupled, _, _ = detect_leiden_temporal(year_graphs, "SomePalette", 0.3, 0.0)
        self.assertNotEqual(per_year_uncoupled[2021]["a"], per_year_uncoupled[2021]["b"])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_plurality_tie_breaks_to_latest_year(self, _mock: MagicMock) -> None:
        # 'x' spends one year with the a-block and one with the d-block (1–1 tie): the
        # full-range summary takes the latest year's assignment.
        from network.community import detect_leiden_temporal

        base = self._two_blocks()
        year_graphs = {
            2020: self._year_graph(base + [("x", "a"), ("a", "x"), ("x", "b"), ("b", "x")]),
            2021: self._year_graph(base + [("x", "d"), ("d", "x"), ("x", "e"), ("e", "x")]),
        }
        per_year, plurality, _ = detect_leiden_temporal(year_graphs, "SomePalette", 0.3, 0.0)
        if per_year[2020]["x"] != per_year[2021]["x"]:  # decoupled years → genuinely different homes
            self.assertEqual(plurality["x"], per_year[2021]["x"])

    def test_requires_two_year_slices(self) -> None:
        from network.community import detect_leiden_temporal

        with self.assertRaisesRegex(ValueError, "at least two"):
            detect_leiden_temporal({2020: self._year_graph(self._two_blocks())}, "SomePalette", 0.3, 1.0)

    def test_detect_rejects_temporal_token(self) -> None:
        from network.community import detect

        with self.assertRaisesRegex(ValueError, "dispatched separately"):
            detect(parse_strategies(["LEIDEN_TEMPORAL"])[0], "vaporwave", nx.DiGraph(), {})


# ---------------------------------------------------------------------------
# community.py — detect_kcore
# ---------------------------------------------------------------------------


class DetectKcoreTests(TestCase):
    def setUp(self) -> None:
        # Triangle a-b-c has coreness 2; d is a leaf with coreness 1
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["a", "b", "c", "d"])
        self.graph.add_edges_from([("a", "b"), ("b", "c"), ("c", "a"), ("a", "d")])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00"])
    def test_all_nodes_assigned(self, _mock: MagicMock) -> None:
        community_map, _ = detect_kcore(self.graph, "SomePalette")
        self.assertEqual(set(community_map.keys()), set(self.graph.nodes()))

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00"])
    def test_shells_produce_distinct_communities(self, _mock: MagicMock) -> None:
        community_map, _ = detect_kcore(self.graph, "SomePalette")
        # a, b, c have coreness 2; d has coreness 1 → 2 distinct community values
        self.assertGreaterEqual(len(set(community_map.values())), 2)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00"])
    def test_community_ids_start_at_1(self, _mock: MagicMock) -> None:
        community_map, _ = detect_kcore(self.graph, "SomePalette")
        self.assertGreaterEqual(min(community_map.values()), 1)


# ---------------------------------------------------------------------------
# community.py — apply_to_graph
# ---------------------------------------------------------------------------


class ApplyToGraphTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_node("1", data={"color": ""})
        self.graph.add_node("2", data={"color": ""})
        self.graph.add_edge("1", "2")
        self.community_map = {"1": 1, "2": 2}
        self.community_palette = {1: (255, 0, 0), 2: (0, 255, 0)}
        self.channel_dict: dict = {
            "1": {"channel": None, "data": {}},
            "2": {"channel": None, "data": {}},
        }

    def test_sets_communities_on_graph_nodes(self) -> None:
        apply_to_graph(self.graph, self.channel_dict, self.community_map, self.community_palette, "LEIDEN")
        for node_id in ["1", "2"]:
            self.assertIn("communities", self.graph.nodes[node_id]["data"])

    def test_sets_color_on_graph_nodes(self) -> None:
        apply_to_graph(self.graph, self.channel_dict, self.community_map, self.community_palette, "LEIDEN")
        for node_id in ["1", "2"]:
            color = self.graph.nodes[node_id]["data"]["color"]
            self.assertIsInstance(color, str)
            self.assertEqual(len(color.split(",")), 3)  # "r,g,b" format

    def test_updates_channel_dict_with_communities_and_color(self) -> None:
        apply_to_graph(self.graph, self.channel_dict, self.community_map, self.community_palette, "LEIDEN")
        for key in ["1", "2"]:
            self.assertIn("communities", self.channel_dict[key]["data"])
            self.assertIn("color", self.channel_dict[key]["data"])

    def test_algorithm_community_label_includes_community_id_and_strategy(self) -> None:
        apply_to_graph(self.graph, self.channel_dict, self.community_map, self.community_palette, "LEIDEN")
        self.assertEqual(self.graph.nodes["1"]["data"]["communities"]["leiden"], "1-leiden")
        self.assertEqual(self.graph.nodes["2"]["data"]["communities"]["leiden"], "2-leiden")

    def test_node_without_community_gets_fallback_color(self) -> None:
        # community_map is empty → no community assigned → nodes use DEFAULT_FALLBACK_COLOR
        apply_to_graph(self.graph, self.channel_dict, {}, self.community_palette, "LEIDEN")
        for node_id in ["1", "2"]:
            color = self.graph.nodes[node_id]["data"]["color"]
            self.assertIsInstance(color, str)
            self.assertEqual(color, "204,204,204")  # DEFAULT_FALLBACK_COLOR

    def test_algorithm_strategy_uses_integer_label(self) -> None:
        for strategy in COMMUNITY_ALGORITHMS:
            apply_to_graph(self.graph, self.channel_dict, self.community_map, self.community_palette, strategy)
            community_label = self.graph.nodes["1"]["data"]["communities"][strategy.lower()]
            # Label must contain the integer community id
            self.assertIn("1", community_label)


# ---------------------------------------------------------------------------
# community.py — apply_edge_colors
# ---------------------------------------------------------------------------


class ApplyEdgeColorsTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_node("1", data={"color": "255,0,0"})
        self.graph.add_node("2", data={"color": "0,0,255"})
        self.graph.add_edge("1", "2", weight=1.0)
        self.channel_dict = {
            "1": {"data": {"color": "255,0,0"}},
            "2": {"data": {"color": "0,0,255"}},
        }
        self.edge_list = [["1", "2", 1.0]]

    def test_assigns_color_to_edges(self) -> None:
        apply_edge_colors(self.graph, self.edge_list, self.channel_dict)
        self.assertIn("color", self.graph.edges["1", "2"])

    def test_color_is_comma_separated_rgb_string(self) -> None:
        apply_edge_colors(self.graph, self.edge_list, self.channel_dict)
        color = self.graph.edges["1", "2"]["color"]
        parts = color.split(",")
        self.assertEqual(len(parts), 3)
        self.assertTrue(all(p.isdigit() for p in parts))

    def test_edge_color_is_darkened_average(self) -> None:
        # red (255,0,0) + blue (0,0,255) → avg (127,0,127) → * 0.75 → (95,0,95)
        apply_edge_colors(self.graph, self.edge_list, self.channel_dict)
        color = self.graph.edges["1", "2"]["color"]
        r, g, b = (int(p) for p in color.split(","))
        self.assertEqual(g, 0)  # green channel stays 0
        self.assertEqual(r, b)  # red and blue channels are symmetric


# ---------------------------------------------------------------------------
# community.py — build_communities_payload
# ---------------------------------------------------------------------------


class BuildCommunitiesPayloadTests(TestCase):
    def test_algorithm_strategy_builds_groups_from_community_map(self) -> None:
        community_map = {"a": 1, "b": 1, "c": 2}
        community_palette = {1: (255, 0, 0), 2: (0, 255, 0)}
        result = build_communities_payload(parse_strategies(["LEIDEN"]), {"leiden": (community_map, community_palette)})
        self.assertIn("leiden", result)
        self.assertIn("groups", result["leiden"])
        self.assertIn("main_groups", result["leiden"])
        self.assertEqual(len(result["leiden"]["groups"]), 2)

    def test_groups_sorted_by_count_descending(self) -> None:
        community_map = {"a": 1, "b": 1, "c": 1, "d": 2}  # community 1: 3 nodes, 2: 1 node
        community_palette = {1: (255, 0, 0), 2: (0, 255, 0)}
        result = build_communities_payload(parse_strategies(["LEIDEN"]), {"leiden": (community_map, community_palette)})
        counts = [g[1] for g in result["leiden"]["groups"]]
        self.assertEqual(counts, sorted(counts, reverse=True))

    def test_algorithm_main_groups_maps_id_to_label(self) -> None:
        community_map = {"a": 1}
        community_palette = {1: (255, 0, 0)}
        result = build_communities_payload(parse_strategies(["KCORE"]), {"kcore": (community_map, community_palette)})
        self.assertIn("1", result["kcore"]["main_groups"])

    def test_labelgroup_strategy_uses_resolved_map(self) -> None:
        group = label_group()
        label = make_label("My Org")
        token = group.token  # "LABELGROUP<pk>"
        (inst,) = parse_strategies([token])
        cmap = {"n1": label.pk, "n2": label.pk}
        result = build_communities_payload([inst], {inst.key: (cmap, {label.pk: (1, 2, 3)})})
        groups = result[inst.key]["groups"]
        self.assertIn("My Org", [g[2] for g in groups])
        # Count comes from the resolved community map (2 nodes), not a raw FK count.
        self.assertEqual([g[1] for g in groups if g[2] == "My Org"][0], 2)

    def test_labelgroup_with_no_resolved_nodes_excluded(self) -> None:
        # A label that owns no node in the window does not appear in the LABELGROUP groups.
        group = label_group()
        make_label("Absent Org")
        (inst,) = parse_strategies([group.token])
        result = build_communities_payload([inst], {inst.key: ({}, {})})
        self.assertNotIn("Absent Org", [g[2] for g in result[inst.key]["groups"]])

    def test_labelgroup_strategy_main_groups_uses_key_and_name(self) -> None:
        group = label_group()
        label = make_label("My Org")
        (inst,) = parse_strategies([group.token])
        result = build_communities_payload([inst], {inst.key: ({"n1": label.pk}, {label.pk: (1, 2, 3)})})
        self.assertEqual(result[inst.key]["main_groups"].get(label.key), label.name)

    def test_multiple_strategies_all_included(self) -> None:
        community_map = {"a": 1}
        community_palette = {1: (255, 0, 0)}
        results = {
            "leiden": (community_map, community_palette),
            "kcore": (community_map, community_palette),
        }
        result = build_communities_payload(parse_strategies(["LEIDEN", "KCORE"]), results)
        self.assertIn("leiden", result)
        self.assertIn("kcore", result)


# ---------------------------------------------------------------------------
# community.py — parse_strategies / StrategyInstance / canonical / labels
# ---------------------------------------------------------------------------


class SplitTokensTests(TestCase):
    def test_plain_tokens(self) -> None:
        self.assertEqual(split_tokens("LEIDEN,KCORE"), ["LEIDEN", "KCORE"])

    def test_comma_inside_params_does_not_split(self) -> None:
        self.assertEqual(
            split_tokens("LEIDEN,LEIDEN_TEMPORAL(resolution=0.05,interslice=1.0),KCORE"),
            ["LEIDEN", "LEIDEN_TEMPORAL(resolution=0.05,interslice=1.0)", "KCORE"],
        )

    def test_multi_param_token_round_trips_through_parse(self) -> None:
        tokens = [t.upper() for t in split_tokens("LEIDEN_TEMPORAL(RESOLUTION=0.05,INTERSLICE=1.0)")]
        (inst,) = parse_strategies(tokens)
        self.assertEqual(inst.params_dict, {"resolution": 0.05, "interslice": 1.0})

    def test_whitespace_and_empty_pieces(self) -> None:
        self.assertEqual(
            split_tokens(" LEIDEN , ,SBM(mode=NESTED, refine=MCMC) ,"), ["LEIDEN", "SBM(mode=NESTED, refine=MCMC)"]
        )

    def test_empty_string(self) -> None:
        self.assertEqual(split_tokens(""), [])

    def test_unbalanced_paren_still_returns_pieces(self) -> None:
        # A malformed token isn't the splitter's problem — parse_tokens rejects it downstream.
        self.assertEqual(split_tokens("LEIDEN_TEMPORAL(resolution=0.05"), ["LEIDEN_TEMPORAL(resolution=0.05"])


class StrategyParserTests(TestCase):
    def test_plain_and_parameterised_keys(self) -> None:
        insts = parse_strategies(["LEIDEN_DIRECTED", "LEIDEN_CPM(resolution=0.01)"])
        self.assertEqual([i.name for i in insts], ["LEIDEN_DIRECTED", "LEIDEN_CPM"])
        self.assertEqual([i.key for i in insts], ["leiden_directed", "leiden_cpm_resolution_0_01"])
        self.assertEqual(insts[1].token(), "LEIDEN_CPM(resolution=0.01)")

    def test_same_strategy_twice_with_different_params(self) -> None:
        insts = parse_strategies(["LEIDEN_CPM(resolution=0.01)", "LEIDEN_CPM(resolution=0.05)"])
        self.assertEqual([i.key for i in insts], ["leiden_cpm_resolution_0_01", "leiden_cpm_resolution_0_05"])

    def test_rejects_identical_parameterised(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["LEIDEN_CPM(resolution=0.05)", "LEIDEN_CPM(resolution=0.05)"])

    def test_rejects_duplicate_plain(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["LEIDEN", "LEIDEN"])

    def test_rejects_params_on_plain(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["LEIDEN(resolution=0.1)"])

    def test_rejects_unknown(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["NOTASTRATEGY"])

    def test_all_expands_to_every_strategy(self) -> None:
        # ALL expands to every algorithm strategy plus one LABELGROUP<id> per partition group
        # (the seed migration creates the primary "Organization" partition group) — minus
        # EXCLUDED_FROM_ALL (LEIDEN_TEMPORAL needs a year timeline, so ALL must not imply it).
        # The metadata LABELGROUP tokens are DB-derived, so build the expected set from the
        # current partitions.
        from network.community import EXCLUDED_FROM_ALL
        from webapp.models import LabelGroup

        names = {i.name for i in parse_strategies(["ALL"])}
        expected = (set(VALID_STRATEGIES) - EXCLUDED_FROM_ALL) | {
            f"LABELGROUP{pk}" for pk in LabelGroup.objects.filter(is_partition=True).values_list("pk", flat=True)
        }
        self.assertEqual(names, expected)
        self.assertNotIn("LEIDEN_TEMPORAL", names)

    def test_bare_token_inherits_default(self) -> None:
        (inst,) = parse_strategies(["LEIDEN_CPM"], defaults={"LEIDEN_CPM": {"resolution": 0.02}})
        self.assertEqual(inst.params_dict["resolution"], 0.02)
        self.assertEqual(inst.key, "leiden_cpm_resolution_0_02")

    def test_canonical_strategy_key(self) -> None:
        self.assertEqual(canonical_strategy_key("leiden_cpm_resolution_0_05"), "leiden_cpm")
        self.assertEqual(canonical_strategy_key("leiden_cpm_resolution_0_01"), "leiden_cpm")
        self.assertEqual(canonical_strategy_key("leiden_directed"), "leiden_directed")  # plain unchanged

    def test_strategy_display_label(self) -> None:
        self.assertEqual(strategy_display_label("leiden_cpm_resolution_0_05"), "Leiden CPM (resolution=0.05)")
        self.assertEqual(strategy_display_label("leiden_cpm_resolution_0_01"), "Leiden CPM (resolution=0.01)")
        self.assertEqual(strategy_display_label("leiden_directed"), "Leiden directed")

    def test_label_group_display_carries_custom_label_tag(self) -> None:
        # Outside its own picker a label group reads as a manual partition: both the key-based display
        # label (exports/tables) and the instance label (basis dropdown, summaries) get "[custom label]".
        group = label_group(name="Region", is_primary=False)
        self.assertEqual(strategy_display_label(f"labelgroup{group.pk}"), "Region [custom label]")
        inst = parse_strategies([f"LABELGROUP{group.pk}"])[0]
        self.assertEqual(inst.label, "Region [custom label]")

    def test_strategy_name_sources_stay_in_sync(self) -> None:
        # Guard: measures.ALL_STRATEGIES (the measure "basis" choices) must match the community list.
        from network import measures

        self.assertEqual(set(measures.ALL_STRATEGIES), set(VALID_STRATEGIES))

    def test_sbm_modes_and_keys(self) -> None:
        # Bare SBM inherits the NESTED default; both modes round-trip to distinct suffixed keys.
        insts = parse_strategies(["SBM", "SBM(mode=FLAT)"])
        self.assertEqual([i.key for i in insts], ["sbm_mode_nested", "sbm_mode_flat"])
        self.assertEqual([i.token() for i in insts], ["SBM(mode=NESTED)", "SBM(mode=FLAT)"])

    def test_sbm_rejects_bad_mode(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["SBM(mode=GIANT)"])

    def test_sbm_canonical_and_label(self) -> None:
        self.assertEqual(canonical_strategy_key("sbm_mode_nested"), "sbm")
        self.assertEqual(strategy_display_label("sbm_mode_flat"), "Stochastic block model (mode=flat)")

    def test_sbm_weights_and_refine_params(self) -> None:
        # Bare SBM keeps its historical key (config round-trip stability); the new empty-default
        # params contribute to the key/token only when set, and combinations are distinct instances.
        insts = parse_strategies(["SBM", "SBM(mode=NESTED,weights=POISSON)", "SBM(weights=POISSON,refine=MCMC)"])
        self.assertEqual(
            [i.key for i in insts],
            ["sbm_mode_nested", "sbm_mode_nested_weights_poisson", "sbm_mode_nested_weights_poisson_refine_mcmc"],
        )
        self.assertEqual(insts[1].token(), "SBM(mode=NESTED,weights=POISSON)")

    def test_sbm_rejects_bad_weights_and_refine(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["SBM(weights=GAMMA)"])
        with self.assertRaises(ValueError):
            parse_strategies(["SBM(refine=EXACT)"])

    def test_sbm_multi_param_display_label(self) -> None:
        # Multi-parameter suffixes parse back into per-parameter annotations (the value runs to the
        # next declared parameter's boundary) — the JS mirror is strategy_label in labels.js.
        self.assertEqual(
            strategy_display_label("sbm_mode_nested_weights_poisson_refine_mcmc"),
            "Stochastic block model (mode=nested, weights=poisson, refine=mcmc)",
        )
        self.assertEqual(
            strategy_display_label("sbm_mode_flat_refine_mcmc"),
            "Stochastic block model (mode=flat, refine=mcmc)",
        )

    def test_sbm_confidence_key_and_label(self) -> None:
        from network.community import sbm_confidence_display_label, sbm_confidence_key

        self.assertEqual(sbm_confidence_key("sbm_mode_nested_refine_mcmc"), "sbm_confidence_mode_nested_refine_mcmc")
        self.assertEqual(
            sbm_confidence_display_label("sbm_mode_nested_refine_mcmc"),
            "SBM confidence (mode=nested, refine=mcmc)",
        )

    def test_consensus_token_key_and_label(self) -> None:
        # Bare CONSENSUS inherits τ=0.5; per-threshold instances get distinct suffixed keys.
        insts = parse_strategies(["CONSENSUS", "CONSENSUS(threshold=0.75)"])
        self.assertEqual([i.key for i in insts], ["consensus_threshold_0_5", "consensus_threshold_0_75"])
        self.assertEqual(insts[0].token(), "CONSENSUS(threshold=0.5)")
        self.assertEqual(canonical_strategy_key("consensus_threshold_0_5"), "consensus")
        self.assertEqual(strategy_display_label("consensus_threshold_0_5"), "Consensus (threshold=0.5)")

    def test_consensus_rejects_out_of_range_threshold(self) -> None:
        with self.assertRaises(ValueError):
            parse_strategies(["CONSENSUS(threshold=1.5)"])
        with self.assertRaises(ValueError):
            parse_strategies(["CONSENSUS(threshold=-0.1)"])

    def test_labelpropagation_removed(self) -> None:
        self.assertNotIn("LABELPROPAGATION", VALID_STRATEGIES)
        with self.assertRaises(ValueError):
            parse_strategies(["LABELPROPAGATION"])

    def test_leiden_temporal_token_key_and_label(self) -> None:
        # Bare token inherits γ=0.05 and ω=1.0; both parameters ride in the key and the label.
        insts = parse_strategies(["LEIDEN_TEMPORAL", "LEIDEN_TEMPORAL(resolution=0.01,interslice=0.5)"])
        self.assertEqual(
            [i.key for i in insts],
            ["leiden_temporal_resolution_0_05_interslice_1_0", "leiden_temporal_resolution_0_01_interslice_0_5"],
        )
        self.assertEqual(canonical_strategy_key(insts[0].key), "leiden_temporal")
        self.assertEqual(
            strategy_display_label("leiden_temporal_resolution_0_05_interslice_1_0"),
            "Leiden temporal (resolution=0.05, interslice=1.0)",
        )

    def test_sbm_assortative_token_key_and_label(self) -> None:
        # Bare token has no visible params (refine defaults empty) → bare stable key; MCMC suffixes.
        insts = parse_strategies(["SBM_ASSORTATIVE", "SBM_ASSORTATIVE(refine=MCMC)"])
        self.assertEqual([i.key for i in insts], ["sbm_assortative", "sbm_assortative_refine_mcmc"])
        self.assertEqual(canonical_strategy_key("sbm_assortative_refine_mcmc"), "sbm_assortative")
        self.assertEqual(strategy_display_label("sbm_assortative"), "Assortative SBM")
        self.assertEqual(strategy_display_label("sbm_assortative_refine_mcmc"), "Assortative SBM (refine=mcmc)")
        from network.community import sbm_confidence_display_label, sbm_confidence_key

        self.assertEqual(sbm_confidence_key("sbm_assortative_refine_mcmc"), "sbm_confidence_assortative_refine_mcmc")
        self.assertEqual(
            sbm_confidence_display_label("sbm_assortative_refine_mcmc"),
            "Assortative SBM confidence (refine=mcmc)",
        )

    def test_temporal_not_consensus_eligible(self) -> None:
        from network.community import consensus_eligible

        self.assertFalse(consensus_eligible("LEIDEN_TEMPORAL"))
        self.assertTrue(consensus_eligible("SBM_ASSORTATIVE"))

    def test_sbm_without_graph_tool_raises_value_error(self) -> None:
        # When graph-tool is absent, detect() surfaces a clear ValueError (caught as CommandError upstream).
        import importlib.util

        if importlib.util.find_spec("graph_tool") is not None:
            self.skipTest("graph-tool installed — error path not exercised")
        from network.community import detect

        import networkx as nx

        graph = nx.DiGraph()
        graph.add_edge("a", "b")
        graph.add_edge("c", "b")
        with self.assertRaisesRegex(ValueError, "graph-tool"):
            detect(parse_strategies(["SBM"])[0], "vaporwave", graph, {})


class DetectSbmTests(TestCase):
    """Live SBM detector tests — skipped unless graph-tool is installed (conda/apt, not pip)."""

    def setUp(self) -> None:
        import importlib.util

        if importlib.util.find_spec("graph_tool") is None:
            self.skipTest("graph-tool not installed")
        import networkx as nx

        # Two clear blocks {a,b,c} and {d,e,f}, densely citing within, sparsely across.
        self.graph = nx.DiGraph()
        within = [("a", "b"), ("b", "c"), ("c", "a"), ("d", "e"), ("e", "f"), ("f", "d")]
        for s, t in within:
            self.graph.add_edge(s, t, weight=1.0)
        self.graph.add_edge("c", "d", weight=1.0)  # single bridge

    def test_returns_partition_over_all_nodes(self) -> None:
        from network.community import detect_sbm

        community_map, palette, confidence = detect_sbm(self.graph, "vaporwave", "FLAT")
        self.assertEqual(set(community_map), set(self.graph.nodes()))
        self.assertTrue(all(isinstance(v, int) for v in community_map.values()))
        self.assertEqual(set(community_map.values()), set(palette))
        self.assertIsNone(confidence)  # no refine → no confidence companion

    def test_modes_are_deterministic(self) -> None:
        from network.community import detect_sbm

        for mode in ("FLAT", "NESTED"):
            first, _, _ = detect_sbm(self.graph, "vaporwave", mode)
            second, _, _ = detect_sbm(self.graph, "vaporwave", mode)
            self.assertEqual(first, second, f"{mode} partition not reproducible under fixed seed")

    def test_weighted_fit_accepts_integer_counts(self) -> None:
        from network.community import detect_sbm

        community_map, _, _ = detect_sbm(self.graph, "vaporwave", "FLAT", weights="POISSON")
        self.assertEqual(set(community_map), set(self.graph.nodes()))

    def test_weighted_fit_accepts_positive_reals(self) -> None:
        from network.community import detect_sbm

        import networkx as nx

        graph = nx.DiGraph()
        for s, t in self.graph.edges():
            graph.add_edge(s, t, weight=0.37)
        community_map, _, _ = detect_sbm(graph, "vaporwave", "FLAT", weights="EXPONENTIAL")
        self.assertEqual(set(community_map), set(graph.nodes()))

    def test_mcmc_refine_returns_confidence_per_node(self) -> None:
        from network.community import detect_sbm

        community_map, _, confidence = detect_sbm(self.graph, "vaporwave", "FLAT", refine="MCMC")
        self.assertEqual(set(community_map), set(self.graph.nodes()))
        self.assertEqual(set(confidence), set(self.graph.nodes()))
        for value in confidence.values():
            self.assertGreaterEqual(value, 0.0)
            self.assertLessEqual(value, 1.0)


class DetectSbmValidationTests(TestCase):
    """Weight-model validation fires before the graph-tool import, so these run everywhere."""

    def _graph(self, weight: float) -> "Any":
        import networkx as nx

        graph = nx.DiGraph()
        graph.add_edge("a", "b", weight=weight)
        graph.add_edge("b", "c", weight=weight)
        return graph

    def test_poisson_rejects_fractional_weights(self) -> None:
        from network.community import detect_sbm

        with self.assertRaisesRegex(ValueError, "POISSON.*TOTAL"):
            detect_sbm(self._graph(0.4), "vaporwave", "FLAT", weights="POISSON")

    def test_exponential_rejects_non_positive_weights(self) -> None:
        from network.community import detect_sbm

        with self.assertRaisesRegex(ValueError, "EXPONENTIAL.*positive"):
            detect_sbm(self._graph(0.0), "vaporwave", "FLAT", weights="EXPONENTIAL")


class DetectSbmAssortativeTests(TestCase):
    """Planted-partition detector — live tests skipped unless graph-tool is installed."""

    def _graph(self) -> "Any":
        graph = nx.DiGraph()
        for s, t in [("a", "b"), ("b", "c"), ("c", "a"), ("d", "e"), ("e", "f"), ("f", "d"), ("c", "d")]:
            graph.add_edge(s, t, weight=1.0)
        return graph

    def test_without_graph_tool_raises_value_error(self) -> None:
        import importlib.util

        if importlib.util.find_spec("graph_tool") is not None:
            self.skipTest("graph-tool installed — error path not exercised")
        from network.community import detect_sbm_assortative

        with self.assertRaisesRegex(ValueError, "graph-tool"):
            detect_sbm_assortative(self._graph(), "vaporwave")

    def test_returns_partition_over_all_nodes(self) -> None:
        import importlib.util

        if importlib.util.find_spec("graph_tool") is None:
            self.skipTest("graph-tool not installed")
        from network.community import detect_sbm_assortative

        community_map, palette, confidence = detect_sbm_assortative(self._graph(), "vaporwave")
        self.assertEqual(set(community_map), set(self._graph().nodes()))
        self.assertEqual(set(community_map.values()), set(palette))
        self.assertIsNone(confidence)

    def test_deterministic_under_fixed_seed(self) -> None:
        import importlib.util

        if importlib.util.find_spec("graph_tool") is None:
            self.skipTest("graph-tool not installed")
        from network.community import detect_sbm_assortative

        first, _, _ = detect_sbm_assortative(self._graph(), "vaporwave")
        second, _, _ = detect_sbm_assortative(self._graph(), "vaporwave")
        self.assertEqual(first, second)

    def test_mcmc_refine_returns_confidence(self) -> None:
        import importlib.util

        if importlib.util.find_spec("graph_tool") is None:
            self.skipTest("graph-tool not installed")
        from network.community import detect_sbm_assortative

        community_map, _, confidence = detect_sbm_assortative(self._graph(), "vaporwave", refine="MCMC")
        self.assertEqual(set(confidence), set(community_map))
        for value in confidence.values():
            self.assertGreaterEqual(value, 0.0)
            self.assertLessEqual(value, 1.0)


# ---------------------------------------------------------------------------
# graph_builder.py — build_graph
# ---------------------------------------------------------------------------


class BuildGraphTests(TestCase):
    def setUp(self) -> None:
        self.label = make_label("Org1", color="#FF0000")
        self.ch1 = make_channel(telegram_id=1, label=self.label, title="Channel 1")
        self.ch2 = make_channel(telegram_id=2, label=self.label, title="Channel 2")

    def _create_forward(self) -> Message:
        """Create a message in ch2 forwarded from ch1 and refresh degrees."""
        msg = Message.objects.create(telegram_id=1, channel=self.ch2, forwarded_from=self.ch1)
        self.ch1.save()
        self.ch2.save()
        return msg

    def test_raises_value_error_when_no_edges(self) -> None:
        with self.assertRaises(ValueError):
            build_graph()

    def test_builds_graph_with_forwarded_message_edges(self) -> None:
        self._create_forward()
        graph, channel_dict, edge_list, channel_qs = build_graph()
        self.assertGreater(len(edge_list), 0)

    def test_both_channels_appear_in_channel_dict(self) -> None:
        self._create_forward()
        _, channel_dict, _, _ = build_graph()
        self.assertIn(str(self.ch1.pk), channel_dict)
        self.assertIn(str(self.ch2.pk), channel_dict)

    def test_edge_weight_normalized_to_max_10(self) -> None:
        self._create_forward()
        graph, _, edge_list, _ = build_graph()
        for _u, _v, data in graph.edges(data=True):
            self.assertLessEqual(data["weight"], 10.0)
            self.assertGreater(data["weight"], 0)

    def test_edges_run_amplifier_to_source(self) -> None:
        """build_graph fixes the citation orientation: amplifier→source (citing→cited)."""
        self._create_forward()  # ch2 forwards content from ch1 → edge ch2 → ch1
        graph, _, _, _ = build_graph()
        self.assertIn((str(self.ch2.pk), str(self.ch1.pk)), graph.edges())
        self.assertNotIn((str(self.ch1.pk), str(self.ch2.pk)), graph.edges())

    def test_builds_graph_with_reference_edges(self) -> None:
        msg = Message.objects.create(telegram_id=1, channel=self.ch2)
        msg.references.add(self.ch1)
        graph, _, edge_list, _ = build_graph()
        self.assertGreater(len(edge_list), 0)

    def test_draw_dead_leaves_includes_channels_with_in_degree(self) -> None:
        ch3 = make_channel(telegram_id=3, label=None, title="Dead Leaf")
        # ch2 (in target) forwards from ch3 → ch3 gets in_degree > 0
        Message.objects.create(telegram_id=10, channel=self.ch2, forwarded_from=ch3)
        ch3.refresh_degrees()
        ch3.refresh_from_db()
        self.assertGreater(ch3.in_degree or 0, 0)
        # Create a ch1↔ch2 edge so graph is valid
        self._create_forward()
        _, channel_dict_dl, _, _ = build_graph(draw_dead_leaves=True)
        self.assertIn(str(ch3.pk), channel_dict_dl)

    def test_draw_dead_leaves_false_excludes_out_of_target(self) -> None:
        ch3 = make_channel(telegram_id=3, label=None, title="Dead Leaf")
        self._create_forward()
        _, channel_dict, _, _ = build_graph(draw_dead_leaves=False)
        self.assertNotIn(str(ch3.pk), channel_dict)

    def test_descriptive_group_partitions_by_out_of_target_labels(self) -> None:
        """Regression: a non-primary partition group whose labels are all out-of-target
        (a "Nation" group used only for colouring) must still resolve a window label per
        node, so its LABELGROUP<id> colouring works. The primary group keeps resolving
        in-target labels only."""
        nation = label_group("Nation", is_primary=False)
        belgium = make_label("Belgium", color="#112233", is_in_target=False, group=nation)
        attribute(self.ch1, belgium)
        self._create_forward()  # ch1↔ch2 edge so the build is valid

        _, channel_dict, _, _ = build_graph()
        node = channel_dict[str(self.ch1.pk)]["data"]
        # The descriptive group is populated despite its label being out-of-target …
        self.assertEqual(node["group_partitions"].get(nation.pk), (belgium.pk, belgium.color))
        # … and detect_labelgroup colours the node by it.
        community_map, palette = detect_labelgroup(nation.pk, channel_dict)
        self.assertEqual(community_map[str(self.ch1.pk)], belgium.pk)
        self.assertIn(belgium.pk, palette)
        # The primary group's representative label is unchanged (in-target identity).
        self.assertEqual(node["resolved_org_id"], self.label.pk)

    def test_draw_dead_leaves_survive_windowed_build(self) -> None:
        """Regression: a dead leaf cited inside the window must survive a windowed
        (e.g. per-year timeline) build, not be dropped by the inactive-channel filter
        — which keys off own-message counts that out-of-target channels never have."""
        leaf = make_channel(telegram_id=3, label=None, title="Dead Leaf")
        # In-target ch2 forwards from the leaf within the 2023 window.
        Message.objects.create(
            telegram_id=10,
            channel=self.ch2,
            forwarded_from=leaf,
            date=datetime.datetime(2023, 6, 15, tzinfo=datetime.timezone.utc),
        )
        leaf.refresh_degrees()
        _, channel_dict, _, channel_qs = build_graph(
            draw_dead_leaves=True,
            start_date=datetime.date(2023, 1, 1),
            end_date=datetime.date(2023, 12, 31),
        )
        self.assertIn(str(leaf.pk), channel_dict)
        self.assertIn(leaf, channel_qs)

    def test_dead_leaf_cited_only_outside_window_is_dropped(self) -> None:
        """A dead leaf whose only citation falls outside the window must not appear:
        the degree-0 orphan sweep still prunes it once windowed edges are known."""
        leaf = make_channel(telegram_id=3, label=None, title="Dead Leaf")
        # Sole citation of the leaf is in 2022, outside the 2023 window.
        Message.objects.create(
            telegram_id=10,
            channel=self.ch2,
            forwarded_from=leaf,
            date=datetime.datetime(2022, 6, 15, tzinfo=datetime.timezone.utc),
        )
        # An in-window edge so the graph is valid and ch1/ch2 stay active.
        Message.objects.create(
            telegram_id=11,
            channel=self.ch2,
            forwarded_from=self.ch1,
            date=datetime.datetime(2023, 6, 15, tzinfo=datetime.timezone.utc),
        )
        Message.objects.create(
            telegram_id=12,
            channel=self.ch1,
            date=datetime.datetime(2023, 3, 1, tzinfo=datetime.timezone.utc),
        )
        leaf.refresh_degrees()
        _, channel_dict, _, _ = build_graph(
            draw_dead_leaves=True,
            start_date=datetime.date(2023, 1, 1),
            end_date=datetime.date(2023, 12, 31),
        )
        self.assertNotIn(str(leaf.pk), channel_dict)

    def test_returns_queryset_of_channels(self) -> None:
        self._create_forward()
        _, _, _, channel_qs = build_graph()
        self.assertIn(self.ch1, channel_qs)
        self.assertIn(self.ch2, channel_qs)

    def _create_forward_on_date(self, date: datetime.datetime) -> Message:
        msg = Message.objects.create(telegram_id=99, channel=self.ch2, forwarded_from=self.ch1, date=date)
        self.ch1.save()
        self.ch2.save()
        return msg

    def test_startdate_excludes_messages_before_date(self) -> None:
        self._create_forward_on_date(datetime.datetime(2023, 1, 15, tzinfo=datetime.timezone.utc))
        with self.assertRaises(ValueError):
            build_graph(start_date=datetime.date(2023, 2, 1))

    def test_enddate_excludes_messages_after_date(self) -> None:
        self._create_forward_on_date(datetime.datetime(2023, 3, 1, tzinfo=datetime.timezone.utc))
        with self.assertRaises(ValueError):
            build_graph(end_date=datetime.date(2023, 2, 28))

    def test_date_range_includes_matching_messages(self) -> None:
        self._create_forward_on_date(datetime.datetime(2023, 6, 15, tzinfo=datetime.timezone.utc))
        # ch1 needs a published message in range so it isn't removed from the graph
        Message.objects.create(
            telegram_id=100, channel=self.ch1, date=datetime.datetime(2023, 6, 10, tzinfo=datetime.timezone.utc)
        )
        graph, _, edge_list, _ = build_graph(
            start_date=datetime.date(2023, 6, 1),
            end_date=datetime.date(2023, 6, 30),
        )
        self.assertGreater(len(edge_list), 0)

    def test_date_filter_removes_channels_with_no_messages_in_range(self) -> None:
        self._create_forward_on_date(datetime.datetime(2022, 1, 1, tzinfo=datetime.timezone.utc))
        # Add a third channel with a message in range to ensure the graph has edges
        ch3 = make_channel(telegram_id=3, label=self.label, title="Channel 3")
        Message.objects.create(
            telegram_id=100,
            channel=ch3,
            forwarded_from=self.ch1,
            date=datetime.datetime(2023, 6, 15, tzinfo=datetime.timezone.utc),
        )
        # ch1 also needs a published message in range so it remains in the filtered graph
        Message.objects.create(
            telegram_id=101,
            channel=self.ch1,
            date=datetime.datetime(2023, 3, 1, tzinfo=datetime.timezone.utc),
        )
        self.ch1.save()
        ch3.save()
        _, channel_dict, _, channel_qs = build_graph(start_date=datetime.date(2023, 1, 1))
        self.assertNotIn(str(self.ch2.pk), channel_dict)
        self.assertNotIn(self.ch2, channel_qs)

    def test_channel_qs_filtered_to_active_channels(self) -> None:
        self._create_forward_on_date(datetime.datetime(2023, 6, 15, tzinfo=datetime.timezone.utc))
        # ch1 needs a published message in range so it isn't removed from the graph
        Message.objects.create(
            telegram_id=100, channel=self.ch1, date=datetime.datetime(2023, 6, 10, tzinfo=datetime.timezone.utc)
        )
        _, _, _, channel_qs = build_graph(
            start_date=datetime.date(2023, 6, 1),
            end_date=datetime.date(2023, 6, 30),
        )
        self.assertIn(self.ch2, channel_qs)


class BuildGraphReferenceCutoffTests(TestCase):
    """t.me/ mention edges must honour a channel's in-target attribution period,
    exactly like forwards — regression for the leak where references made in a
    channel's out-of-period messages slipped into the graph (the reference_counts
    query in graph_builder used to skip channel_cutoff_q)."""

    def setUp(self) -> None:
        self.label = make_label("Org", color="#FF0000")
        # ch1 is in-target only during 2023.
        self.ch1 = make_channel(
            telegram_id=1,
            label=self.label,
            attribution_start=datetime.date(2023, 1, 1),
            attribution_end=datetime.date(2023, 12, 31),
            title="Bounded",
        )
        # ch2 / ch3 are in-target for all time (open period).
        self.ch2 = make_channel(telegram_id=2, label=self.label, title="Cited A")
        self.ch3 = make_channel(telegram_id=3, label=self.label, title="Cited B")

    @staticmethod
    def _connected(graph: nx.DiGraph, a: object, b: object) -> bool:
        edges = graph.edges()
        return (str(a.pk), str(b.pk)) in edges or (str(b.pk), str(a.pk)) in edges

    def test_in_period_mention_creates_edge(self) -> None:
        msg = Message.objects.create(
            telegram_id=10,
            channel=self.ch1,
            date=datetime.datetime(2023, 6, 1, tzinfo=datetime.timezone.utc),
        )
        msg.references.add(self.ch3)
        graph, _, _, _ = build_graph()
        self.assertTrue(self._connected(graph, self.ch1, self.ch3))

    def test_out_of_period_mention_excluded(self) -> None:
        # In-period mention of ch2 → a valid edge and a non-zero PARTIAL_REFERENCES denominator.
        in_period = Message.objects.create(
            telegram_id=11,
            channel=self.ch1,
            date=datetime.datetime(2023, 6, 1, tzinfo=datetime.timezone.utc),
        )
        in_period.references.add(self.ch2)
        # Mention of ch3 dated AFTER ch1's in-target period ends → must not create an edge.
        out_of_period = Message.objects.create(
            telegram_id=12,
            channel=self.ch1,
            date=datetime.datetime(2024, 6, 1, tzinfo=datetime.timezone.utc),
        )
        out_of_period.references.add(self.ch3)
        graph, _, _, _ = build_graph()
        self.assertTrue(self._connected(graph, self.ch1, self.ch2))  # in-period mention kept
        self.assertFalse(self._connected(graph, self.ch1, self.ch3))  # out-of-period mention dropped


# ---------------------------------------------------------------------------
# exporter.py — build_graph_data
# ---------------------------------------------------------------------------


def _make_graph_with_positions() -> tuple[nx.DiGraph, dict, dict[str, tuple[float, float]]]:
    """Return a minimal 2-node directed graph, channel_dict, and positions."""
    graph = nx.DiGraph()
    node_data_1 = {
        "pk": "1",
        "label": "Ch1",
        "organization": "Org A",
        "communities": {"leiden": "1-leiden"},
        "color": "255,0,0",
        "pic": "",
        "url": "https://t.me/ch1",
        "activity_period": "Unknown",
        "fans": 0,
        "in_deg": 0,
        "is_lost": False,
        "is_private": False,
        "messages_count": 0,
        "out_deg": 0,
    }
    node_data_2 = {
        "pk": "2",
        "label": "Ch2",
        "organization": "Org B",
        "communities": {"leiden": "2-leiden"},
        "color": "0,0,255",
        "pic": "",
        "url": "https://t.me/ch2",
        "activity_period": "Unknown",
        "fans": 0,
        "in_deg": 0,
        "is_lost": False,
        "is_private": False,
        "messages_count": 0,
        "out_deg": 0,
    }
    graph.add_node("1", data=node_data_1)
    graph.add_node("2", data=node_data_2)
    graph.add_edge("1", "2", weight=5.0, color="100,100,100")
    positions = {"1": (1.0, 2.0), "2": (3.0, 4.0)}
    return graph, {}, positions


class BuildGraphDataTests(TestCase):
    def setUp(self) -> None:
        self.graph, self.channel_dict, self.positions = _make_graph_with_positions()

    def test_nodes_list_length_matches_graph(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        self.assertEqual(len(graph_data["nodes"]), 2)

    def test_edges_list_length_matches_graph(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        self.assertEqual(len(graph_data["edges"]), 1)

    def test_node_has_all_required_keys(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        required = {
            "id",
            "x",
            "y",
            "label",
            "organization",
            "communities",
            "color",
            "pic",
            "url",
            "activity_period",
            "fans",
            "in_deg",
            "is_lost",
            "is_private",
            "messages_count",
            "out_deg",
        }
        for node in graph_data["nodes"]:
            missing = required - node.keys()
            self.assertFalse(missing, f"Node is missing keys: {missing}")

    def test_edge_has_all_required_keys(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        edge = graph_data["edges"][0]
        for key in ("source", "target", "weight", "color", "id"):
            self.assertIn(key, edge)

    def test_node_positions_are_applied(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        self.assertAlmostEqual(node_map["1"]["x"], 1.0)
        self.assertAlmostEqual(node_map["1"]["y"], 2.0)
        self.assertAlmostEqual(node_map["2"]["x"], 3.0)
        self.assertAlmostEqual(node_map["2"]["y"], 4.0)

    def test_edge_ids_are_sequential(self) -> None:
        graph_data = build_graph_data(self.graph, self.positions)
        edge_ids = [e["id"] for e in graph_data["edges"]]
        self.assertEqual(edge_ids, list(range(len(edge_ids))))


# ---------------------------------------------------------------------------
# exporter.py — apply_base_node_measures
# ---------------------------------------------------------------------------


class ApplyBaseNodeMeasuresTests(TestCase):
    def setUp(self) -> None:
        label = make_label("Org1", color="#FF0000")
        self.ch1 = make_channel(telegram_id=1, label=label, title="Chan1", participants_count=500)
        self.ch2 = make_channel(telegram_id=2, label=label, title="Chan2", participants_count=300)
        # ch1 has a message forwarded from ch2
        Message.objects.create(telegram_id=1, channel=self.ch1, forwarded_from=self.ch2)
        self.ch1.save()
        self.ch2.save()

        self.graph = nx.DiGraph()
        self.graph.add_node(str(self.ch1.pk), data={"pk": str(self.ch1.pk)})
        self.graph.add_node(str(self.ch2.pk), data={"pk": str(self.ch2.pk)})
        self.graph.add_edge(str(self.ch2.pk), str(self.ch1.pk), weight=5.0)
        self.channel_dict = {
            str(self.ch1.pk): {"channel": self.ch1},
            str(self.ch2.pk): {"channel": self.ch2},
        }
        self.graph_data = {
            "nodes": [
                {"id": str(self.ch1.pk), "in_deg": 0, "out_deg": 0, "fans": 0, "messages_count": 0, "label": ""},
                {"id": str(self.ch2.pk), "in_deg": 0, "out_deg": 0, "fans": 0, "messages_count": 0, "label": ""},
            ],
            "edges": [],
        }

    def test_returns_non_empty_measures_labels(self) -> None:
        labels = apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        self.assertIsInstance(labels, list)
        self.assertGreater(len(labels), 0)

    def test_labels_are_key_description_pairs(self) -> None:
        labels = apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        for key, description in labels:
            self.assertIsInstance(key, str)
            self.assertIsInstance(description, str)

    def test_fans_filled_from_participants_count(self) -> None:
        apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertEqual(node_map[str(self.ch1.pk)]["fans"], 500)
        self.assertEqual(node_map[str(self.ch2.pk)]["fans"], 300)

    def test_label_filled_from_channel_title(self) -> None:
        apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertEqual(node_map[str(self.ch1.pk)]["label"], "Chan1")
        self.assertEqual(node_map[str(self.ch2.pk)]["label"], "Chan2")

    def test_in_degree_filled_from_graph(self) -> None:
        apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        # ch1 has one incoming edge from ch2 (weight 5.0)
        self.assertGreater(node_map[str(self.ch1.pk)]["in_deg"], 0)
        self.assertEqual(node_map[str(self.ch2.pk)]["in_deg"], 0)

    def test_messages_count_filled_from_db(self) -> None:
        apply_base_node_measures(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertEqual(node_map[str(self.ch1.pk)]["messages_count"], 1)
        self.assertEqual(node_map[str(self.ch2.pk)]["messages_count"], 0)


# ---------------------------------------------------------------------------
# exporter.py — apply_pagerank
# ---------------------------------------------------------------------------


class ApplyPageRankTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["1", "2", "3"])
        self.graph.add_edges_from([("1", "2"), ("2", "3"), ("3", "1")])
        self.graph_data = {
            "nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}],
            "edges": [],
        }

    def test_adds_pagerank_key_to_all_nodes(self) -> None:
        apply_pagerank(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("pagerank", node)

    def test_returns_list_with_pagerank_label(self) -> None:
        labels = apply_pagerank(self.graph_data, self.graph)
        keys = [k for k, _ in labels]
        self.assertIn("pagerank", keys)

    def test_pagerank_values_are_floats(self) -> None:
        apply_pagerank(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIsInstance(node["pagerank"], float)

    def test_pagerank_values_sum_to_one(self) -> None:
        apply_pagerank(self.graph_data, self.graph)
        total = sum(n["pagerank"] for n in self.graph_data["nodes"])
        self.assertAlmostEqual(total, 1.0, places=5)


# ---------------------------------------------------------------------------
# exporter.py — find_main_component
# ---------------------------------------------------------------------------


class FindMainComponentTests(TestCase):
    def test_returns_a_set(self) -> None:
        graph = nx.DiGraph()
        graph.add_edges_from([("a", "b"), ("b", "c")])
        result = find_main_component(graph)
        self.assertIsInstance(result, set)

    def test_returns_largest_weakly_connected_component(self) -> None:
        graph = nx.DiGraph()
        # Component 1: a-b-c (size 3)
        graph.add_edges_from([("a", "b"), ("b", "c")])
        # Component 2: d-e (size 2)
        graph.add_edge("d", "e")
        result = find_main_component(graph)
        self.assertEqual(result, {"a", "b", "c"})

    def test_single_component_graph_returns_all_nodes(self) -> None:
        graph = nx.DiGraph()
        graph.add_edges_from([("x", "y"), ("y", "z"), ("z", "x")])
        result = find_main_component(graph)
        self.assertEqual(result, {"x", "y", "z"})


# ---------------------------------------------------------------------------
# exporter.py — reposition_isolated_nodes
# ---------------------------------------------------------------------------


class RepositionIsolatedNodesTests(TestCase):
    def _make_graph_data(self) -> dict:
        return {
            "nodes": [
                {"id": "1", "x": 10.0, "y": 10.0},  # main component
                {"id": "2", "x": 20.0, "y": 20.0},  # main component
                {"id": "3", "x": 0.0, "y": 0.0},  # isolated
            ],
            "edges": [],
        }

    def test_main_component_node_positions_unchanged(self) -> None:
        graph_data = self._make_graph_data()
        reposition_isolated_nodes(graph_data, {"1", "2"})
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        self.assertAlmostEqual(node_map["1"]["x"], 10.0)
        self.assertAlmostEqual(node_map["1"]["y"], 10.0)
        self.assertAlmostEqual(node_map["2"]["x"], 20.0)
        self.assertAlmostEqual(node_map["2"]["y"], 20.0)

    def test_isolated_node_repositioned_near_max(self) -> None:
        graph_data = self._make_graph_data()
        reposition_isolated_nodes(graph_data, {"1", "2"})
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        # First isolated node lands at (max_x - 0*d, max_y - 0*d) = (max_x, max_y)
        self.assertAlmostEqual(node_map["3"]["x"], 20.0)
        self.assertAlmostEqual(node_map["3"]["y"], 20.0)

    def test_all_nodes_retain_x_y_coordinates(self) -> None:
        graph_data = self._make_graph_data()
        reposition_isolated_nodes(graph_data, {"1", "2"})
        for node in graph_data["nodes"]:
            self.assertIn("x", node)
            self.assertIn("y", node)

    def test_empty_isolated_set_leaves_all_nodes_in_place(self) -> None:
        graph_data = self._make_graph_data()
        reposition_isolated_nodes(graph_data, {"1", "2", "3"})
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        self.assertAlmostEqual(node_map["3"]["x"], 0.0)
        self.assertAlmostEqual(node_map["3"]["y"], 0.0)


# ---------------------------------------------------------------------------
# exporter.py — ensure_graph_root
# ---------------------------------------------------------------------------


class EnsureGraphRootTests(TestCase):
    def test_creates_directory_if_not_present(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "graph_root")
            ensure_graph_root(target)
            self.assertTrue(os.path.isdir(target))

    def test_removes_existing_contents(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "graph_root")
            os.makedirs(target)
            sentinel = os.path.join(target, "old_file.txt")
            with open(sentinel, "w") as f:
                f.write("stale data")
            ensure_graph_root(target)
            self.assertFalse(os.path.exists(sentinel))
            self.assertTrue(os.path.isdir(target))

    def test_succeeds_even_without_map_template(self) -> None:
        # webapp_engine/map may not exist in test environment; ensure_graph_root
        # catches the OSError and should not propagate it.
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "graph_root")
            try:
                ensure_graph_root(target)
            except OSError:
                self.fail("ensure_graph_root raised OSError unexpectedly")


# ---------------------------------------------------------------------------
# exporter.py — write_graph_files
# ---------------------------------------------------------------------------


class WriteGraphFilesTests(TestCase):
    def setUp(self) -> None:
        label = make_label("Org1", color="#FF0000")
        self.ch = make_channel(telegram_id=1, label=label, title="Chan1")
        self.channel_qs = Channel.objects.filter(pk=self.ch.pk)
        self.graph_data = {"nodes": [{"id": "1", "x": 0.0, "y": 0.0}], "edges": []}
        self.communities_data = {
            "leiden": {
                "main_groups": {"1": "1-leiden"},
                "groups": [("1", 5, "1-leiden", "#FF0000")],
            }
        }
        self.measures_labels = [("in_deg", "Inbound connections"), ("pagerank", "PageRank")]

    def test_writes_output_json_with_nodes_and_edges(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            write_graph_files(self.graph_data, self.communities_data, self.measures_labels, self.channel_qs, tmpdir)
            out_file = os.path.join(tmpdir, "data", "channel_position.json")
            self.assertTrue(os.path.exists(out_file))
            with open(out_file) as f:
                data = json.load(f)
            self.assertIn("nodes", data)
            self.assertIn("edges", data)

    def test_writes_accessory_json_with_required_structure(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            write_graph_files(self.graph_data, self.communities_data, self.measures_labels, self.channel_qs, tmpdir)
            channels_file = os.path.join(tmpdir, "data", "channels.json")
            self.assertTrue(os.path.exists(channels_file))
            with open(channels_file) as f:
                acc_data = json.load(f)
            for key in ("measures", "total_pages_count"):
                self.assertIn(key, acc_data)
            communities_file = os.path.join(tmpdir, "data", "communities.json")
            self.assertTrue(os.path.exists(communities_file))

    def test_accessory_total_pages_count_matches_queryset(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            write_graph_files(self.graph_data, self.communities_data, self.measures_labels, self.channel_qs, tmpdir)
            with open(os.path.join(tmpdir, "data", "channels.json")) as f:
                acc_data = json.load(f)
            self.assertEqual(acc_data["total_pages_count"], self.channel_qs.count())


# ---------------------------------------------------------------------------
# community.py — detect_leiden
# ---------------------------------------------------------------------------


class DetectLeidenTests(TestCase):
    def setUp(self) -> None:
        # Two clear clusters connected by a weak bridge
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["a", "b", "c", "d", "e", "f"])
        self.graph.add_edges_from(
            [
                ("a", "b"),
                ("b", "c"),
                ("c", "a"),  # cluster 1
                ("d", "e"),
                ("e", "f"),
                ("f", "d"),  # cluster 2
                ("c", "d"),  # bridge
            ]
        )

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_returns_community_map_and_palette(self, _mock: MagicMock) -> None:
        community_map, palette = detect_leiden(self.graph, "SomePalette")
        self.assertIsInstance(community_map, dict)
        self.assertIsInstance(palette, dict)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_all_nodes_assigned(self, _mock: MagicMock) -> None:
        community_map, _ = detect_leiden(self.graph, "SomePalette")
        self.assertEqual(set(community_map.keys()), set(self.graph.nodes()))

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_community_ids_start_at_1(self, _mock: MagicMock) -> None:
        community_map, _ = detect_leiden(self.graph, "SomePalette")
        self.assertGreaterEqual(min(community_map.values()), 1)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_palette_covers_all_detected_communities(self, _mock: MagicMock) -> None:
        community_map, palette = detect_leiden(self.graph, "SomePalette")
        for community_id in community_map.values():
            self.assertIn(community_id, palette)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_isolated_nodes_merged_into_same_community(self, _mock: MagicMock) -> None:
        graph = nx.DiGraph()
        graph.add_edge("a", "b")
        graph.add_node("iso1")
        graph.add_node("iso2")
        community_map, _ = detect_leiden(graph, "SomePalette")
        self.assertEqual(community_map["iso1"], community_map["iso2"])


# ---------------------------------------------------------------------------
# community.py — detect_louvain
# ---------------------------------------------------------------------------


class DetectLouvainTests(TestCase):
    def setUp(self) -> None:
        # Two clear clusters connected by a weak bridge (same fixture as Leiden).
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["a", "b", "c", "d", "e", "f"])
        self.graph.add_edges_from([("a", "b"), ("b", "c"), ("c", "a"), ("d", "e"), ("e", "f"), ("f", "d"), ("c", "d")])

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_returns_community_map_and_palette(self, _mock: MagicMock) -> None:
        community_map, palette = detect_louvain(self.graph, "SomePalette")
        self.assertIsInstance(community_map, dict)
        self.assertIsInstance(palette, dict)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_all_nodes_assigned(self, _mock: MagicMock) -> None:
        community_map, _ = detect_louvain(self.graph, "SomePalette")
        self.assertEqual(set(community_map.keys()), set(self.graph.nodes()))

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_palette_covers_all_detected_communities(self, _mock: MagicMock) -> None:
        community_map, palette = detect_louvain(self.graph, "SomePalette")
        for community_id in community_map.values():
            self.assertIn(community_id, palette)

    @patch("network.community.palette_colors", return_value=["#ff0000", "#00ff00", "#0000ff"])
    def test_seed_makes_partition_reproducible(self, _mock: MagicMock) -> None:
        first, _ = detect_louvain(self.graph, "SomePalette")
        second, _ = detect_louvain(self.graph, "SomePalette")
        self.assertEqual(first, second)


# ---------------------------------------------------------------------------
# community.py — detect() dispatcher
# ---------------------------------------------------------------------------


class DetectDispatcherTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["a", "b"])
        self.graph.add_edge("a", "b")
        self.group = label_group()
        self.label = make_label("Org")
        self.ch1 = make_channel(telegram_id=1, label=self.label)
        self.ch2 = make_channel(telegram_id=2, label=self.label)
        self.channel_dict = {
            str(self.ch1.pk): {"channel": self.ch1, "data": {}},
            str(self.ch2.pk): {"channel": self.ch2, "data": {}},
        }

    @patch("network.community.detect_kcore")
    def test_kcore_strategy_calls_detect_kcore(self, mock_detect: MagicMock) -> None:
        from network.community import detect

        mock_detect.return_value = ({}, {})
        detect("KCORE", "palette", self.graph, self.channel_dict)
        mock_detect.assert_called_once_with(self.graph, "palette", reverse=False)

    @patch("network.community.detect_leiden")
    def test_leiden_strategy_calls_detect_leiden(self, mock_detect: MagicMock) -> None:
        from network.community import detect

        mock_detect.return_value = ({}, {})
        detect("LEIDEN", "palette", self.graph, self.channel_dict)
        mock_detect.assert_called_once_with(self.graph, "palette", reverse=False)

    @patch("network.community.detect_louvain")
    def test_louvain_strategy_calls_detect_louvain(self, mock_detect: MagicMock) -> None:
        from network.community import detect

        mock_detect.return_value = ({}, {})
        detect("LOUVAIN", "palette", self.graph, self.channel_dict)
        mock_detect.assert_called_once_with(self.graph, "palette", reverse=False)

    @patch("network.community.detect_labelgroup")
    def test_labelgroup_strategy_dispatches_to_detect_labelgroup(self, mock_detect: MagicMock) -> None:
        from network.community import detect

        mock_detect.return_value = ({}, {})
        detect(self.group.token, "palette", self.graph, self.channel_dict)
        # LABELGROUP<id> resolves the group pk and reads the partition from channel_dict;
        # the palette_name / reverse flags do not apply (palette comes from label colours).
        mock_detect.assert_called_once_with(self.group.pk, self.channel_dict)

    def test_unknown_strategy_raises_value_error(self) -> None:
        from network.community import detect

        with self.assertRaises(ValueError, msg="Unknown community strategy"):
            detect("", "palette", self.graph, self.channel_dict)


# ---------------------------------------------------------------------------
# exporter.py — copy_channel_media
# ---------------------------------------------------------------------------


class CopyChannelMediaTests(TestCase):
    def setUp(self) -> None:
        self.label = make_label("Org")

    def test_channel_without_username_is_skipped(self) -> None:
        ch = make_channel(telegram_id=1, label=self.label, username="")
        with tempfile.TemporaryDirectory() as tmpdir:
            from network.exporter import copy_channel_media

            # No error, nothing copied
            copy_channel_media(Channel.objects.filter(pk=ch.pk), tmpdir)
            self.assertFalse(os.path.exists(os.path.join(tmpdir, "channels")))

    def test_missing_source_dir_is_silently_ignored(self) -> None:
        ch = make_channel(telegram_id=2, label=self.label, username="testchan")
        with tempfile.TemporaryDirectory() as tmpdir:
            from network.exporter import copy_channel_media

            # MEDIA_ROOT/channels/testchan/profile doesn't exist → FileNotFoundError silently caught
            copy_channel_media(Channel.objects.filter(pk=ch.pk), tmpdir)
            self.assertFalse(os.path.exists(os.path.join(tmpdir, "channels", "testchan")))

    def test_existing_source_dir_is_copied(self) -> None:
        ch = make_channel(telegram_id=3, label=self.label, username="copychan")
        with tempfile.TemporaryDirectory() as media_root, tempfile.TemporaryDirectory() as output_root:
            src = os.path.join(media_root, "channels", "copychan", "profile")
            os.makedirs(src)
            sentinel = os.path.join(src, "photo.jpg")
            with open(sentinel, "w") as f:
                f.write("fake image")
            from django.test import override_settings

            from network.exporter import copy_channel_media

            with override_settings(MEDIA_ROOT=media_root):
                copy_channel_media(Channel.objects.filter(pk=ch.pk), output_root)
            dst = os.path.join(output_root, "media", "channels", "copychan", "profile", "photo.jpg")
            self.assertTrue(os.path.exists(dst))

    def test_oserror_on_copy_is_logged_not_raised(self) -> None:
        ch = make_channel(telegram_id=4, label=self.label, username="errchan")
        with tempfile.TemporaryDirectory() as media_root, tempfile.TemporaryDirectory() as output_root:
            src = os.path.join(media_root, "channels", "errchan", "profile")
            os.makedirs(src)
            dst_parent = os.path.join(output_root, "channels", "errchan", "profile")
            os.makedirs(dst_parent)  # pre-create dst so copytree raises OSError (already exists)
            from network.exporter import copy_channel_media

            with override_settings(MEDIA_ROOT=media_root):
                try:
                    copy_channel_media(Channel.objects.filter(pk=ch.pk), output_root)
                except OSError:
                    self.fail("copy_channel_media raised OSError unexpectedly")


# ---------------------------------------------------------------------------
# structural_analysis management command
# ---------------------------------------------------------------------------


_EXPORT_CMD = "network.management.commands.structural_analysis"


class TemporalPrecomputeTests(TestCase):
    """_compute_temporal_partitions over real messages in two calendar years."""

    def setUp(self) -> None:
        label = make_label("Org1", color="#FF0000")
        self.channels = [make_channel(telegram_id=i, label=label, title=f"C{i}") for i in range(1, 5)]
        telegram_id = 100
        # Two mutual-forward pairs, repeated in 2023 and 2024: {C1,C2} and {C3,C4}.
        for year in (2023, 2024):
            for src, dst in [
                (self.channels[0], self.channels[1]),
                (self.channels[1], self.channels[0]),
                (self.channels[2], self.channels[3]),
                (self.channels[3], self.channels[2]),
            ]:
                telegram_id += 1
                Message.objects.create(
                    telegram_id=telegram_id,
                    channel=dst,
                    forwarded_from=src,
                    date=datetime.datetime(year, 6, 1, tzinfo=datetime.timezone.utc),
                )

    def test_precompute_builds_coupled_per_year_maps(self) -> None:
        from django.core.management.base import OutputWrapper

        from network.management.commands.structural_analysis import Command

        cmd = Command()
        cmd.stdout = OutputWrapper(io.StringIO())
        parser = cmd.create_parser("manage.py", "structural_analysis")
        options = vars(
            parser.parse_args(
                [
                    "--community-strategies",
                    "LEIDEN,LEIDEN_TEMPORAL",
                    "--timeline-step",
                    "year",
                    "--edge-weight-strategy",
                    "PARTIAL_REFERENCES",
                ]
            )
        )
        opts = cmd._resolve_options(options)
        temporal_instances = [i for i in opts.communities_strategy if i.name == "LEIDEN_TEMPORAL"]
        results = cmd._compute_temporal_partitions(opts, temporal_instances)

        (key,) = [inst.key for inst in temporal_instances]
        per_year, plurality, palette = results[key]
        self.assertEqual(set(per_year), {2023, 2024})
        pk1, pk2, pk3, pk4 = (str(c.pk) for c in self.channels)
        for pk in (pk1, pk2, pk3, pk4):
            # Same structure both years + identity coupling → stable ids, and the plurality
            # summary agrees with them.
            self.assertEqual(per_year[2023][pk], per_year[2024][pk])
            self.assertEqual(plurality[pk], per_year[2023][pk])
        self.assertEqual(per_year[2023][pk1], per_year[2023][pk2])
        self.assertEqual(per_year[2023][pk3], per_year[2023][pk4])
        self.assertNotEqual(per_year[2023][pk1], per_year[2023][pk3])
        for cid in plurality.values():
            self.assertIn(cid, palette)

    def test_precompute_fails_with_single_year(self) -> None:
        from django.core.management.base import CommandError, OutputWrapper

        from network.management.commands.structural_analysis import Command

        Message.objects.filter(date__year=2024).delete()
        cmd = Command()
        cmd.stdout = OutputWrapper(io.StringIO())
        parser = cmd.create_parser("manage.py", "structural_analysis")
        options = vars(parser.parse_args(["--community-strategies", "LEIDEN_TEMPORAL", "--timeline-step", "year"]))
        opts = cmd._resolve_options(options)
        temporal_instances = list(opts.communities_strategy)
        with self.assertRaisesRegex(CommandError, "at least two"):
            cmd._compute_temporal_partitions(opts, temporal_instances)


def _patch_export_pipeline() -> list:
    """Return a list of patch decorators for all structural_analysis submodule calls."""
    targets = [
        f"{_EXPORT_CMD}.graph_builder.build_graph",
        f"{_EXPORT_CMD}.community.detect",
        f"{_EXPORT_CMD}.community.apply_to_graph",
        f"{_EXPORT_CMD}.community.apply_edge_colors",
        f"{_EXPORT_CMD}.community.build_communities_payload",
        f"{_EXPORT_CMD}.layout.forceatlas2_positions",
        f"{_EXPORT_CMD}.exporter.build_graph_data",
        f"{_EXPORT_CMD}.exporter.find_main_component",
        f"{_EXPORT_CMD}.measures.apply_base_node_measures",
        f"{_EXPORT_CMD}.measures.apply_pagerank",
        f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes",
        f"{_EXPORT_CMD}.exporter.ensure_graph_root",
        f"{_EXPORT_CMD}.exporter.write_graph_files",
        f"{_EXPORT_CMD}.tables.write_table_html",
        f"{_EXPORT_CMD}.tables.write_table_xlsx",
        f"{_EXPORT_CMD}.exporter.copy_channel_media",
    ]
    # Apply in reverse so the first target becomes the first positional arg
    decorators = [patch(t) for t in reversed(targets)]
    return decorators


class ExportNetworkCommandTests(TestCase):
    def setUp(self) -> None:
        # The community pipeline (detect / apply_to_graph / build_communities_payload) is mocked,
        # so any valid --community-strategies token exercises it. LABELGROUP<id> partitions are
        # rejected by the command's own _validate_settings (it only accepts the algorithm strategies
        # in VALID_STRATEGIES), so an algorithm token is used here — LEIDEN, replacing the old
        # ORGANIZATION metadata token.
        self.strategy_token = "LEIDEN"
        _b = _EXPORT_CMD
        for target in [
            f"{_b}.exporter.write_meta_json",
            f"{_b}.exporter.write_robots_txt",
            f"{_b}.exporter.write_summary_json",
            f"{_b}.tables.write_index_html",
            f"{_b}.tables.write_network_metrics_json",
            f"{_b}.tables.write_community_metrics_json",
            f"{_b}.tables.write_network_table_html",
            f"{_b}.tables.write_community_table_html",
            f"{_b}.tables.write_network_table_xlsx",
            f"{_b}.tables.write_community_table_xlsx",
            f"{_b}.community_stats.compute_community_metrics",
            f"{_b}.os.makedirs",
            f"{_b}.os.rename",
        ]:
            p = patch(target)
            p.start()
            self.addCleanup(p.stop)

    def _configure_happy_path(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        mock_communities_payload: MagicMock,
    ) -> None:
        fake_graph = nx.DiGraph()
        fake_graph.add_node("1")
        fake_qs = MagicMock()
        fake_qs.filter.return_value = fake_qs
        fake_qs.count.return_value = 0
        mock_build.return_value = (fake_graph, {"1": {}}, [["1", "2", 1.0]], fake_qs)
        mock_detect.return_value = ({"1": 1}, {1: (255, 0, 0)})
        mock_layout.return_value = {"1": (0.0, 0.0)}
        mock_graph_data.return_value = {"nodes": [], "edges": []}
        mock_main_comp.return_value = {"1"}
        mock_measures.return_value = [("in_deg", "Inbound")]
        mock_pagerank.return_value = [("pagerank", "PageRank")]
        mock_communities_payload.return_value = {"leiden": {"groups": [], "main_groups": {}}}

    def test_raises_command_error_on_invalid_startdate(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        with self.assertRaises(CommandError):
            call_command("structural_analysis", startdate="not-a-date", stdout=io.StringIO(), stderr=io.StringIO())

    def test_raises_command_error_on_invalid_enddate(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        with self.assertRaises(CommandError):
            call_command("structural_analysis", enddate="2023-13-01", stdout=io.StringIO(), stderr=io.StringIO())

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_raises_command_error_when_no_edges(
        self,
        mock_build: MagicMock,
        *_mocks: MagicMock,
    ) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        mock_build.side_effect = ValueError("There are no relationships between channels.")
        with self.assertRaises(CommandError):
            # graph=True bypasses the bare-CLI early-exit so build_graph is reached.
            call_command("structural_analysis", graph=True, stdout=io.StringIO(), stderr=io.StringIO())

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_all_pipeline_steps_called(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        mock_reposition: MagicMock,
        mock_ensure: MagicMock,
        mock_write: MagicMock,
        mock_table_html: MagicMock,
        mock_table_xls: MagicMock,
        mock_copy: MagicMock,
    ) -> None:
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        # The defaults are factory-empty; pass community_strategies + measures
        # explicitly so the community-detection and measures pipelines run.
        call_command(
            "structural_analysis",
            graph=True,
            html=True,
            community_strategies=self.strategy_token,
            measures="PAGERANK",
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )

        mock_build.assert_called_once()
        mock_detect.assert_called()
        mock_apply_to_graph.assert_called()
        mock_edge_colors.assert_called_once()
        mock_layout.assert_called_once()
        mock_graph_data.assert_called_once()
        mock_main_comp.assert_called_once()
        mock_measures.assert_called_once()
        mock_pagerank.assert_called_once()
        mock_reposition.assert_called_once()
        mock_ensure.assert_called_once()
        mock_write.assert_called_once()
        mock_table_html.assert_called_once()
        mock_table_xls.assert_not_called()
        mock_copy.assert_called_once()

    def test_rejects_out_of_range_backbone_alpha(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        for bad in (-0.1, 1.0, 1.5):
            with self.assertRaises(CommandError):
                call_command(
                    "structural_analysis",
                    community_strategies="LEIDEN",
                    community_backbone_alpha=bad,
                    stdout=io.StringIO(),
                    stderr=io.StringIO(),
                )

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_backbone_alpha_filters_detection_graph(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        *_mocks: MagicMock,
    ) -> None:
        """--community-backbone-alpha runs detection on the disparity-filtered backbone.

        3×3 bipartite fixture with one dominant target per source (weight 100 vs 0.001):
        the six off-diagonal edges are insignificant from both endpoints at α=0.05, so
        detection must see the 3-edge backbone while the full 9-edge graph flows on to
        layout / measures / exports untouched.
        """
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        fake_graph = nx.DiGraph()
        for i, source in enumerate(("s1", "s2", "s3")):
            for j, target in enumerate(("t1", "t2", "t3")):
                fake_graph.add_edge(source, target, weight=100.0 if i == j else 0.001)
        channel_dict = {n: {"data": {}} for n in fake_graph.nodes()}
        fake_qs = MagicMock()
        fake_qs.filter.return_value = fake_qs
        fake_qs.count.return_value = 0
        mock_build.return_value = (fake_graph, channel_dict, [["s1", "t1", 1.0]], fake_qs)

        call_command(
            "structural_analysis",
            graph=True,
            community_strategies="LEIDEN",
            community_backbone_alpha=0.05,
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )

        detect_graph = mock_detect.call_args[0][2]
        self.assertIsNot(detect_graph, fake_graph)
        self.assertEqual(set(detect_graph.nodes()), set(fake_graph.nodes()))  # vertex set preserved
        self.assertEqual(detect_graph.number_of_edges(), 3)
        self.assertEqual(set(detect_graph.edges()), {("s1", "t1"), ("s2", "t2"), ("s3", "t3")})
        self.assertEqual(fake_graph.number_of_edges(), 9)  # full graph untouched
        # apply_to_graph still writes onto the FULL graph, not the backbone.
        self.assertIs(mock_apply_to_graph.call_args[0][0], fake_graph)

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_no_backbone_alpha_detects_on_full_graph(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        *_mocks: MagicMock,
    ) -> None:
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        fake_graph = mock_build.return_value[0]

        call_command(
            "structural_analysis",
            graph=True,
            community_strategies="LEIDEN",
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )
        # Without the flag, detection runs on the very graph build_graph returned.
        self.assertIs(mock_detect.call_args[0][2], fake_graph)

    def test_consensus_requires_two_eligible_inputs(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        # One eligible input (LEIDEN) — KCORE and LABELGROUPs don't count.
        with self.assertRaisesRegex(CommandError, "consensus-eligible"):
            call_command(
                "structural_analysis",
                community_strategies="LEIDEN,KCORE,CONSENSUS",
                stdout=io.StringIO(),
                stderr=io.StringIO(),
            )

    def test_temporal_requires_year_timeline(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        with self.assertRaisesRegex(CommandError, "timeline-step year"):
            call_command(
                "structural_analysis",
                community_strategies="LEIDEN,LEIDEN_TEMPORAL",
                stdout=io.StringIO(),
                stderr=io.StringIO(),
            )

    def test_temporal_not_eligible_as_consensus_input(self) -> None:
        from django.core.management import call_command
        from django.core.management.base import CommandError

        # LEIDEN + LEIDEN_TEMPORAL = one eligible input only (temporal doesn't count).
        with self.assertRaisesRegex(CommandError, "consensus-eligible"):
            call_command(
                "structural_analysis",
                community_strategies="LEIDEN,LEIDEN_TEMPORAL,CONSENSUS",
                timeline_step="year",
                stdout=io.StringIO(),
                stderr=io.StringIO(),
            )

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_consensus_dispatched_after_direct_strategies(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        *_mocks: MagicMock,
    ) -> None:
        """CONSENSUS runs the real detect_consensus over the direct strategies' (mocked) partitions.

        detect() is mocked to hand LEIDEN and LOUVAIN identical two-block partitions over a
        4-node graph, so the real consensus aggregation must recover those blocks and
        apply_to_graph must be called once per strategy (3 times), with the CONSENSUS instance
        receiving a map that co-assigns the agreed pairs.
        """
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        fake_graph = nx.DiGraph()
        fake_graph.add_edges_from([("a", "b"), ("c", "d"), ("b", "c")])
        channel_dict = {n: {"data": {}} for n in fake_graph.nodes()}
        fake_qs = MagicMock()
        fake_qs.filter.return_value = fake_qs
        fake_qs.count.return_value = 0
        mock_build.return_value = (fake_graph, channel_dict, [["a", "b", 1.0]], fake_qs)
        mock_detect.return_value = ({"a": 1, "b": 1, "c": 2, "d": 2}, {1: (255, 0, 0), 2: (0, 255, 0)})

        call_command(
            "structural_analysis",
            graph=True,
            community_strategies="CONSENSUS,LEIDEN,LOUVAIN",  # order must not matter
            community_palette="vaporwave",
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )

        self.assertEqual(mock_detect.call_count, 2)  # LEIDEN + LOUVAIN only; CONSENSUS not via detect()
        applied = {call.args[4].key: call.args[2] for call in mock_apply_to_graph.call_args_list}
        self.assertEqual(set(applied), {"leiden", "louvain", "consensus_threshold_0_5"})
        consensus_map = applied["consensus_threshold_0_5"]
        self.assertEqual(consensus_map["a"], consensus_map["b"])
        self.assertEqual(consensus_map["c"], consensus_map["d"])
        self.assertNotEqual(consensus_map["a"], consensus_map["c"])

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_table_format_none_skips_both(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        mock_reposition: MagicMock,
        mock_ensure: MagicMock,
        mock_write: MagicMock,
        mock_table_html: MagicMock,
        mock_table_xls: MagicMock,
        mock_copy: MagicMock,
    ) -> None:
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        # graph=True bypasses the bare-CLI early-exit; html=False suppresses tables.
        call_command(
            "structural_analysis",
            graph=True,
            html=False,
            community_strategies=self.strategy_token,
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )
        mock_table_html.assert_not_called()
        mock_table_xls.assert_not_called()

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_table_format_xls_only(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        mock_reposition: MagicMock,
        mock_ensure: MagicMock,
        mock_write: MagicMock,
        mock_table_html: MagicMock,
        mock_table_xls: MagicMock,
        mock_copy: MagicMock,
    ) -> None:
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        call_command(
            "structural_analysis",
            html=False,
            xlsx=True,
            community_strategies=self.strategy_token,
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )
        mock_table_html.assert_not_called()
        mock_table_xls.assert_called_once()

    @patch(f"{_EXPORT_CMD}.exporter.copy_channel_media")
    @patch(f"{_EXPORT_CMD}.tables.write_table_xlsx")
    @patch(f"{_EXPORT_CMD}.tables.write_table_html")
    @patch(f"{_EXPORT_CMD}.exporter.write_graph_files")
    @patch(f"{_EXPORT_CMD}.exporter.ensure_graph_root")
    @patch(f"{_EXPORT_CMD}.exporter.reposition_isolated_nodes")
    @patch(f"{_EXPORT_CMD}.measures.apply_pagerank")
    @patch(f"{_EXPORT_CMD}.measures.apply_base_node_measures")
    @patch(f"{_EXPORT_CMD}.exporter.find_main_component")
    @patch(f"{_EXPORT_CMD}.exporter.build_graph_data")
    @patch(f"{_EXPORT_CMD}.layout.forceatlas2_positions")
    @patch(f"{_EXPORT_CMD}.community.build_communities_payload")
    @patch(f"{_EXPORT_CMD}.community.apply_edge_colors")
    @patch(f"{_EXPORT_CMD}.community.apply_to_graph")
    @patch(f"{_EXPORT_CMD}.community.detect")
    @patch(f"{_EXPORT_CMD}.graph_builder.build_graph")
    def test_table_format_html_xls_calls_both(
        self,
        mock_build: MagicMock,
        mock_detect: MagicMock,
        mock_apply_to_graph: MagicMock,
        mock_edge_colors: MagicMock,
        mock_communities_payload: MagicMock,
        mock_layout: MagicMock,
        mock_graph_data: MagicMock,
        mock_main_comp: MagicMock,
        mock_measures: MagicMock,
        mock_pagerank: MagicMock,
        mock_reposition: MagicMock,
        mock_ensure: MagicMock,
        mock_write: MagicMock,
        mock_table_html: MagicMock,
        mock_table_xls: MagicMock,
        mock_copy: MagicMock,
    ) -> None:
        from django.core.management import call_command

        self._configure_happy_path(
            mock_build,
            mock_detect,
            mock_layout,
            mock_graph_data,
            mock_main_comp,
            mock_measures,
            mock_pagerank,
            mock_communities_payload,
        )
        call_command(
            "structural_analysis",
            html=True,
            xlsx=True,
            community_strategies=self.strategy_token,
            edge_weight_strategy="PARTIAL_REFERENCES",
            stdout=io.StringIO(),
            stderr=io.StringIO(),
        )
        mock_table_html.assert_called_once()
        mock_table_xls.assert_called_once()


# ---------------------------------------------------------------------------
# measures/_centrality.py — apply_hits
# ---------------------------------------------------------------------------


class ApplyHitsTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_nodes_from(["1", "2", "3"])
        self.graph.add_edges_from([("1", "2"), ("2", "3"), ("3", "1")])
        self.graph_data: dict = {"nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}], "edges": []}

    def test_adds_hub_and_authority_to_all_nodes(self) -> None:
        apply_hits(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("hits_hub", node)
            self.assertIn("hits_authority", node)

    def test_returns_two_labels(self) -> None:
        labels = apply_hits(self.graph_data, self.graph)
        keys = [k for k, _ in labels]
        self.assertIn("hits_hub", keys)
        self.assertIn("hits_authority", keys)

    def test_hub_and_authority_are_floats(self) -> None:
        apply_hits(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIsInstance(node["hits_hub"], float)
            self.assertIsInstance(node["hits_authority"], float)


# ---------------------------------------------------------------------------
# measures/_centrality.py — in/out degree centralities
# ---------------------------------------------------------------------------


class ApplyInDegreeCentralityTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("3", "2")])
        self.graph_data: dict = {"nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}], "edges": []}

    def test_adds_in_degree_centrality_key(self) -> None:
        apply_in_degree_centrality(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("in_degree_centrality", node)

    def test_sink_node_has_higher_in_degree_than_source(self) -> None:
        apply_in_degree_centrality(self.graph_data, self.graph)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertGreater(node_map["2"]["in_degree_centrality"], node_map["1"]["in_degree_centrality"])


class ApplyOutDegreeCentralityTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("1", "3")])
        self.graph_data: dict = {"nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}], "edges": []}

    def test_adds_out_degree_centrality_key(self) -> None:
        apply_out_degree_centrality(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("out_degree_centrality", node)

    def test_source_node_has_higher_out_degree_than_sinks(self) -> None:
        apply_out_degree_centrality(self.graph_data, self.graph)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertGreater(node_map["1"]["out_degree_centrality"], node_map["2"]["out_degree_centrality"])


# ---------------------------------------------------------------------------
# measures/_centrality.py — apply_module_role (Guimerà-Amaral within-module role)
# ---------------------------------------------------------------------------


class ApplyModuleRoleTests(TestCase):
    _ROLES = {
        "Ultra-peripheral",
        "Peripheral",
        "Connector",
        "Kinless",
        "Provincial hub",
        "Connector hub",
        "Kinless hub",
    }

    def _graph_with_communities(self):
        graph = nx.DiGraph()
        graph.add_edges_from([("a", "b"), ("b", "c"), ("c", "a"), ("a", "d")])
        for n in graph.nodes():
            comm = "2-ld" if n == "d" else "1-ld"
            graph.nodes[n]["data"] = {"communities": {"leiden_directed": comm}}
        graph_data: dict = {"nodes": [{"id": n} for n in graph.nodes()], "edges": []}
        return graph, graph_data

    def test_emits_numeric_z_and_participation_and_role_label(self) -> None:
        graph, graph_data = self._graph_with_communities()
        labels = apply_module_role(graph_data, graph, "leiden_directed")
        self.assertEqual(
            labels,
            [("within_module_z", "Within-module z"), ("participation", "Participation Coefficient")],
        )
        for node in graph_data["nodes"]:
            self.assertIn("within_module_z", node)
            self.assertGreaterEqual(node["participation"], 0.0)
            self.assertLessEqual(node["participation"], 1.0)
            self.assertIn(node["module_role"], self._ROLES)
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        # "a" bridges the two communities (ties to b, c in its own and d outside);
        # "b" and "c" stay entirely inside theirs.
        self.assertGreater(node_map["a"]["participation"], 0.0)
        self.assertEqual(node_map["b"]["participation"], 0.0)

    def test_node_without_community_gets_none(self) -> None:
        graph, graph_data = self._graph_with_communities()
        graph.add_node("loner")
        graph.nodes["loner"]["data"] = {"communities": {}}
        graph_data["nodes"].append({"id": "loner"})
        apply_module_role(graph_data, graph, "leiden_directed")
        node_map = {n["id"]: n for n in graph_data["nodes"]}
        self.assertIsNone(node_map["loner"]["within_module_z"])
        self.assertIsNone(node_map["loner"]["participation"])
        self.assertIsNone(node_map["loner"]["module_role"])


# ---------------------------------------------------------------------------
# measures/_centrality.py — apply_burt_constraint
# ---------------------------------------------------------------------------


class ApplyBurtConstraintTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("2", "3"), ("1", "3")])
        self.graph.add_node("isolated")
        self.graph_data: dict = {
            "nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}, {"id": "isolated"}],
            "edges": [],
        }

    def test_adds_burt_constraint_key(self) -> None:
        apply_burt_constraint(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("burt_constraint", node)

    def test_isolated_node_gets_none(self) -> None:
        apply_burt_constraint(self.graph_data, self.graph)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertIsNone(node_map["isolated"]["burt_constraint"])

    def test_connected_node_gets_numeric_value(self) -> None:
        apply_burt_constraint(self.graph_data, self.graph)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertIsNotNone(node_map["1"]["burt_constraint"])
        self.assertIsInstance(node_map["1"]["burt_constraint"], float)


# ---------------------------------------------------------------------------
# measures/_centrality.py — apply_local_clustering
# ---------------------------------------------------------------------------


class ApplyLocalClusteringTests(TestCase):
    def setUp(self) -> None:
        # A directed 3-cycle: every node participates in exactly one directed triangle.
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("2", "3"), ("3", "1")])
        self.graph_data: dict = {"nodes": [{"id": "1"}, {"id": "2"}, {"id": "3"}], "edges": []}

    def test_adds_local_clustering_key(self) -> None:
        apply_local_clustering(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertIn("local_clustering", node)

    def test_values_in_unit_interval(self) -> None:
        apply_local_clustering(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertGreaterEqual(node["local_clustering"], 0.0)
            self.assertLessEqual(node["local_clustering"], 1.0)

    def test_cycle_nodes_have_nonzero_clustering(self) -> None:
        # Every node in a 3-cycle is part of a directed triangle.
        apply_local_clustering(self.graph_data, self.graph)
        for node in self.graph_data["nodes"]:
            self.assertGreater(node["local_clustering"], 0.0)

    def test_isolated_node_gets_zero(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("isolated")
        graph_data: dict = {"nodes": [{"id": "isolated"}], "edges": []}
        apply_local_clustering(graph_data, graph)
        self.assertEqual(graph_data["nodes"][0]["local_clustering"], 0.0)

    def test_returns_label(self) -> None:
        labels = apply_local_clustering(self.graph_data, self.graph)
        keys = [k for k, _ in labels]
        self.assertIn("local_clustering", keys)


# ---------------------------------------------------------------------------
# measures/_centrality.py — apply_reciprocity
# ---------------------------------------------------------------------------


class ApplyReciprocityTests(TestCase):
    def _apply(self, edges, nodes=None):
        graph = nx.DiGraph()
        if nodes:
            graph.add_nodes_from(nodes)
        graph.add_edges_from(edges)
        graph_data: dict = {"nodes": [{"id": n} for n in graph.nodes()], "edges": []}
        labels = apply_reciprocity(graph_data, graph)
        return {n["id"]: n["reciprocity"] for n in graph_data["nodes"]}, labels

    def test_mutual_pair_scores_one(self) -> None:
        values, labels = self._apply([("a", "b"), ("b", "a")])
        self.assertEqual(labels, [("reciprocity", "Reciprocity")])
        self.assertAlmostEqual(values["a"], 1.0)
        self.assertAlmostEqual(values["b"], 1.0)

    def test_one_way_tie_scores_zero(self) -> None:
        values, _ = self._apply([("a", "b")])
        self.assertEqual(values["a"], 0.0)
        self.assertEqual(values["b"], 0.0)

    def test_mixed_partners(self) -> None:
        # a↔b mutual, a→c one-way: r(a) = 2·1 / (1 + 2) = 0.6667
        values, _ = self._apply([("a", "b"), ("b", "a"), ("a", "c")])
        self.assertAlmostEqual(values["a"], 0.6667)
        self.assertAlmostEqual(values["b"], 1.0)
        self.assertEqual(values["c"], 0.0)

    def test_isolated_node_gets_none(self) -> None:
        values, _ = self._apply([("a", "b"), ("b", "a")], nodes=["loner"])
        self.assertIsNone(values["loner"])

    def test_self_loop_is_excluded(self) -> None:
        # A self-loop must count neither as a partner nor as a reciprocated tie.
        values, _ = self._apply([("a", "a"), ("a", "b")])
        self.assertEqual(values["a"], 0.0)

    def test_in_edges_only_scores_zero(self) -> None:
        # The dead-leaf shape: cited by others, never citing (out-edges out of scope).
        values, _ = self._apply([("a", "leaf"), ("b", "leaf")])
        self.assertEqual(values["leaf"], 0.0)


# ---------------------------------------------------------------------------
# measures/_content.py — apply_amplification_factor
# ---------------------------------------------------------------------------


class ApplyAmplificationFactorTests(TestCase):
    def setUp(self) -> None:
        label = make_label("Org", color="#FF0000")
        self.ch1 = make_channel(telegram_id=10, label=label, title="Source")
        self.ch2 = make_channel(telegram_id=11, label=label, title="Amplifier")
        # ch1 has 4 own messages; ch2 forwards 2 of ch1's messages into ch2's channel
        # → ch1's content is "amplified" 2 times; ch1 has 4 messages → factor = 2/4 = 0.5
        Message.objects.create(telegram_id=1, channel=self.ch1)
        Message.objects.create(telegram_id=2, channel=self.ch1)
        Message.objects.create(telegram_id=3, channel=self.ch1)
        Message.objects.create(telegram_id=4, channel=self.ch1)
        Message.objects.create(telegram_id=5, channel=self.ch2, forwarded_from=self.ch1)
        Message.objects.create(telegram_id=6, channel=self.ch2, forwarded_from=self.ch1)
        self.channel_dict = {
            str(self.ch1.pk): {"channel": self.ch1},
            str(self.ch2.pk): {"channel": self.ch2},
        }
        self.graph_data: dict = {
            "nodes": [{"id": str(self.ch1.pk)}, {"id": str(self.ch2.pk)}],
            "edges": [],
        }
        self.graph = nx.DiGraph()

    def test_amplification_factor_computed_correctly(self) -> None:
        apply_amplification_factor(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        # ch1 has 4 messages and is forwarded 2 times → 2/4 = 0.5
        self.assertAlmostEqual(node_map[str(self.ch1.pk)]["amplification_factor"], 0.5)
        # ch2 is never forwarded from → 0.0
        self.assertEqual(node_map[str(self.ch2.pk)]["amplification_factor"], 0.0)

    def test_channel_with_no_messages_gets_zero(self) -> None:
        apply_amplification_factor(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertEqual(node_map[str(self.ch2.pk)]["amplification_factor"], 0.0)


# ---------------------------------------------------------------------------
# measures/_content.py — apply_content_originality
# ---------------------------------------------------------------------------


class ApplyContentOriginalityTests(TestCase):
    def setUp(self) -> None:
        label = make_label("Org2", color="#00FF00")
        self.ch1 = make_channel(telegram_id=20, label=label, title="Original")
        self.ch2 = make_channel(telegram_id=21, label=label, title="Forwarder")
        # ch1: 4 messages, 0 forwarded → originality 1.0
        for i in range(4):
            Message.objects.create(telegram_id=100 + i, channel=self.ch1)
        # ch2: 4 messages, 2 forwarded → originality 0.5
        Message.objects.create(telegram_id=200, channel=self.ch2, forwarded_from=self.ch1)
        Message.objects.create(telegram_id=201, channel=self.ch2, forwarded_from=self.ch1)
        Message.objects.create(telegram_id=202, channel=self.ch2)
        Message.objects.create(telegram_id=203, channel=self.ch2)
        graph = nx.DiGraph()
        self.channel_dict = {
            str(self.ch1.pk): {"channel": self.ch1},
            str(self.ch2.pk): {"channel": self.ch2},
        }
        self.graph_data: dict = {
            "nodes": [{"id": str(self.ch1.pk)}, {"id": str(self.ch2.pk)}],
            "edges": [],
        }
        self.graph = graph

    def test_all_original_messages_scores_one(self) -> None:
        apply_content_originality(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertAlmostEqual(node_map[str(self.ch1.pk)]["content_originality"], 1.0)

    def test_half_forwarded_scores_half(self) -> None:
        apply_content_originality(self.graph_data, self.graph, self.channel_dict)
        node_map = {n["id"]: n for n in self.graph_data["nodes"]}
        self.assertAlmostEqual(node_map[str(self.ch2.pk)]["content_originality"], 0.5)

    def test_channel_with_no_messages_gets_none(self) -> None:
        label = make_label("OrgEmpty", color="#0000FF")
        empty_ch = make_channel(telegram_id=99, label=label, title="Empty")
        channel_dict = {str(empty_ch.pk): {"channel": empty_ch}}
        graph_data: dict = {"nodes": [{"id": str(empty_ch.pk)}], "edges": []}
        apply_content_originality(graph_data, nx.DiGraph(), channel_dict)
        self.assertIsNone(graph_data["nodes"][0]["content_originality"])


# ---------------------------------------------------------------------------
# coordination.py — compute_coordination / build_nx_graph
# ---------------------------------------------------------------------------


class ComputeCoordinationTests(TestCase):
    WINDOW = 300

    def setUp(self) -> None:
        label = make_label("CoordOrg", color="#112233")
        self.origin = make_channel(telegram_id=900, label=label, title="Origin")
        self.a = make_channel(telegram_id=901, label=label, title="A")
        self.b = make_channel(telegram_id=902, label=label, title="B")
        self.c = make_channel(telegram_id=903, label=label, title="C")
        self.ids = [self.a.pk, self.b.pk, self.c.pk]
        self.t0 = datetime.datetime(2024, 5, 1, 12, 0, tzinfo=datetime.timezone.utc)
        self._tg = 5000

    def _fwd(self, channel, origin_post, seconds, *, origin=None, fwd_date=None):
        """One forward of ``origin_post`` (origin-message id) at t0 + ``seconds``."""
        self._tg += 1
        return Message.objects.create(
            telegram_id=self._tg,
            channel=channel,
            forwarded_from=origin or self.origin,
            fwd_from_channel_post=origin_post,
            fwd_from_date=fwd_date or (self.t0 - datetime.timedelta(days=1)),
            date=self.t0 + datetime.timedelta(seconds=seconds),
        )

    def _compute(self, min_events: int = 3, ids=None):
        return compute_coordination(
            ids if ids is not None else self.ids, window_seconds=self.WINDOW, min_events=min_events
        )

    def test_repeated_within_window_pair_is_kept(self) -> None:
        for post in (1, 2, 3):
            self._fwd(self.a, post, 0)
            self._fwd(self.b, post, 100)
        result = self._compute(min_events=3)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 3)])
        for node_id in (str(self.a.pk), str(self.b.pk)):
            self.assertEqual(result.node_scores[node_id]["partners"], 1)
            self.assertEqual(result.node_scores[node_id]["strength"], 3)
            self.assertAlmostEqual(result.node_scores[node_id]["ratio"], 1.0)
        self.assertEqual(result.channels_seen, 2)
        self.assertEqual(result.origins_seen, 3)

    def test_forwards_outside_window_do_not_pair(self) -> None:
        self._fwd(self.a, 1, 0)
        self._fwd(self.b, 1, self.WINDOW + 1)
        result = self._compute(min_events=1)
        self.assertEqual(result.edges, [])
        self.assertEqual(result.node_ids, [])

    def test_exactly_window_apart_still_counts(self) -> None:
        self._fwd(self.a, 1, 0)
        self._fwd(self.b, 1, self.WINDOW)
        result = self._compute(min_events=1)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 1)])

    def test_min_events_filters_incidental_pairs(self) -> None:
        for post in (1, 2):
            self._fwd(self.a, post, 0)
            self._fwd(self.b, post, 50)
        result = self._compute(min_events=3)
        self.assertEqual(result.edges, [])

    def test_repeat_forwards_of_same_origin_count_once(self) -> None:
        self._fwd(self.a, 1, 0)
        self._fwd(self.a, 1, 40)  # A re-shares the same origin — not an extra event
        self._fwd(self.b, 1, 100)
        result = self._compute(min_events=1)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 1)])

    def test_self_forwards_are_ignored(self) -> None:
        self._fwd(self.origin, 1, 0, origin=self.origin)  # origin re-broadcasting itself
        self._fwd(self.a, 1, 10)
        result = self._compute(min_events=1, ids=[self.origin.pk, *self.ids])
        self.assertEqual(result.edges, [])

    def test_fallback_origin_key_uses_fwd_date(self) -> None:
        shared_origin_date = self.t0 - datetime.timedelta(days=2)
        self._fwd(self.a, None, 0, fwd_date=shared_origin_date)
        self._fwd(self.b, None, 60, fwd_date=shared_origin_date)
        # Same channels, no post id and a *different* origin date — a different origin.
        self._fwd(self.a, None, 200, fwd_date=self.t0 - datetime.timedelta(days=3))
        result = self._compute(min_events=1)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 1)])

    def test_unidentifiable_origin_rows_are_skipped(self) -> None:
        for channel, seconds in ((self.a, 0), (self.b, 30)):
            self._tg += 1
            Message.objects.create(
                telegram_id=self._tg,
                channel=channel,
                forwarded_from=self.origin,
                fwd_from_channel_post=None,
                fwd_from_date=None,
                date=self.t0 + datetime.timedelta(seconds=seconds),
            )
        result = self._compute(min_events=1)
        self.assertEqual(result.edges, [])
        self.assertEqual(result.origins_seen, 0)

    def test_ratio_is_coordinated_share_of_forwarded_origins(self) -> None:
        # A forwards 4 origins; B co-forwards 2 of them inside the window.
        for post in (1, 2, 3, 4):
            self._fwd(self.a, post, 0)
        self._fwd(self.b, 1, 50)
        self._fwd(self.b, 2, 80)
        result = self._compute(min_events=2)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 2)])
        self.assertAlmostEqual(result.node_scores[str(self.a.pk)]["ratio"], 0.5)
        self.assertAlmostEqual(result.node_scores[str(self.b.pk)]["ratio"], 1.0)

    def test_filtered_partner_does_not_appear(self) -> None:
        for post in (1, 2, 3):
            self._fwd(self.a, post, 0)
            self._fwd(self.b, post, 100)
        self._fwd(self.c, 1, 150)  # C syncs on a single origin only
        result = self._compute(min_events=2)
        self.assertEqual(result.edges, [(str(self.a.pk), str(self.b.pk), 3)])
        self.assertNotIn(str(self.c.pk), result.node_ids)
        # A's coordinated origins are counted over retained pairs only.
        self.assertEqual(result.node_scores[str(self.a.pk)]["partners"], 1)

    def test_build_nx_graph_is_bidirectional_with_shared_node_data(self) -> None:
        for post in (1, 2, 3):
            self._fwd(self.a, post, 0)
            self._fwd(self.b, post, 100)
        result = self._compute(min_events=3)
        main_graph = nx.DiGraph()
        for channel in (self.a, self.b):
            main_graph.add_node(str(channel.pk), data={"pk": str(channel.pk), "color": "10,20,30"})
        co_graph = build_nx_graph(result, main_graph)
        a_id, b_id = str(self.a.pk), str(self.b.pk)
        self.assertTrue(co_graph.has_edge(a_id, b_id))
        self.assertTrue(co_graph.has_edge(b_id, a_id))
        self.assertEqual(co_graph.edges[a_id, b_id]["weight"], 3.0)
        self.assertEqual(co_graph.nodes[a_id]["data"], main_graph.nodes[a_id]["data"])


class BuildCoordinationGraphDataTests(TestCase):
    def setUp(self) -> None:
        label = make_label("CoordGD", color="#445566")
        self.a = make_channel(telegram_id=910, label=label, title="A")
        self.b = make_channel(telegram_id=911, label=label, title="B")
        self.origin = make_channel(telegram_id=912, label=label, title="Origin")
        t0 = datetime.datetime(2024, 6, 1, 9, 0, tzinfo=datetime.timezone.utc)
        for post, offset in ((1, 0), (2, 0), (3, 0)):
            for channel, extra in ((self.a, 0), (self.b, 30)):
                Message.objects.create(
                    telegram_id=6000 + post * 10 + extra,
                    channel=channel,
                    forwarded_from=self.origin,
                    fwd_from_channel_post=post,
                    fwd_from_date=t0 - datetime.timedelta(days=1),
                    date=t0 + datetime.timedelta(seconds=offset + extra),
                )
        self.result = compute_coordination([self.a.pk, self.b.pk], window_seconds=300, min_events=3)

    def test_nodes_copy_main_dicts_and_carry_scores(self) -> None:
        a_id, b_id = str(self.a.pk), str(self.b.pk)
        graph_data = {
            "nodes": [
                {"id": a_id, "label": "A", "color": "255,0,0", "fans": 5},
                {"id": b_id, "label": "B", "color": "0,0,255", "fans": 7},
            ],
            "edges": [],
        }
        positions = {a_id: (1.0, 2.0), b_id: (3.0, 4.0)}
        coord_data = build_coordination_graph_data(graph_data, self.result, positions)
        self.assertEqual(len(coord_data["nodes"]), 2)
        node_a = next(n for n in coord_data["nodes"] if n["id"] == a_id)
        self.assertEqual((node_a["x"], node_a["y"]), (1.0, 2.0))
        self.assertEqual(node_a["fans"], 5)  # copied from the main graph's node dict
        self.assertEqual(node_a["coordination_strength"], 3)
        self.assertEqual(node_a["coordination_partners"], 1)
        self.assertAlmostEqual(node_a["coordination_ratio"], 1.0)
        # One tie, materialised in both directions, with a dimmed averaged colour.
        self.assertEqual(len(coord_data["edges"]), 2)
        sources = {e["source"] for e in coord_data["edges"]}
        self.assertEqual(sources, {a_id, b_id})
        for edge in coord_data["edges"]:
            self.assertEqual(edge["weight"], 3.0)
            self.assertEqual(len(edge["color"].split(",")), 3)


class WriteCoordinationOutputsTests(TestCase):
    """Integration: the coordination data dir + both viewer pages, against the real map templates."""

    def setUp(self) -> None:
        self.coord_graph_data = {
            "nodes": [
                {
                    "id": "1",
                    "x": 0.0,
                    "y": 1.0,
                    "label": "A",
                    "color": "255,0,0",
                    "coordination_strength": 3,
                    "coordination_partners": 1,
                    "coordination_ratio": 1.0,
                },
                {
                    "id": "2",
                    "x": 2.0,
                    "y": 3.0,
                    "label": "B",
                    "color": "0,0,255",
                    "coordination_strength": 3,
                    "coordination_partners": 1,
                    "coordination_ratio": 0.5,
                },
            ],
            "edges": [
                {"source": "1", "target": "2", "weight": 3.0, "color": "95,0,95", "id": 0},
                {"source": "2", "target": "1", "weight": 3.0, "color": "95,0,95", "id": 1},
            ],
        }
        self.positions_3d = {"1": (0.0, 1.0, 2.0), "2": (3.0, 4.0, 5.0)}

    def test_data_files_written_and_parseable(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_files(
                self.coord_graph_data,
                self.positions_3d,
                [("coordination_strength", "Coordinated co-forwards")],
                tmp,
                communities_data={"leiden": {"groups": [], "main_groups": {}}},
            )
            data_dir = os.path.join(tmp, "data_coordination")
            with open(os.path.join(data_dir, "channel_position.json")) as f:
                positions = json.load(f)
            self.assertEqual(len(positions["nodes"]), 2)
            self.assertEqual(len(positions["edges"]), 2)
            with open(os.path.join(data_dir, "channel_position_3d.json")) as f:
                positions_3d = json.load(f)
            self.assertEqual(positions_3d["nodes"][0]["z"], 2.0)
            with open(os.path.join(data_dir, "channels.json")) as f:
                channels = json.load(f)
            self.assertEqual(channels["total_pages_count"], 2)
            self.assertEqual(channels["measures"][0][0], "coordination_strength")
            self.assertEqual(channels["nodes"][0]["coordination_partners"], 1)
            # The viewers unconditionally fetch communities.json — it must exist.
            with open(os.path.join(data_dir, "communities.json")) as f:
                communities = json.load(f)
            self.assertIn("leiden", communities["strategies"])

    def test_dir_name_selects_per_year_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_files(
                self.coord_graph_data,
                self.positions_3d,
                [("coordination_strength", "Coordinated co-forwards")],
                tmp,
                dir_name="data_coordination_2023",
            )
            year_dir = os.path.join(tmp, "data_coordination_2023")
            for name in ("channel_position.json", "channel_position_3d.json", "channels.json", "communities.json"):
                self.assertTrue(os.path.exists(os.path.join(year_dir, name)), name)
            # Absent communities_data still yields a parseable, empty payload.
            with open(os.path.join(year_dir, "communities.json")) as f:
                self.assertEqual(json.load(f), {"strategies": {}})

    def test_no_3d_positions_skips_the_3d_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_files(
                self.coord_graph_data,
                None,
                [("coordination_strength", "Coordinated co-forwards")],
                tmp,
            )
            data_dir = os.path.join(tmp, "data_coordination")
            self.assertTrue(os.path.exists(os.path.join(data_dir, "channel_position.json")))
            self.assertFalse(os.path.exists(os.path.join(data_dir, "channel_position_3d.json")))

    def test_timeline_json_lists_only_years_with_coordination(self) -> None:
        entries = [
            {"year": 2022, "has_coordination": True, "coordination_nodes": 4, "coordination_ties": 3},
            {"year": 2023, "has_coordination": False},
            {"year": 2024, "has_coordination": True, "coordination_nodes": 2, "coordination_ties": 1},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_timeline_json(entries, tmp)
            with open(os.path.join(tmp, "data_coordination", "timeline.json")) as f:
                timeline = json.load(f)
        self.assertEqual([y["year"] for y in timeline["years"]], [2022, 2024])
        self.assertTrue(all(y["has_graph"] for y in timeline["years"]))
        self.assertEqual(timeline["years"][0]["edges"], 3)

    def test_pages_written_from_real_templates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_pages(
                tmp,
                seo=False,
                project_title="Test Project",
                node_count=2,
                tie_count=1,
                strategy_labels={"labelgroup1": "Org"},
            )
            for page in ("coordination.html", "coordination3d.html"):
                path = os.path.join(tmp, page)
                self.assertTrue(os.path.exists(path), page)
                with open(path) as f:
                    content = f.read()
                # Re-pointed at the coordination data dir, with the shims injected.
                self.assertIn('window.DATA_DIR = "data_coordination/"', content)
                self.assertIn("window.EXTRA_LAYOUTS = []", content)
                self.assertIn('window.STRATEGY_LABELS = {"labelgroup1": "Org"}', content)
                # Placeholders resolved and reworded for mutual ties.
                self.assertNotIn("__NODE_COUNT__", content)
                self.assertNotIn("__EDGE_COUNT__", content)
                self.assertIn("mutual co-forwarding ties", content)
                self.assertNotIn("directed edges", content)
                # Titled after the project, still noindex without --seo.
                self.assertIn("<title>Test Project — Coordination</title>", content)
                self.assertIn('content="noindex"', content)

    def test_include_flags_gate_the_pages(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_pages(
                tmp,
                seo=False,
                project_title="Test Project",
                node_count=2,
                tie_count=1,
                strategy_labels={},
                include_2d=True,
                include_3d=False,
            )
            self.assertTrue(os.path.exists(os.path.join(tmp, "coordination.html")))
            self.assertFalse(os.path.exists(os.path.join(tmp, "coordination3d.html")))
        with tempfile.TemporaryDirectory() as tmp:
            write_coordination_pages(
                tmp,
                seo=False,
                project_title="Test Project",
                node_count=2,
                tie_count=1,
                strategy_labels={},
                include_2d=False,
                include_3d=True,
            )
            self.assertFalse(os.path.exists(os.path.join(tmp, "coordination.html")))
            self.assertTrue(os.path.exists(os.path.join(tmp, "coordination3d.html")))


# ---------------------------------------------------------------------------
# community_stats.py — _freeman_centralization
# ---------------------------------------------------------------------------


class FreemanCentralizationTests(TestCase):
    def test_returns_none_for_empty_list(self) -> None:
        self.assertIsNone(_freeman_centralization([]))

    def test_returns_none_for_single_value(self) -> None:
        self.assertIsNone(_freeman_centralization([0.5]))

    def test_returns_none_when_all_zero(self) -> None:
        self.assertIsNone(_freeman_centralization([0.0, 0.0, 0.0]))

    def test_returns_zero_for_equal_values(self) -> None:
        result = _freeman_centralization([0.5, 0.5, 0.5])
        self.assertAlmostEqual(result, 0.0)

    def test_returns_one_for_star_pattern(self) -> None:
        # One node with max centrality, all others at zero → centralization = 1.0
        result = _freeman_centralization([1.0, 0.0, 0.0])
        self.assertAlmostEqual(result, 1.0)

    def test_none_entries_ignored(self) -> None:
        # None entries should be filtered; result same as [1.0, 0.0, 0.0]
        result = _freeman_centralization([1.0, None, 0.0, None, 0.0])
        self.assertAlmostEqual(result, 1.0)


# ---------------------------------------------------------------------------
# community_stats.py — _network_summary
# ---------------------------------------------------------------------------


class NetworkSummaryTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("a", "b"), ("b", "c"), ("c", "a"), ("a", "c")])

    def test_returns_expected_keys(self) -> None:
        summary = _network_summary(self.graph)
        for key in ("n", "e", "density", "reciprocity", "avg_clustering", "wcc_count", "scc_count"):
            self.assertIn(key, summary)

    def test_node_and_edge_counts_correct(self) -> None:
        summary = _network_summary(self.graph)
        self.assertEqual(summary["n"], 3)
        self.assertEqual(summary["e"], 4)

    def test_density_in_unit_interval(self) -> None:
        summary = _network_summary(self.graph)
        self.assertGreaterEqual(summary["density"], 0.0)
        self.assertLessEqual(summary["density"], 1.0)

    def test_empty_graph_does_not_raise(self) -> None:
        summary = _network_summary(nx.DiGraph())
        self.assertEqual(summary["n"], 0)
        self.assertEqual(summary["e"], 0)


# ---------------------------------------------------------------------------
# community_stats.py — network_summary_rows
# ---------------------------------------------------------------------------


class NetworkSummaryRowsTests(TestCase):
    def _make_summary(self) -> dict:
        return {
            "n": 5,
            "e": 8,
            "density": 0.4,
            "reciprocity": 0.25,
            "avg_clustering": 0.3,
            "avg_path_length": 2.1,
            "diameter": 4,
            "path_on_full": True,
            "wcc_count": 1,
            "wcc_fraction": 1.0,
            "scc_count": 2,
            "scc_fraction": 0.6,
            "assortativity": {"in_in": None, "in_out": None, "out_in": None, "out_out": None},
            "centralizations": {},
        }

    def test_returns_list_of_three_tuples(self) -> None:
        rows = network_summary_rows(self._make_summary())
        self.assertIsInstance(rows, list)
        for row in rows:
            self.assertIsInstance(row, tuple)
            self.assertEqual(len(row), 3)

    def test_nodes_row_present(self) -> None:
        rows = network_summary_rows(self._make_summary())
        labels = [r[0] for r in rows]
        self.assertIn("Nodes", labels)

    def test_edges_row_value_matches_summary(self) -> None:
        rows = network_summary_rows(self._make_summary())
        row_map = {r[0]: r[1] for r in rows}
        self.assertEqual(row_map["Nodes"], 5)
        self.assertEqual(row_map["Edges"], 8)


# ---------------------------------------------------------------------------
# community_stats.py — compute_community_metrics
# ---------------------------------------------------------------------------


class ComputeCommunityMetricsTests(TestCase):
    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("2", "3"), ("3", "1")])
        for node_id in ["1", "2", "3"]:
            self.graph.nodes[node_id]["data"] = {"communities": {"leiden": f"comm_{node_id}"}}
        self.graph_data: dict = {
            "nodes": [
                {"id": "1", "communities": {"leiden": "comm_1"}},
                {"id": "2", "communities": {"leiden": "comm_2"}},
                {"id": "3", "communities": {"leiden": "comm_3"}},
            ],
            "edges": [],
        }
        self.communities_data = {
            "leiden": {
                "groups": [
                    (1, 1, "comm_1", "#ff0000"),
                    (2, 1, "comm_2", "#00ff00"),
                    (3, 1, "comm_3", "#0000ff"),
                ]
            }
        }

    def test_returns_network_summary_and_strategies_keys(self) -> None:
        result = compute_community_metrics(self.graph_data, self.communities_data, self.graph, ["leiden"])
        self.assertIn("network_summary", result)
        self.assertIn("strategies", result)

    def test_network_summary_has_correct_node_count(self) -> None:
        result = compute_community_metrics(self.graph_data, self.communities_data, self.graph, ["leiden"])
        self.assertEqual(result["network_summary"]["n"], 3)

    def test_strategy_entry_present_for_requested_strategy(self) -> None:
        result = compute_community_metrics(self.graph_data, self.communities_data, self.graph, ["leiden"])
        self.assertIn("leiden", result["strategies"])

    def test_status_callback_called_for_each_step(self) -> None:
        calls: list[str] = []
        compute_community_metrics(
            self.graph_data,
            self.communities_data,
            self.graph,
            ["leiden"],
            status_callback=calls.append,
        )
        # callback called once for "network" and once per strategy
        self.assertGreaterEqual(len(calls), 2)


class ComparePartitionsTests(TestCase):
    """The four-index partition-comparison helper (ARI, AMI, NMI, VI)."""

    def test_returns_all_four_indices(self) -> None:
        scores = _compare_partitions([0, 0, 1, 1], [0, 0, 1, 1])
        self.assertEqual(set(scores), {key for key, *_ in PARTITION_COMPARISON_METRICS})

    def test_identical_partitions_are_perfect(self) -> None:
        # Relabeled-identical: the indices are invariant to how communities are named.
        scores = _compare_partitions([0, 0, 1, 1, 2, 2], [7, 7, 3, 3, 9, 9])
        self.assertEqual(scores, {"ari": 1.0, "ami": 1.0, "nmi": 1.0, "vi": 0.0})

    def test_independent_partitions_score_low(self) -> None:
        scores = _compare_partitions([0, 0, 1, 1, 2, 2], [0, 1, 2, 0, 1, 2])
        self.assertLessEqual(scores["ari"], 0.0)  # chance-corrected: ≤ 0 when independent
        self.assertLessEqual(scores["ami"], 0.0)
        self.assertGreater(scores["vi"], 0.0)  # VI is a positive distance

    def test_empty_input_returns_none(self) -> None:
        self.assertIsNone(_compare_partitions([], []))


class PartitionComparisonMatrixTests(TestCase):
    """compute_community_metrics builds the strategy×strategy comparison matrices."""

    def setUp(self) -> None:
        self.graph = nx.DiGraph()
        self.graph.add_edges_from([("1", "2"), ("3", "4"), ("5", "6"), ("2", "3"), ("4", "5")])
        # leiden and leiden_directed are the same partition under different labels; kcore is a
        # shell decomposition that must be excluded from the comparison matrices.
        self.graph_data = {
            "nodes": [
                {"id": "1", "communities": {"leiden": "A", "leiden_directed": "X", "kcore": "0"}},
                {"id": "2", "communities": {"leiden": "A", "leiden_directed": "X", "kcore": "0"}},
                {"id": "3", "communities": {"leiden": "B", "leiden_directed": "Y", "kcore": "1"}},
                {"id": "4", "communities": {"leiden": "B", "leiden_directed": "Y", "kcore": "1"}},
                {"id": "5", "communities": {"leiden": "C", "leiden_directed": "Z", "kcore": "2"}},
                {"id": "6", "communities": {"leiden": "C", "leiden_directed": "Z", "kcore": "2"}},
            ],
            "edges": [],
        }
        self.communities_data = {
            "leiden": {"groups": [(1, 2, "A", "#f00"), (2, 2, "B", "#0f0"), (3, 2, "C", "#00f")]},
            "leiden_directed": {"groups": [(1, 2, "X", "#f00"), (2, 2, "Y", "#0f0"), (3, 2, "Z", "#00f")]},
            "kcore": {"groups": [(1, 2, "0", "#f00"), (2, 2, "1", "#0f0"), (3, 2, "2", "#00f")]},
        }
        self.strategies = ["leiden", "leiden_directed", "kcore"]

    def _comparison(self) -> dict:
        result = compute_community_metrics(self.graph_data, self.communities_data, self.graph, self.strategies)
        self.assertIn("partition_comparison", result)
        return result["partition_comparison"]

    def test_kcore_excluded_from_strategies(self) -> None:
        comparison = self._comparison()
        self.assertEqual(comparison["strategies"], ["leiden", "leiden_directed"])

    def test_all_four_metric_matrices_present_and_square(self) -> None:
        comparison = self._comparison()
        k = len(comparison["strategies"])
        for key, *_ in PARTITION_COMPARISON_METRICS:
            cells = comparison["metrics"][key]
            self.assertEqual(len(cells), k)
            self.assertTrue(all(len(row) == k for row in cells))

    def test_diagonal_is_identity_per_metric(self) -> None:
        metrics = self._comparison()["metrics"]
        for key, _abbr, _name, is_distance in PARTITION_COMPARISON_METRICS:
            diag = 0.0 if is_distance else 1.0
            self.assertTrue(all(metrics[key][i][i] == diag for i in range(len(metrics[key]))))

    def test_matrices_symmetric(self) -> None:
        metrics = self._comparison()["metrics"]
        for key, *_ in PARTITION_COMPARISON_METRICS:
            cells = metrics[key]
            for i in range(len(cells)):
                for j in range(len(cells)):
                    self.assertEqual(cells[i][j], cells[j][i])

    def test_identical_partitions_agree_off_diagonal(self) -> None:
        # leiden vs leiden_directed are the same grouping → perfect agreement off the diagonal.
        metrics = self._comparison()["metrics"]
        self.assertEqual(metrics["ari"][0][1], 1.0)
        self.assertEqual(metrics["ami"][0][1], 1.0)
        self.assertEqual(metrics["nmi"][0][1], 1.0)
        self.assertEqual(metrics["vi"][0][1], 0.0)

    def test_absent_when_fewer_than_two_comparable_strategies(self) -> None:
        # Only kcore (excluded) plus one strategy leaves a single comparable partition.
        result = compute_community_metrics(self.graph_data, self.communities_data, self.graph, ["leiden", "kcore"])
        self.assertNotIn("partition_comparison", result)


class DisparityFilterTests(TestCase):
    def _star(self, n_leaves: int, weights: "list[float] | None" = None) -> nx.DiGraph:
        """Directed star ``A → L0, L1, …`` with the given outgoing weights."""
        g = nx.DiGraph()
        if weights is None:
            weights = [1.0] * n_leaves
        for i, w in enumerate(weights):
            g.add_edge("A", f"L{i}", weight=w)
        return g

    def test_single_outgoing_edge_kept_under_strict_threshold(self) -> None:
        # A → B (only outgoing), B has only incoming from A: k_out=k_in=1 → α=0 both sides
        g = nx.DiGraph()
        g.add_edge("A", "B", weight=42.0)
        from network.robustness import disparity_filter

        backbone = disparity_filter(g, alpha=1e-12)
        self.assertEqual(set(backbone.edges()), {("A", "B")})

    def test_uniform_distribution_is_filtered_out(self) -> None:
        # Source A fans out to 4 targets with equal weight; each target also has
        # 4 incoming edges of equal weight from background sources. With p = 1/4
        # and k = 4 the disparity α = (3/4)^3 = 0.4219… well above the 0.05
        # threshold from *both* sides, so A's outgoing edges should be removed.
        from network.robustness import disparity_filter

        g = nx.DiGraph()
        for tgt in ("B", "C", "D", "E"):
            g.add_edge("A", tgt, weight=1.0)
            for src in range(3):  # 3 extra background sources → target k_in = 4
                g.add_edge(f"S{tgt}{src}", tgt, weight=1.0)
        backbone = disparity_filter(g, alpha=0.05)
        for tgt in ("B", "C", "D", "E"):
            self.assertNotIn(("A", tgt), backbone.edges())

    def test_dominant_edge_survives_strict_threshold(self) -> None:
        # Hub A: one dominant edge to B (weight 100), nine background edges of weight 1.
        # B also has many incoming edges so α_in is not trivially 0.
        # α_out for (A, B) = (1 - 100/109)^9 = (9/109)^9 ≈ 2.4e-10 ≪ 0.05.
        from network.robustness import disparity_filter

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=100.0)
        for i in range(9):
            g.add_edge("A", f"X{i}", weight=1.0)
        for i in range(9):
            g.add_edge(f"Y{i}", "B", weight=1.0)
        backbone = disparity_filter(g, alpha=0.05)
        self.assertIn(("A", "B"), backbone.edges())
        # The nine background edges of A are not concentrated enough on either side
        # (each target has k_in = 1, though, so α_in = 0 from that side → kept).
        # So this test only checks the dominant edge — see next test for full filtering.

    def test_alpha_threshold_at_one_keeps_every_edge(self) -> None:
        # min(α_in, α_out) is always < 1 for any positive-weight edge, so α=1 → everything kept.
        from network.robustness import disparity_filter

        g = self._star(5, weights=[1.0, 2.0, 3.0, 4.0, 5.0])
        backbone = disparity_filter(g, alpha=1.0)
        self.assertEqual(g.number_of_edges(), backbone.number_of_edges())

    def test_invalid_alpha_raises(self) -> None:
        from network.robustness import disparity_filter

        g = self._star(2)
        with self.assertRaises(ValueError):
            disparity_filter(g, alpha=0.0)
        with self.assertRaises(ValueError):
            disparity_filter(g, alpha=1.5)
        with self.assertRaises(ValueError):
            disparity_filter(g, alpha=-0.1)

    def test_nodes_and_node_attributes_preserved(self) -> None:
        from network.robustness import disparity_filter

        g = nx.DiGraph()
        g.add_node("A", color="red", label="source")
        g.add_node("B", color="blue", label="target")
        g.add_edge("A", "B", weight=1.0)
        backbone = disparity_filter(g, alpha=0.5)
        self.assertEqual(set(backbone.nodes()), {"A", "B"})
        self.assertEqual(backbone.nodes["A"]["color"], "red")
        self.assertEqual(backbone.nodes["B"]["label"], "target")

    def test_edge_attributes_preserved_on_retained_edges(self) -> None:
        from network.robustness import disparity_filter

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0, color="green", tag="forward")
        backbone = disparity_filter(g, alpha=0.5)
        self.assertEqual(backbone.edges["A", "B"]["color"], "green")
        self.assertEqual(backbone.edges["A", "B"]["tag"], "forward")

    def test_isolated_nodes_remain_in_backbone(self) -> None:
        # Filtering may strip all of a node's edges; the node itself is kept.
        from network.robustness import disparity_filter

        g = nx.DiGraph()
        # uniform 4-fan with k_in = 4 on every target → α ≈ 0.42 > 0.05 from both sides
        for tgt in ("B", "C", "D", "E"):
            g.add_edge("A", tgt, weight=1.0)
            for src in range(3):
                g.add_edge(f"S{tgt}{src}", tgt, weight=1.0)
        backbone = disparity_filter(g, alpha=0.05)
        self.assertIn("A", backbone.nodes())  # A may be isolated but must be present

    def test_compute_alpha_values_returns_pair_per_edge_in_unit_interval(self) -> None:
        from network.robustness import compute_alpha_values

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=10.0)
        g.add_edge("A", "C", weight=1.0)
        g.add_edge("D", "B", weight=1.0)
        alphas = compute_alpha_values(g)
        self.assertEqual(set(alphas.keys()), {("A", "B"), ("A", "C"), ("D", "B")})
        for a_in, a_out in alphas.values():
            self.assertGreaterEqual(a_in, 0.0)
            self.assertLessEqual(a_in, 1.0)
            self.assertGreaterEqual(a_out, 0.0)
            self.assertLessEqual(a_out, 1.0)

    def test_compute_alpha_values_matches_serrano_formula(self) -> None:
        # Source A has k_out = 3, weights [1, 1, 1] → p = 1/3, α = (2/3)^2 ≈ 0.4444
        # Target B/C/D each has k_in = 1 → α_in = 0
        from network.robustness import compute_alpha_values

        g = nx.DiGraph()
        for tgt in ("B", "C", "D"):
            g.add_edge("A", tgt, weight=1.0)
        alphas = compute_alpha_values(g)
        expected_alpha_out = (2.0 / 3.0) ** 2
        for tgt in ("B", "C", "D"):
            a_in, a_out = alphas[("A", tgt)]
            self.assertAlmostEqual(a_out, expected_alpha_out, places=10)
            self.assertEqual(a_in, 0.0)


class AttackCurveTests(TestCase):
    def test_wcc_curve_on_directed_chain_matches_closed_form(self) -> None:
        # Chain A → B → C → D, removed in order. WCC is computed undirected, so
        # the residual chain shrinks one node at a time: S(q) = (N - q) / N.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C"), ("C", "D")])
        curve = attack_curve(g, ["A", "B", "C", "D"], "WCC")
        self.assertEqual(curve, [1.0, 0.75, 0.5, 0.25, 0.0])

    def test_wcc_curve_on_clique_decreases_linearly(self) -> None:
        # In a directed clique every removal leaves a smaller clique, fully connected.
        from network.robustness import attack_curve

        g = nx.complete_graph(5, create_using=nx.DiGraph)
        curve = attack_curve(g, list(g.nodes()), "WCC")
        for q, s in enumerate(curve):
            self.assertAlmostEqual(s, (5 - q) / 5)

    def test_wcc_curve_on_star_center_first_collapses_immediately(self) -> None:
        # Removing the centre of a star isolates every leaf → S(1) = 1/N.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("C", "L1"), ("C", "L2"), ("C", "L3"), ("C", "L4")])
        curve = attack_curve(g, ["C", "L1", "L2", "L3", "L4"], "WCC")
        # S(0)=1, S(1..4)=1/5 (each leaf is its own singleton WCC), S(5)=0
        self.assertEqual(curve, [1.0, 0.2, 0.2, 0.2, 0.2, 0.0])

    def test_wcc_curve_on_star_leaves_first_decreases_smoothly(self) -> None:
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("C", "L1"), ("C", "L2"), ("C", "L3"), ("C", "L4")])
        curve = attack_curve(g, ["L1", "L2", "L3", "L4", "C"], "WCC")
        self.assertEqual(curve, [1.0, 0.8, 0.6, 0.4, 0.2, 0.0])

    def test_scc_curve_on_directed_cycle(self) -> None:
        # A → B → C → A: one SCC of size 3, then only singletons after any removal.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C"), ("C", "A")])
        curve = attack_curve(g, ["A", "B", "C"], "SCC")
        self.assertAlmostEqual(curve[0], 1.0)
        self.assertAlmostEqual(curve[1], 1.0 / 3.0)
        self.assertAlmostEqual(curve[2], 1.0 / 3.0)
        self.assertAlmostEqual(curve[3], 0.0)

    def test_reach_curve_on_directed_chain_matches_closed_form(self) -> None:
        # A → B → C → D: reachable pairs = AB, AC, AD, BC, BD, CD = 6.
        # Total ordered pairs = 4*3 = 12. S(0) = 6/12 = 0.5.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C"), ("C", "D")])
        curve = attack_curve(g, ["A", "B", "C", "D"], "REACH", reach_sample=None)
        self.assertAlmostEqual(curve[0], 0.5)
        # Removing A drops 3 pairs (A→B, A→C, A→D) → 3/12 = 0.25
        self.assertAlmostEqual(curve[1], 0.25)

    def test_reach_curve_with_sampling_stays_in_unit_interval(self) -> None:
        from network.robustness import attack_curve

        g = nx.gnp_random_graph(50, 0.1, seed=42, directed=True)
        rng = np.random.default_rng(7)
        curve = attack_curve(g, list(g.nodes()), "REACH", reach_sample=10, rng=rng)
        for s in curve:
            self.assertGreaterEqual(s, 0.0)
            self.assertLessEqual(s, 1.0)

    def test_reach_sampling_reproducible_with_seed(self) -> None:
        from network.robustness import attack_curve

        g = nx.gnp_random_graph(40, 0.15, seed=42, directed=True)
        order = list(g.nodes())
        c1 = attack_curve(g, order, "REACH", reach_sample=8, rng=np.random.default_rng(0))
        c2 = attack_curve(g, order, "REACH", reach_sample=8, rng=np.random.default_rng(0))
        self.assertEqual(c1, c2)

    def test_strength_curve_tracks_heaviest_component_weight(self) -> None:
        # Two components: heavy pair (weight 10) and light chain (total 2).
        # S_strength(0) = 10/12; removing "H1" leaves the light chain as the
        # heaviest component → 2/12; removing "L1" splits it → 1/12.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edge("H1", "H2", weight=10.0)
        g.add_edge("L1", "L2", weight=1.0)
        g.add_edge("L2", "L3", weight=1.0)
        curve = attack_curve(g, ["H1", "L1"], "STRENGTH")
        self.assertAlmostEqual(curve[0], 10.0 / 12.0)
        self.assertAlmostEqual(curve[1], 2.0 / 12.0)
        self.assertAlmostEqual(curve[2], 1.0 / 12.0)

    def test_strength_curve_sees_damage_wcc_misses(self) -> None:
        # A hub carrying nearly all the weight inside one big WCC: removing it
        # barely dents the node count but guts the surviving strength.
        from network.robustness import attack_curve

        g = nx.DiGraph()
        for i in range(5):
            g.add_edge(f"S{i}", "HUB", weight=10.0)
            g.add_edge(f"S{i}", f"S{(i + 1) % 5}", weight=0.1)
        wcc = attack_curve(g, ["HUB"], "WCC")
        strength = attack_curve(g, ["HUB"], "STRENGTH")
        self.assertAlmostEqual(wcc[1], 5.0 / 6.0)  # 5 of 6 nodes still connected
        self.assertAlmostEqual(strength[1], 0.5 / 50.5)  # 0.99% of the weight survives

    def test_invalid_metric_raises(self) -> None:
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edge("A", "B")
        with self.assertRaises(ValueError):
            attack_curve(g, ["A"], "INVALID")  # type: ignore[arg-type]

    def test_empty_graph_returns_single_zero(self) -> None:
        from network.robustness import attack_curve

        self.assertEqual(attack_curve(nx.DiGraph(), [], "WCC"), [0.0])

    def test_removal_order_with_missing_nodes_is_silently_skipped(self) -> None:
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edge("A", "B")
        curve = attack_curve(g, ["A", "X", "B"], "WCC")
        # S(0)=1 (one WCC of 2 over 2), S(1)=0.5 after A, X is missing → 0.5, S(3)=0
        self.assertEqual(curve, [1.0, 0.5, 0.5, 0.0])

    def test_does_not_mutate_input_graph(self) -> None:
        from network.robustness import attack_curve

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C")])
        before = (set(g.nodes()), set(g.edges()))
        attack_curve(g, ["A", "B"], "WCC")
        after = (set(g.nodes()), set(g.edges()))
        self.assertEqual(before, after)


class RIndexAndThresholdTests(TestCase):
    def test_r_index_on_chain_matches_closed_form(self) -> None:
        # Closed form for a chain removed sequentially: R = (N-1) / (2N).
        from network.robustness import attack_curve, r_index

        g = nx.DiGraph()
        g.add_edges_from([(str(i), str(i + 1)) for i in range(3)])
        curve = attack_curve(g, ["0", "1", "2", "3"], "WCC")
        self.assertAlmostEqual(r_index(curve), 3.0 / 8.0)

    def test_r_index_on_clique_matches_closed_form(self) -> None:
        from network.robustness import attack_curve, r_index

        g = nx.complete_graph(5, create_using=nx.DiGraph)
        curve = attack_curve(g, list(g.nodes()), "WCC")
        self.assertAlmostEqual(r_index(curve), (5 - 1) / (2 * 5))

    def test_r_index_star_center_first_lower_than_leaves_first(self) -> None:
        # Targeted hub removal must give a strictly lower R than peripheral.
        from network.robustness import attack_curve, r_index

        g = nx.DiGraph()
        g.add_edges_from([("C", "L1"), ("C", "L2"), ("C", "L3"), ("C", "L4")])
        r_center = r_index(attack_curve(g, ["C", "L1", "L2", "L3", "L4"], "WCC"))
        r_leaves = r_index(attack_curve(g, ["L1", "L2", "L3", "L4", "C"], "WCC"))
        self.assertLess(r_center, r_leaves)
        # And matches the analytical values exactly: 0.16 vs 0.40.
        self.assertAlmostEqual(r_center, 0.16)
        self.assertAlmostEqual(r_leaves, 0.40)

    def test_r_index_zero_on_empty_or_single_point_curve(self) -> None:
        from network.robustness import r_index

        self.assertEqual(r_index([]), 0.0)
        self.assertEqual(r_index([1.0]), 0.0)

    def test_critical_threshold_returns_fraction_at_collapse(self) -> None:
        from network.robustness import critical_threshold

        # N=4, S(0)=1, threshold = 0.05*1 = 0.05; first drop below at q=2.
        curve = [1.0, 0.5, 0.04, 0.0, 0.0]
        self.assertAlmostEqual(critical_threshold(curve, drop_to=0.05), 0.5)

    def test_critical_threshold_none_when_never_reached(self) -> None:
        from network.robustness import critical_threshold

        self.assertIsNone(critical_threshold([1.0, 0.9, 0.8, 0.7, 0.6, 0.5], drop_to=0.05))

    def test_critical_threshold_handles_degenerate_inputs(self) -> None:
        from network.robustness import critical_threshold

        self.assertIsNone(critical_threshold([], 0.05))
        self.assertIsNone(critical_threshold([0.0, 0.0], 0.05))
        self.assertIsNone(critical_threshold([0.0], 0.05))


class WeightedGlobalEfficiencyTests(TestCase):
    def test_returns_zero_for_acyclic_graph(self) -> None:
        # A → B has no two-node SCC, so efficiency is 0.
        from network.robustness import weighted_global_efficiency

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=2.0)
        self.assertEqual(weighted_global_efficiency(g), 0.0)

    def test_two_node_bidirectional_matches_formula(self) -> None:
        # Distance = 1/weight = 0.5; 1/d = 2.0 per ordered pair; sum / (n*(n-1)) = 4/2 = 2.0.
        from network.robustness import weighted_global_efficiency

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=2.0)
        g.add_edge("B", "A", weight=2.0)
        self.assertAlmostEqual(weighted_global_efficiency(g), 2.0)

    def test_two_node_unit_weight_matches_unweighted_value(self) -> None:
        # With weight 1 the weighted form coincides with the unweighted efficiency = 1.
        from network.robustness import weighted_global_efficiency

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        g.add_edge("B", "A", weight=1.0)
        self.assertAlmostEqual(weighted_global_efficiency(g), 1.0)

    def test_empty_and_singleton_graphs_return_zero(self) -> None:
        from network.robustness import weighted_global_efficiency

        self.assertEqual(weighted_global_efficiency(nx.DiGraph()), 0.0)
        g = nx.DiGraph()
        g.add_node("A")
        self.assertEqual(weighted_global_efficiency(g), 0.0)

    def test_nodes_argument_restricts_before_scc_search(self) -> None:
        # Full graph: bidirectional triangle A↔B↔C↔A → one SCC = {A, B, C}.
        # Restricted to {A, B}: only the A↔B pair, SCC = {A, B}, efficiency = 1.
        from network.robustness import weighted_global_efficiency

        g = nx.DiGraph()
        for u, v in [("A", "B"), ("B", "A"), ("B", "C"), ("C", "B"), ("A", "C"), ("C", "A")]:
            g.add_edge(u, v, weight=1.0)
        self.assertAlmostEqual(weighted_global_efficiency(g), 1.0)  # full triangle
        self.assertAlmostEqual(weighted_global_efficiency(g, nodes={"A", "B"}), 1.0)


class ResidualSizesTests(TestCase):
    def test_matches_attack_curve_final_point(self) -> None:
        # One-shot removal of a node set must agree with the corresponding
        # attack-curve point for every metric (REACH exact — no sampling).
        from network.robustness import attack_curve, residual_sizes

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=2.0)
        g.add_edge("B", "C", weight=1.0)
        g.add_edge("C", "A", weight=1.0)
        g.add_edge("C", "D", weight=0.5)
        sizes = residual_sizes(g, ["A", "D"], reach_sample=None)
        for metric_key, metric in (("wcc", "WCC"), ("scc", "SCC"), ("reach", "REACH"), ("strength", "STRENGTH")):
            curve = attack_curve(g, ["A", "D"], metric, reach_sample=None)
            self.assertAlmostEqual(sizes[metric_key], curve[-1], msg=metric_key)

    def test_missing_nodes_silently_skipped_and_graph_unmutated(self) -> None:
        from network.robustness import residual_sizes

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        before = (set(g.nodes()), set(g.edges()))
        sizes = residual_sizes(g, ["X", "A"], reach_sample=None)
        self.assertAlmostEqual(sizes["wcc"], 0.5)
        self.assertEqual((set(g.nodes()), set(g.edges())), before)

    def test_empty_graph_returns_zeroes(self) -> None:
        from network.robustness import residual_sizes

        sizes = residual_sizes(nx.DiGraph(), ["A"])
        self.assertEqual(sizes, {"wcc": 0.0, "scc": 0.0, "reach": 0.0, "strength": 0.0})


class EfficiencyCurveTests(TestCase):
    def test_grid_includes_endpoints_and_is_monotone_in_f(self) -> None:
        from network.robustness import efficiency_curve

        g = nx.gnp_random_graph(30, 0.2, seed=42, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0
        fractions, values = efficiency_curve(g, list(g.nodes()), n_points=10)
        self.assertEqual(fractions[0], 0.0)
        self.assertEqual(fractions[-1], 1.0)
        self.assertEqual(len(fractions), len(values))
        self.assertEqual(fractions, sorted(fractions))
        self.assertEqual(values[-1], 0.0)  # empty residual has no SCC

    def test_baseline_point_matches_weighted_global_efficiency(self) -> None:
        from network.robustness import efficiency_curve, weighted_global_efficiency

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=2.0)
        g.add_edge("B", "A", weight=2.0)
        g.add_edge("B", "C", weight=1.0)
        fractions, values = efficiency_curve(g, ["A", "B", "C"], n_points=3)
        self.assertAlmostEqual(values[0], weighted_global_efficiency(g))

    def test_empty_order_returns_single_baseline_point(self) -> None:
        from network.robustness import efficiency_curve

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        fractions, values = efficiency_curve(g, [], n_points=5)
        self.assertEqual(fractions, [0.0])
        self.assertEqual(len(values), 1)

    def test_grid_never_exceeds_n_points_plus_one(self) -> None:
        from network.robustness import efficiency_curve

        g = nx.gnp_random_graph(50, 0.1, seed=7, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0
        fractions, _ = efficiency_curve(g, list(g.nodes()), n_points=20)
        self.assertLessEqual(len(fractions), 21)


class RemovalOrderTests(TestCase):
    def test_strategy_constants_partition_correctly(self) -> None:
        from network.robustness import ALL_STRATEGIES, DEFAULT_STRATEGIES, DYNAMIC_STRATEGIES, STATIC_STRATEGIES

        self.assertEqual(set(ALL_STRATEGIES), STATIC_STRATEGIES | DYNAMIC_STRATEGIES)
        self.assertEqual(STATIC_STRATEGIES & DYNAMIC_STRATEGIES, set())
        # Static: random + degree (2) + prestige (1) + bridges (1) + dismantling CI (1) + visibility (1) = 7
        self.assertEqual(len(STATIC_STRATEGIES), 7)
        # Dynamic: in/out_strength_dyn, pagerank_dyn, betweenness_dyn, collective_influence_dyn, fragmentation_dyn
        self.assertEqual(len(DYNAMIC_STRATEGIES), 6)
        self.assertEqual(len(ALL_STRATEGIES), 13)
        self.assertEqual(DEFAULT_STRATEGIES, ["random", "in_strength", "out_strength", "pagerank", "betweenness"])

    def test_random_returns_permutation_of_nodes(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C"), ("C", "D")])
        order = removal_order(g, "random", rng=np.random.default_rng(0))
        self.assertEqual(sorted(order), sorted(g.nodes()))

    def test_random_reproducible_with_same_seed(self) -> None:
        from network.robustness import removal_order

        g = nx.gnp_random_graph(20, 0.2, seed=42, directed=True)
        o1 = removal_order(g, "random", rng=np.random.default_rng(0))
        o2 = removal_order(g, "random", rng=np.random.default_rng(0))
        self.assertEqual(o1, o2)
        # Different seeds should diverge (probability ≈ 1 for a 20-node permutation)
        o3 = removal_order(g, "random", rng=np.random.default_rng(1))
        self.assertNotEqual(o1, o3)

    def test_in_strength_static_sorted_descending(self) -> None:
        from network.robustness import removal_order

        # B in_str=3, C in_str=2, A and X both in_str=0
        g = nx.DiGraph()
        g.add_edge("X", "B", weight=3.0)
        g.add_edge("X", "C", weight=2.0)
        g.add_node("A")
        order = removal_order(g, "in_strength")
        # Sort: B(3), C(2), A(0), X(0)  — A before X by ascending ID
        self.assertEqual(order, ["B", "C", "A", "X"])

    def test_out_strength_static_sorted_descending(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=5.0)
        g.add_edge("A", "C", weight=3.0)
        g.add_edge("D", "C", weight=1.0)
        # A out_str=8, D out_str=1, B and C out_str=0
        order = removal_order(g, "out_strength")
        self.assertEqual(order, ["A", "D", "B", "C"])

    def test_static_tie_breaking_is_ascending_node_id(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        for nid in ("Z", "A", "M"):
            g.add_node(nid)
        order = removal_order(g, "in_strength")
        self.assertEqual(order, ["A", "M", "Z"])

    def test_in_strength_dyn_differs_from_static_on_chain(self) -> None:
        # Chain A → B → C → D, all weight 1.
        # Static in_strength: B, C, D tied at 1; sort by ID; A=0 last → [B, C, D, A].
        # Dynamic: B picked first (tie at 1, smallest ID). Removing B disconnects C, so C and D
        # now both have in_str 0 except D which still has in_edge from C → D picked next.
        # Then A and C both 0 → A (smaller ID), then C.
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        g.add_edge("B", "C", weight=1.0)
        g.add_edge("C", "D", weight=1.0)
        self.assertEqual(removal_order(g, "in_strength"), ["B", "C", "D", "A"])
        self.assertEqual(removal_order(g, "in_strength_dyn"), ["B", "D", "A", "C"])

    def test_pagerank_returns_permutation_of_nodes(self) -> None:
        from network.robustness import removal_order

        g = nx.gnp_random_graph(15, 0.2, seed=42, directed=True)
        order = removal_order(g, "pagerank")
        self.assertEqual(sorted(order), sorted(g.nodes()))

    def test_pagerank_dyn_returns_permutation_of_nodes(self) -> None:
        from network.robustness import removal_order

        g = nx.gnp_random_graph(15, 0.2, seed=42, directed=True)
        order = removal_order(g, "pagerank_dyn")
        self.assertEqual(sorted(order), sorted(g.nodes()))

    def test_empty_graph_returns_empty_order(self) -> None:
        from network.robustness import removal_order

        self.assertEqual(removal_order(nx.DiGraph(), "in_strength"), [])
        self.assertEqual(removal_order(nx.DiGraph(), "random", rng=np.random.default_rng(0)), [])
        self.assertEqual(removal_order(nx.DiGraph(), "pagerank_dyn"), [])

    def test_invalid_strategy_raises(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_node("A")
        with self.assertRaises(ValueError):
            removal_order(g, "no-such-strategy")

    def test_does_not_mutate_input_graph(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C"), ("C", "A")])
        before_nodes, before_edges = set(g.nodes()), set(g.edges())
        for strat in ("in_strength", "pagerank", "pagerank_dyn"):
            removal_order(g, strat)
        self.assertEqual((set(g.nodes()), set(g.edges())), (before_nodes, before_edges))

    def test_dynamic_scorers_return_permutations(self) -> None:
        from network.robustness import removal_order

        g = nx.gnp_random_graph(10, 0.3, seed=42, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0
        for strat in ("in_strength_dyn", "out_strength_dyn", "pagerank_dyn"):
            with self.subTest(strategy=strat):
                order = removal_order(g, strat)
                self.assertEqual(sorted(order), sorted(g.nodes()))

    def test_strategy_names_are_case_insensitive(self) -> None:
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        self.assertEqual(removal_order(g, "PageRank"), removal_order(g, "pagerank"))
        self.assertEqual(removal_order(g, "IN_STRENGTH"), removal_order(g, "in_strength"))

    def test_betweenness_removes_bridge_first(self) -> None:
        # Two directed triangles joined by the bridge B1 ↔ B2: every cross-side
        # shortest path runs through both bridge endpoints, so they rank first.
        from network.robustness import removal_order

        g = nx.DiGraph()
        for a, b in [("L1", "L2"), ("L2", "L3"), ("L3", "L1"), ("R1", "R2"), ("R2", "R3"), ("R3", "R1")]:
            g.add_edge(a, b, weight=1.0)
        g.add_edge("L1", "R1", weight=1.0)
        g.add_edge("R1", "L1", weight=1.0)
        order = removal_order(g, "betweenness")
        self.assertEqual(set(order[:2]), {"L1", "R1"})

    def test_betweenness_weighted_prefers_heavy_path(self) -> None:
        # A reaches C via B (heavy edges, distance 1/10 each) or via D (light
        # edges, distance 1/0.1 each).  Shortest paths take the heavy route, so
        # B has positive betweenness and D none.
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=10.0)
        g.add_edge("B", "C", weight=10.0)
        g.add_edge("A", "D", weight=0.1)
        g.add_edge("D", "C", weight=0.1)
        order = removal_order(g, "betweenness")
        self.assertEqual(order[0], "B")

    def test_betweenness_dyn_returns_permutation_of_nodes(self) -> None:
        from network.robustness import removal_order

        g = nx.gnp_random_graph(12, 0.25, seed=42, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0
        order = removal_order(g, "betweenness_dyn")
        self.assertEqual(sorted(order), sorted(g.nodes()))

    def test_subscribers_sorted_by_fans_descending(self) -> None:
        # Audience size lives on the node's ``data`` dict (graph_builder's
        # "fans"); unknown audiences (missing dict or None) go last, by ID.
        from network.robustness import removal_order

        g = nx.DiGraph()
        g.add_node("A", data={"fans": 500})
        g.add_node("B", data={"fans": 10_000})
        g.add_node("C", data={"fans": None})
        g.add_node("D")
        order = removal_order(g, "subscribers")
        self.assertEqual(order, ["B", "A", "C", "D"])


class RewireStrengthPreservingTests(TestCase):
    def _gnp(self, n: int = 30, p: float = 0.15, seed: int = 42) -> nx.DiGraph:
        g = nx.gnp_random_graph(n, p, seed=seed, directed=True)
        rng = np.random.default_rng(seed)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = float(rng.uniform(0.5, 5.0))
        return g

    def test_degree_sequence_preserved_exactly(self) -> None:
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        h = rewire_strength_preserving(g, rng=np.random.default_rng(0))
        self.assertEqual(set(g.nodes()), set(h.nodes()))
        self.assertEqual(dict(g.in_degree()), dict(h.in_degree()))
        self.assertEqual(dict(g.out_degree()), dict(h.out_degree()))

    def test_strength_sequence_preserved(self) -> None:
        # IPF restores each node's in/out strength onto the observed values.
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        h = rewire_strength_preserving(g, rng=np.random.default_rng(0))
        for n, s in dict(g.out_degree(weight="weight")).items():
            self.assertAlmostEqual(s, h.out_degree(n, weight="weight"), places=4)
        for n, s in dict(g.in_degree(weight="weight")).items():
            self.assertAlmostEqual(s, h.in_degree(n, weight="weight"), places=4)

    def test_total_weight_preserved(self) -> None:
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        before = sum(d["weight"] for _, _, d in g.edges(data=True))
        h = rewire_strength_preserving(g, rng=np.random.default_rng(0))
        after = sum(d["weight"] for _, _, d in h.edges(data=True))
        self.assertAlmostEqual(before, after, places=4)

    def test_topology_is_randomised(self) -> None:
        # Same edge count, but the wiring changes (that is the point of the null).
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        h = rewire_strength_preserving(g, rng=np.random.default_rng(0))
        self.assertEqual(g.number_of_edges(), h.number_of_edges())
        self.assertNotEqual(set(g.edges()), set(h.edges()))

    def test_reproducible_with_same_seed(self) -> None:
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        h1 = rewire_strength_preserving(g, rng=np.random.default_rng(7))
        h2 = rewire_strength_preserving(g, rng=np.random.default_rng(7))
        self.assertEqual(set(h1.edges()), set(h2.edges()))
        for e in h1.edges():
            self.assertEqual(h1.edges[e]["weight"], h2.edges[e]["weight"])

    def test_does_not_mutate_input(self) -> None:
        from network.robustness import rewire_strength_preserving

        g = self._gnp()
        edges_before = set(g.edges())
        weights_before = {e: g.edges[e]["weight"] for e in g.edges()}
        rewire_strength_preserving(g, rng=np.random.default_rng(0))
        self.assertEqual(set(g.edges()), edges_before)
        for e, w in weights_before.items():
            self.assertEqual(g.edges[e]["weight"], w)

    def test_small_graph_returns_copy_unchanged(self) -> None:
        from network.robustness import rewire_strength_preserving

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=3.0)  # only 1 edge → nothing to swap
        h = rewire_strength_preserving(g, rng=np.random.default_rng(0))
        self.assertEqual(h.edges["A", "B"]["weight"], 3.0)
        # And empty graph
        h2 = rewire_strength_preserving(nx.DiGraph(), rng=np.random.default_rng(0))
        self.assertEqual(h2.number_of_nodes(), 0)


class NullDistributionTests(TestCase):
    def test_yields_requested_number_of_graphs(self) -> None:
        from network.robustness import null_distribution

        g = nx.gnp_random_graph(20, 0.2, seed=42, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0
        result = list(null_distribution(g, n_simulations=5, rng=np.random.default_rng(0)))
        self.assertEqual(len(result), 5)
        for h in result:
            # Strength-preserving null randomises wiring but keeps the degree sequence.
            self.assertEqual(g.number_of_edges(), h.number_of_edges())
            self.assertEqual(dict(g.in_degree()), dict(h.in_degree()))
            self.assertEqual(dict(g.out_degree()), dict(h.out_degree()))

    def test_zero_simulations_yields_nothing(self) -> None:
        from network.robustness import null_distribution

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        self.assertEqual(list(null_distribution(g, n_simulations=0)), [])

    def test_successive_simulations_differ_under_shared_rng(self) -> None:
        # Two consecutive nulls drawn from the same rng should differ (the rng
        # state advances between calls, producing different rewirings).
        from network.robustness import null_distribution

        g = nx.gnp_random_graph(30, 0.3, seed=42, directed=True)
        rng_seed = np.random.default_rng(42)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = float(rng_seed.uniform(0.5, 5.0))
        nulls = list(null_distribution(g, n_simulations=2, rng=np.random.default_rng(0)))
        self.assertNotEqual(set(nulls[0].edges()), set(nulls[1].edges()))


class ZScoreTests(TestCase):
    def test_basic_computation(self) -> None:
        from network.robustness import z_score

        # samples have mean 5, sample std (ddof=1) = sqrt(2.5) ≈ 1.5811
        z, mu, sigma = z_score(observed=10.0, null_samples=[3.0, 4.0, 5.0, 6.0, 7.0])
        self.assertAlmostEqual(mu, 5.0)
        self.assertAlmostEqual(sigma, np.std([3.0, 4.0, 5.0, 6.0, 7.0], ddof=1))
        self.assertAlmostEqual(z, (10.0 - 5.0) / sigma)

    def test_zero_std_returns_nan_z(self) -> None:
        from network.robustness import z_score

        z, mu, sigma = z_score(observed=10.0, null_samples=[5.0, 5.0, 5.0])
        self.assertEqual(mu, 5.0)
        self.assertEqual(sigma, 0.0)
        self.assertTrue(np.isnan(z))

    def test_single_sample_treated_as_zero_std(self) -> None:
        from network.robustness import z_score

        z, mu, sigma = z_score(observed=10.0, null_samples=[3.0])
        self.assertEqual(mu, 3.0)
        self.assertEqual(sigma, 0.0)
        self.assertTrue(np.isnan(z))

    def test_empty_samples_returns_nan_triple(self) -> None:
        from network.robustness import z_score

        z, mu, sigma = z_score(observed=10.0, null_samples=[])
        self.assertTrue(np.isnan(z))
        self.assertTrue(np.isnan(mu))
        self.assertTrue(np.isnan(sigma))

    def test_negative_z_when_observed_below_mean(self) -> None:
        from network.robustness import z_score

        z, _, _ = z_score(observed=1.0, null_samples=[5.0, 5.5, 6.0, 6.5, 7.0])
        self.assertLess(z, 0)


class EmpiricalPTests(TestCase):
    def test_observed_below_all_samples_hits_resolution_floor(self) -> None:
        from network.robustness import empirical_p

        samples = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]  # K = 9
        # Lower tail: (0 + 1) / 10 = 0.1 → two-sided 0.2, the floor 2/(K+1).
        self.assertAlmostEqual(empirical_p(0.1, samples), 0.2)

    def test_observed_in_the_middle_is_capped_at_one(self) -> None:
        from network.robustness import empirical_p

        samples = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
        self.assertEqual(empirical_p(0.5, samples), 1.0)

    def test_ties_count_in_both_tails(self) -> None:
        from network.robustness import empirical_p

        # Observed equals the minimum sample, so it counts in both tails:
        # lower tail (1+1)/4 = 0.5, upper tail (3+1)/4 = 1.0 → 2 · 0.5 = 1.0.
        self.assertEqual(empirical_p(0.1, [0.1, 0.2, 0.3]), 1.0)
        # Observed strictly below every sample: (0+1)/4 = 0.25 → 0.5.
        self.assertAlmostEqual(empirical_p(0.05, [0.1, 0.2, 0.3]), 0.5)

    def test_defined_on_zero_variance_null(self) -> None:
        # z_score returns nan here; the empirical p stays informative.
        # (0.5 is dyadic, so the mean of identical samples is exact and σ is
        # exactly zero — 0.4 would leave a ~1e-17 residual std.)
        from network.robustness import empirical_p, z_score

        samples = [0.5, 0.5, 0.5]
        z, _, _ = z_score(0.1, samples)
        self.assertTrue(np.isnan(z))
        self.assertAlmostEqual(empirical_p(0.1, samples), 0.5)  # floor 2/(K+1)
        self.assertEqual(empirical_p(0.5, samples), 1.0)  # inside the degenerate null

    def test_empty_samples_return_nan(self) -> None:
        from network.robustness import empirical_p

        self.assertTrue(np.isnan(empirical_p(0.5, [])))

    def test_never_zero_and_never_above_one(self) -> None:
        from network.robustness import empirical_p

        rng = np.random.default_rng(0)
        samples = list(rng.uniform(0, 1, size=19))
        for observed in (-1.0, 0.0, 0.5, 1.0, 2.0):
            p = empirical_p(observed, samples)
            self.assertGreaterEqual(p, 2.0 / 20.0)
            self.assertLessEqual(p, 1.0)


class ModularRobustnessCurvesTests(TestCase):
    def _two_cliques_with_bridge(self) -> tuple[nx.DiGraph, dict[str, int]]:
        # Two 3-node directed cliques + a single inter-community bridge A → D.
        # Each clique contributes 6 directed edges (3 pairs × 2 directions).
        g = nx.DiGraph()
        for u, v in [("A", "B"), ("B", "A"), ("A", "C"), ("C", "A"), ("B", "C"), ("C", "B")]:
            g.add_edge(u, v)
        for u, v in [("D", "E"), ("E", "D"), ("D", "F"), ("F", "D"), ("E", "F"), ("F", "E")]:
            g.add_edge(u, v)
        g.add_edge("A", "D")  # the only inter-community edge
        partition = {"A": 0, "B": 0, "C": 0, "D": 1, "E": 1, "F": 1}
        return g, partition

    def test_curve_length_matches_removal_order_plus_one(self) -> None:
        from network.robustness import modular_robustness_curves

        g, part = self._two_cliques_with_bridge()
        curves = modular_robustness_curves(g, ["A", "B", "C"], part)
        for key in ("intra", "inter", "ratio"):
            self.assertEqual(len(curves[key]), 4)

    def test_baselines_at_q_zero(self) -> None:
        # 12 intra edges (6 per clique) + 1 inter; both baselines = 1.0
        from network.robustness import modular_robustness_curves

        g, part = self._two_cliques_with_bridge()
        curves = modular_robustness_curves(g, [], part)
        self.assertEqual(curves["intra"], [1.0])
        self.assertEqual(curves["inter"], [1.0])
        # ratio = 12 / 1 = 12.0
        self.assertEqual(curves["ratio"], [12.0])

    def test_bridge_endpoint_removal_strips_inter_edge(self) -> None:
        # Removing A drops the single inter edge (A→D) and four intra edges
        # incident on A (A→B, A→C, B→A, C→A).  Expected after q=1:
        #   intra_q = 12 - 4 = 8  →  intra[1] = 8/12 ≈ 0.667
        #   inter_q = 1 - 1 = 0    →  inter[1] = 0.0; ratio[1] = None
        from network.robustness import modular_robustness_curves

        g, part = self._two_cliques_with_bridge()
        curves = modular_robustness_curves(g, ["A"], part)
        self.assertAlmostEqual(curves["intra"][1], 8 / 12)
        self.assertEqual(curves["inter"][1], 0.0)
        self.assertIsNone(curves["ratio"][1])

    def test_all_same_community_inter_curve_is_zero(self) -> None:
        from network.robustness import modular_robustness_curves

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C")])
        part = {"A": 0, "B": 0, "C": 0}
        curves = modular_robustness_curves(g, ["A", "B", "C"], part)
        # inter_0 == 0 → entire inter curve is 0.0
        self.assertEqual(curves["inter"], [0.0, 0.0, 0.0, 0.0])
        # ratio is always None when inter is 0 throughout
        self.assertEqual(curves["ratio"], [None, None, None, None])

    def test_all_different_communities_intra_curve_is_zero(self) -> None:
        from network.robustness import modular_robustness_curves

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C")])
        part = {"A": 0, "B": 1, "C": 2}
        curves = modular_robustness_curves(g, ["A"], part)
        self.assertEqual(curves["intra"], [0.0, 0.0])
        # inter_0 = 2, after removing A → 1 inter edge gone
        self.assertAlmostEqual(curves["inter"][0], 1.0)
        self.assertAlmostEqual(curves["inter"][1], 0.5)
        self.assertAlmostEqual(curves["ratio"][0], 0.0)
        self.assertAlmostEqual(curves["ratio"][1], 0.0)

    def test_self_loop_counted_once(self) -> None:
        # A→A (self-loop), A→B, B→A. All same community.  Removing A should
        # drop all 3 edges, not 4 (which would be the bug from counting the
        # self-loop both in out_edges and in_edges).
        from network.robustness import modular_robustness_curves

        g = nx.DiGraph()
        g.add_edges_from([("A", "A"), ("A", "B"), ("B", "A")])
        part = {"A": 0, "B": 0}
        curves = modular_robustness_curves(g, ["A"], part)
        # intra_0 = 3 → intra[0] = 1.0, intra[1] = 0/3 = 0.0
        self.assertEqual(curves["intra"], [1.0, 0.0])

    def test_unassigned_nodes_treated_as_inter(self) -> None:
        # B has no entry in partition → A→B counts as inter even though B's
        # community is conceptually "unknown".
        from network.robustness import modular_robustness_curves

        g = nx.DiGraph()
        g.add_edge("A", "B")
        part = {"A": 0}  # B missing
        curves = modular_robustness_curves(g, [], part)
        self.assertEqual(curves["intra"], [0.0])  # intra_0 == 0
        self.assertEqual(curves["inter"], [1.0])  # the one edge is inter

    def test_missing_nodes_in_removal_order_are_skipped(self) -> None:
        from network.robustness import modular_robustness_curves

        g = nx.DiGraph()
        g.add_edges_from([("A", "B"), ("B", "C")])
        part = {"A": 0, "B": 0, "C": 0}
        # "X" is not in the graph
        curves = modular_robustness_curves(g, ["A", "X", "B"], part)
        # After A: lose A→B → intra = 1, fraction = 1/2 = 0.5
        # After X (skipped): unchanged → 0.5
        # After B: lose B→C → intra = 0, fraction = 0.0
        self.assertEqual(curves["intra"], [1.0, 0.5, 0.5, 0.0])

    def test_does_not_mutate_input_graph(self) -> None:
        from network.robustness import modular_robustness_curves

        g, part = self._two_cliques_with_bridge()
        before_nodes, before_edges = set(g.nodes()), set(g.edges())
        modular_robustness_curves(g, ["A", "B", "D"], part)
        self.assertEqual((set(g.nodes()), set(g.edges())), (before_nodes, before_edges))

    def test_empty_graph_returns_single_zero_baselines(self) -> None:
        from network.robustness import modular_robustness_curves

        curves = modular_robustness_curves(nx.DiGraph(), [], {})
        self.assertEqual(curves["intra"], [0.0])
        self.assertEqual(curves["inter"], [0.0])
        self.assertEqual(curves["ratio"], [None])


class BanWaveScenarioTests(TestCase):
    def _bridged_cliques(self) -> tuple[nx.DiGraph, dict[str, str]]:
        g = nx.DiGraph()
        for side in ("L", "R"):
            for i in range(3):
                for j in range(3):
                    if i != j:
                        g.add_edge(f"{side}{i}", f"{side}{j}", weight=1.0)
        g.add_edge("L0", "R0", weight=1.0)
        partition = {n: n[0] for n in g.nodes()}
        return g, partition

    def test_one_row_per_block_with_all_fields(self) -> None:
        from network.robustness import ban_wave_rows

        g, partition = self._bridged_cliques()
        random_curves = {m: [1.0] * (g.number_of_nodes() + 1) for m in ("wcc", "scc", "reach", "strength")}
        rows = ban_wave_rows(g, partition, random_curves, reach_sample=None, rng=np.random.default_rng(0))
        self.assertEqual([r["community"] for r in rows], ["L", "R"])
        for row in rows:
            self.assertEqual(row["n"], 3)
            self.assertAlmostEqual(row["fraction"], 0.5)
            for m in ("wcc", "scc", "reach", "strength"):
                self.assertIn(f"s_{m}", row)
                self.assertEqual(row[f"random_{m}"], 1.0)
        # Removing either 3-clique leaves the other intact: half the nodes survive.
        self.assertAlmostEqual(rows[0]["s_wcc"], 0.5)

    def test_small_blocks_and_missing_baseline_handled(self) -> None:
        from network.robustness import ban_wave_rows

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        g.add_edge("B", "C", weight=1.0)
        g.add_node("D")
        partition = {"A": 1, "B": 1, "C": 2, "D": 3}  # blocks of size 2, 1, 1
        rows = ban_wave_rows(g, partition, {}, reach_sample=None, rng=np.random.default_rng(0))
        self.assertEqual(len(rows), 1)  # singleton blocks skipped
        self.assertEqual(rows[0]["community"], "1")
        self.assertIsNone(rows[0]["random_wcc"])  # no baseline curve provided

    def test_block_covering_whole_graph_is_skipped(self) -> None:
        from network.robustness import ban_wave_rows

        g = nx.DiGraph()
        g.add_edge("A", "B", weight=1.0)
        rows = ban_wave_rows(g, {"A": 0, "B": 0}, {}, reach_sample=None, rng=np.random.default_rng(0))
        self.assertEqual(rows, [])

    def test_does_not_mutate_input_graph(self) -> None:
        from network.robustness import ban_wave_rows

        g, partition = self._bridged_cliques()
        before = (set(g.nodes()), set(g.edges()))
        ban_wave_rows(g, partition, {}, reach_sample=None, rng=np.random.default_rng(0))
        self.assertEqual((set(g.nodes()), set(g.edges())), before)


class RobustnessRunnerTests(TestCase):
    def _toy_graph(self, n: int = 20, p: float = 0.2, seed: int = 42) -> nx.DiGraph:
        g = nx.gnp_random_graph(n, p, seed=seed, directed=True)
        rng = np.random.default_rng(seed)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = float(rng.uniform(0.5, 5.0))
        return g

    def _fast_cfg(self, **kwargs: Any) -> Any:
        from network.robustness import RobustnessConfig

        defaults: dict[str, Any] = {
            "alpha": None,
            "n_random_runs": 3,
            "n_null": 0,
            "seed": 0,
            "reach_sample": 10,
        }
        defaults.update(kwargs)
        return RobustnessConfig(**defaults)

    # -- config validation ----------------------------------------------------

    def test_config_defaults_match_spec(self) -> None:
        from network.robustness import RobustnessConfig

        c = RobustnessConfig()
        self.assertEqual(
            (c.alpha, c.n_random_runs, c.n_null, c.seed, c.reach_sample, c.strategies),
            (0.05, 100, 20, 42, 500, None),
        )

    def test_config_rejects_invalid_values(self) -> None:
        from network.robustness import RobustnessConfig

        with self.assertRaises(ValueError):
            RobustnessConfig(n_random_runs=0)
        with self.assertRaises(ValueError):
            RobustnessConfig(n_null=-1)
        with self.assertRaises(ValueError):
            RobustnessConfig(alpha=1.5)
        with self.assertRaises(ValueError):
            RobustnessConfig(reach_sample=0)

    def test_config_rejects_empty_strategies_list(self) -> None:
        from network.robustness import RobustnessConfig

        with self.assertRaises(ValueError):
            RobustnessConfig(strategies=[])

    def test_config_rejects_unknown_strategy(self) -> None:
        from network.robustness import RobustnessConfig

        with self.assertRaises(ValueError):
            RobustnessConfig(strategies=["no-such-strategy"])

    def test_config_accepts_surviving_strategies(self) -> None:
        from network.robustness import RobustnessConfig

        RobustnessConfig(strategies=["pagerank", "in_strength_dyn"])

    # -- payload shape --------------------------------------------------------

    def test_payload_top_level_keys(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(), config=self._fast_cfg())
        self.assertEqual(
            set(out.keys()),
            {"config", "graph", "efficiency", "strategies", "modular", "ban_waves", "alpha_sensitivity"},
        )

    def test_payload_strategy_keys_defaults_to_default_set(self) -> None:
        from network.robustness import DEFAULT_STRATEGIES, run_robustness

        out = run_robustness(self._toy_graph(), config=self._fast_cfg())
        self.assertEqual(set(out["strategies"].keys()), set(DEFAULT_STRATEGIES))

    def test_payload_strategy_keys_match_explicit_selection(self) -> None:
        from network.robustness import run_robustness

        chosen = ["pagerank", "in_strength_dyn", "out_strength", "pagerank_dyn"]
        out = run_robustness(self._toy_graph(n=10), config=self._fast_cfg(strategies=chosen))
        self.assertEqual(set(out["strategies"].keys()), set(chosen))

    def test_payload_strategy_label_present_per_strategy(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(), config=self._fast_cfg(strategies=["pagerank"]))
        self.assertEqual(out["strategies"]["pagerank"]["label"], "PageRank")

    def test_each_strategy_has_four_curves_and_r_fc(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(), config=self._fast_cfg())
        for s, payload in out["strategies"].items():
            for m in ("wcc", "scc", "reach", "strength"):
                self.assertIn(f"curve_{m}", payload, msg=f"strategy {s!r} missing curve_{m}")
                self.assertIn(f"r_{m}", payload)
                self.assertIn(f"fc_{m}", payload)
            self.assertIsNone(payload["null"])  # n_null=0

    def test_curve_length_matches_backbone_node_count_plus_one(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(n=12), config=self._fast_cfg())
        n_plus_1 = out["graph"]["backbone_n"] + 1
        for payload in out["strategies"].values():
            for m in ("wcc", "scc", "reach", "strength"):
                self.assertEqual(len(payload[f"curve_{m}"]), n_plus_1)

    def test_efficiency_block_has_curves_on_shared_grid(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(n=12), config=self._fast_cfg())
        eff = out["efficiency"]
        self.assertEqual(set(eff.keys()), {"baseline", "fractions", "curves"})
        self.assertEqual(set(eff["curves"].keys()), set(out["strategies"].keys()))
        self.assertEqual(eff["fractions"][0], 0.0)
        self.assertEqual(eff["fractions"][-1], 1.0)
        for curve in eff["curves"].values():
            self.assertEqual(len(curve), len(eff["fractions"]))
        # The q = 0 grid point is the baseline efficiency by construction.
        for curve in eff["curves"].values():
            self.assertAlmostEqual(curve[0], eff["baseline"])

    # -- disparity filter ------------------------------------------------------

    def test_alpha_none_skips_filter(self) -> None:
        from network.robustness import run_robustness

        g = self._toy_graph()
        out = run_robustness(g, config=self._fast_cfg(alpha=None))
        self.assertFalse(out["graph"]["filtered"])
        self.assertEqual(out["graph"]["backbone_m"], g.number_of_edges())

    def test_alpha_zero_skips_filter(self) -> None:
        from network.robustness import run_robustness

        g = self._toy_graph()
        out = run_robustness(g, config=self._fast_cfg(alpha=0))
        self.assertFalse(out["graph"]["filtered"])
        self.assertEqual(out["graph"]["backbone_m"], g.number_of_edges())

    def test_alpha_within_range_applies_filter(self) -> None:
        from network.robustness import run_robustness

        g = self._toy_graph()
        out = run_robustness(g, config=self._fast_cfg(alpha=0.5))
        self.assertTrue(out["graph"]["filtered"])
        self.assertLessEqual(out["graph"]["backbone_m"], g.number_of_edges())

    # -- null model -----------------------------------------------------------

    def test_null_populated_when_n_null_positive(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(n=12), config=self._fast_cfg(n_null=3))
        for payload in out["strategies"].values():
            self.assertIsNotNone(payload["null"])
            for m in ("wcc", "scc", "reach", "strength"):
                self.assertIn(f"r_{m}", payload["null"])
                self.assertEqual(set(payload["null"][f"r_{m}"].keys()), {"mean", "std", "z", "p", "q"})
                # 3 null draws → the add-one two-sided p lives in [2/4, 1].
                p = payload["null"][f"r_{m}"]["p"]
                self.assertGreaterEqual(p, 0.5)
                self.assertLessEqual(p, 1.0)
                self.assertIn(f"curve_{m}_mean", payload["null"])
                self.assertIn(f"curve_{m}_std", payload["null"])

    def test_null_curves_have_matching_length(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(n=10), config=self._fast_cfg(n_null=2))
        n_plus_1 = out["graph"]["backbone_n"] + 1
        for payload in out["strategies"].values():
            for m in ("wcc", "scc", "reach", "strength"):
                self.assertEqual(len(payload["null"][f"curve_{m}_mean"]), n_plus_1)
                self.assertEqual(len(payload["null"][f"curve_{m}_std"]), n_plus_1)

    # -- modular / ban waves ---------------------------------------------------

    def test_modular_none_when_no_partitions(self) -> None:
        from network.robustness import run_robustness

        out = run_robustness(self._toy_graph(), config=self._fast_cfg())
        self.assertIsNone(out["modular"])
        self.assertIsNone(out["ban_waves"])

    def test_ban_waves_populated_when_partitions_given(self) -> None:
        from network.robustness import run_robustness

        g = self._toy_graph(n=10)
        partition = {n: (0 if n < 5 else 1) for n in g.nodes()}
        out = run_robustness(g, partitions={"hand": partition}, config=self._fast_cfg())
        self.assertIsNotNone(out["ban_waves"])
        rows = out["ban_waves"]["hand"]
        self.assertEqual({r["community"] for r in rows}, {"0", "1"})
        # random is in the default strategy set, so the equal-q baseline is present.
        for row in rows:
            self.assertIsNotNone(row["random_wcc"])

    def test_ban_waves_baseline_computed_without_random_strategy(self) -> None:
        from network.robustness import run_robustness

        g = self._toy_graph(n=10)
        partition = {n: (0 if n < 5 else 1) for n in g.nodes()}
        out = run_robustness(g, partitions={"hand": partition}, config=self._fast_cfg(strategies=["pagerank"]))
        rows = out["ban_waves"]["hand"]
        for row in rows:
            self.assertIsNotNone(row["random_wcc"])

    def test_modular_populated_when_partitions_given(self) -> None:
        from network.robustness import DEFAULT_STRATEGIES, run_robustness

        g = self._toy_graph(n=10)
        # Hand-built two-block partition
        partition = {n: (0 if n < 5 else 1) for n in g.nodes()}
        out = run_robustness(g, partitions={"hand": partition}, config=self._fast_cfg())
        self.assertIsNotNone(out["modular"])
        self.assertEqual(set(out["modular"].keys()), {"hand"})
        self.assertEqual(set(out["modular"]["hand"].keys()), set(DEFAULT_STRATEGIES))
        # Each per-strategy entry is the modular_robustness_curves dict.
        for payload in out["modular"]["hand"].values():
            self.assertEqual(set(payload.keys()), {"intra", "inter", "ratio"})

    # -- reproducibility ------------------------------------------------------

    def test_same_seed_produces_identical_payloads(self) -> None:
        import math

        from network.robustness import run_robustness

        def _same(a: float, b: float) -> bool:
            # z is legitimately NaN when the null R distribution has zero variance
            # (z_score documents this); NaN != NaN, so equal-or-both-NaN is the
            # reproducibility check we actually want.
            return a == b or (math.isnan(a) and math.isnan(b))

        g = self._toy_graph()
        out1 = run_robustness(g, config=self._fast_cfg(n_null=2, seed=99))
        out2 = run_robustness(g, config=self._fast_cfg(n_null=2, seed=99))
        # Compare every strategy's R values; both whole payloads must match.
        for s in out1["strategies"]:
            for m in ("wcc", "scc", "reach"):
                self.assertEqual(out1["strategies"][s][f"r_{m}"], out2["strategies"][s][f"r_{m}"])
                z1 = out1["strategies"][s]["null"][f"r_{m}"]["z"]
                z2 = out2["strategies"][s]["null"][f"r_{m}"]["z"]
                self.assertTrue(_same(z1, z2), f"z mismatch for {s}/{m}: {z1!r} vs {z2!r}")

    # -- progress callback ----------------------------------------------------

    def test_progress_callback_receives_expected_labels(self) -> None:
        from network.robustness import run_robustness

        labels: list[str] = []
        g = self._toy_graph(n=8)
        partition = {n: (n % 2) for n in g.nodes()}
        run_robustness(g, partitions={"hand": partition}, config=self._fast_cfg(n_null=1), progress=labels.append)
        self.assertIn("disparity", labels)
        self.assertIn("baseline-efficiency", labels)
        self.assertIn("pagerank", labels)
        self.assertIn("efficiency/pagerank", labels)
        self.assertTrue(any(s.startswith("null/") for s in labels))
        self.assertTrue(any(s.startswith("modular/") for s in labels))
        self.assertTrue(any(s.startswith("banwave/") for s in labels))

    # -- JSON serialisability -------------------------------------------------

    def test_payload_is_json_serialisable(self) -> None:
        import json

        from network.robustness import run_robustness

        g = self._toy_graph(n=8)
        out = run_robustness(
            g,
            partitions={"hand": {n: n % 2 for n in g.nodes()}},
            config=self._fast_cfg(n_null=1),
        )
        # Round-trip must not fail (no inf/nan in the curves; None used for undefined ratios).
        encoded = json.dumps(out)
        self.assertIsInstance(encoded, str)
        decoded = json.loads(encoded)
        self.assertEqual(
            set(decoded.keys()),
            {"config", "graph", "efficiency", "strategies", "modular", "ban_waves", "alpha_sensitivity"},
        )


class DismantlingAttackTests(TestCase):
    def _bridge_graph(self) -> nx.DiGraph:
        # Two triangles joined only through the articulation node "x".
        g = nx.DiGraph()
        for a, b in [("a", "b"), ("b", "c"), ("c", "a"), ("d", "e"), ("e", "f"), ("f", "d"), ("a", "x"), ("x", "d")]:
            g.add_edge(a, b, weight=1.0)
        return g

    def test_collective_influence_ranks_hubs_first(self) -> None:
        from network.robustness import removal_order

        order = removal_order(self._bridge_graph(), "collective_influence")
        self.assertEqual(set(order), set(self._bridge_graph().nodes()))
        # The three high-degree connectors (a, d, x) lead the ranking.
        self.assertEqual(set(order[:3]), {"a", "d", "x"})

    def test_fragmentation_dyn_targets_articulation_point(self) -> None:
        from network.robustness import removal_order

        order = removal_order(self._bridge_graph(), "fragmentation_dyn")
        # "x" is the single articulation point — removing it does the most damage.
        self.assertEqual(order[0], "x")
        self.assertEqual(set(order), set(self._bridge_graph().nodes()))

    def test_dismantling_orders_handle_trivial_graphs(self) -> None:
        from network.robustness import removal_order

        self.assertEqual(removal_order(nx.DiGraph(), "collective_influence_dyn"), [])
        single = nx.DiGraph()
        single.add_node("z")
        self.assertEqual(removal_order(single, "fragmentation_dyn"), ["z"])

    def test_new_strategies_are_registered(self) -> None:
        from network.robustness import ALL_STRATEGIES

        for name in ("collective_influence", "collective_influence_dyn", "fragmentation_dyn"):
            self.assertIn(name, ALL_STRATEGIES)


class ReciprocalNullTests(TestCase):
    def _mixed_graph(self) -> nx.DiGraph:
        g = nx.DiGraph()
        for u, v in [("a", "b"), ("c", "d"), ("e", "f"), ("g", "h")]:
            g.add_edge(u, v, weight=2.0)
            g.add_edge(v, u, weight=3.0)
        for u, v in [("a", "c"), ("b", "d"), ("e", "g"), ("f", "h"), ("a", "e"), ("c", "g")]:
            g.add_edge(u, v, weight=1.5)
        return g

    def test_preserves_degree_and_reciprocity(self) -> None:
        from network.robustness.null_model import rewire_reciprocity_preserving

        g = self._mixed_graph()
        h = rewire_reciprocity_preserving(g, rng=np.random.default_rng(1))
        self.assertEqual(dict(g.out_degree()), dict(h.out_degree()))
        self.assertEqual(dict(g.in_degree()), dict(h.in_degree()))
        recip = lambda G: {n: sum(1 for w in G.successors(n) if G.has_edge(w, n)) for n in G.nodes()}  # noqa: E731
        self.assertEqual(recip(g), recip(h))
        self.assertFalse(any(u == v for u, v in h.edges()))

    def test_distribution_selects_model(self) -> None:
        from network.robustness.null_model import null_distribution

        g = self._mixed_graph()
        draws = list(null_distribution(g, 2, rng=np.random.default_rng(0), model="reciprocal"))
        self.assertEqual(len(draws), 2)
        with self.assertRaises(ValueError):
            list(null_distribution(g, 1, model="bogus"))


class BHAdjustTests(TestCase):
    def test_monotone_and_bounded(self) -> None:
        from network.robustness.null_model import bh_adjust

        qs = bh_adjust([0.01, 0.02, 0.5, 1.0])
        self.assertTrue(all(q <= 1.0 for q in qs))
        # Each q is at least its p (BH only inflates).
        for p, q in zip([0.01, 0.02, 0.5, 1.0], qs, strict=True):
            self.assertGreaterEqual(q + 1e-12, p)

    def test_nan_carried_through(self) -> None:
        import math

        from network.robustness.null_model import bh_adjust

        qs = bh_adjust([0.01, float("nan"), 0.5])
        self.assertTrue(math.isnan(qs[1]))
        self.assertFalse(math.isnan(qs[0]))


class BanReplayTests(TestCase):
    def _graphs(self) -> dict[int, nx.DiGraph]:
        def mk(nodes, edges):
            g = nx.DiGraph()
            g.add_nodes_from(nodes)
            for u, v in edges:
                g.add_edge(u, v, weight=1.0)
            return g

        g19 = mk(list("abcdeh"), [("a", "b"), ("b", "c"), ("c", "d"), ("d", "e"), ("h", "a"), ("h", "c"), ("h", "e")])
        g21 = mk(list("abcdez"), [("a", "b"), ("b", "c"), ("c", "d"), ("d", "e"), ("z", "a"), ("e", "a")])
        return {2019: g19, 2021: g21}

    def test_interior_wave_year_produces_full_row(self) -> None:
        from network.robustness import ban_replay_rows

        rows = ban_replay_rows(self._graphs(), {2020: {"h"}}, n_random_runs=10, rng=np.random.default_rng(1))
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual((row["year"], row["n_closed"]), (2020, 1))
        for m in ("wcc", "scc", "reach", "strength"):
            self.assertIn(f"predicted_{m}", row)
            self.assertIsNotNone(row[f"observed_{m}"])

    def test_skips_without_pre_or_closure(self) -> None:
        from network.robustness import ban_replay_rows

        graphs = self._graphs()
        # No pre-wave graph (2020 wave needs 2019 — remove it).
        self.assertEqual(ban_replay_rows({2021: graphs[2021]}, {2020: {"h"}}), [])
        # Closure absent from the pre-wave graph.
        self.assertEqual(ban_replay_rows(graphs, {2020: {"zzz"}}), [])

    def test_observed_none_when_no_post_year(self) -> None:
        from network.robustness import ban_replay_rows

        graphs = {2019: self._graphs()[2019]}
        rows = ban_replay_rows(graphs, {2020: {"h"}}, n_random_runs=3, rng=np.random.default_rng(1))
        self.assertEqual(len(rows), 1)
        self.assertIsNone(rows[0]["observed_wcc"])


class AlphaSensitivityTests(TestCase):
    def test_sweep_populates_rows_per_alpha(self) -> None:
        from network.robustness import RobustnessConfig, run_robustness

        g = nx.gnp_random_graph(25, 0.25, seed=3, directed=True)
        for u, v in g.edges():
            g.edges[u, v]["weight"] = 1.0 + (hash((u, v)) % 5)
        cfg = RobustnessConfig(
            alpha=0.05,
            strategies=["random", "pagerank"],
            n_random_runs=3,
            n_null=0,
            seed=1,
            reach_sample=10,
            alpha_grid=[0.0, 0.05, 0.1],
        )
        out = run_robustness(g, config=cfg)
        sens = out["alpha_sensitivity"]
        self.assertIsNotNone(sens)
        self.assertEqual([r["alpha"] for r in sens["rows"]], [0.0, 0.05, 0.1])
        for r in sens["rows"]:
            self.assertIn("pagerank", r["r"])
            self.assertIn("wcc", r["r"]["pagerank"])

    def test_config_rejects_out_of_range_grid(self) -> None:
        from network.robustness import RobustnessConfig

        with self.assertRaises(ValueError):
            RobustnessConfig(alpha_grid=[1.0])
        with self.assertRaises(ValueError):
            RobustnessConfig(null_model="bogus")


class BanWaveLabelAnnotationTests(TestCase):
    def _comm(self) -> dict:
        # communities_data shape: {strategy_key: {"groups": [(cid, count, name, colour), ...]}}
        return {
            "labelgroup2": {"groups": [(18, 10, "Germany", "#fff"), (9, 8, "UK", "#eee")]},
            "leiden_directed": {"groups": [("1", 13, "1-leiden_directed", "#ddd")]},
        }

    def test_labelgroup_rows_get_names_algorithmic_left_bare(self) -> None:
        from network.management.commands.structural_analysis import _annotate_ban_wave_labels

        payload = {
            "ban_waves": {
                "labelgroup2": [{"community": "18"}, {"community": "9"}],
                "leiden_directed": [{"community": "1"}],
            }
        }
        _annotate_ban_wave_labels(payload, self._comm())
        lg = payload["ban_waves"]["labelgroup2"]
        self.assertEqual([r.get("community_label") for r in lg], ["Germany", "UK"])
        # Algorithmic auto-name "<id>-<strategy>" is not applied.
        self.assertNotIn("community_label", payload["ban_waves"]["leiden_directed"][0])

    def test_tolerates_none_and_unknown_strategies(self) -> None:
        from network.management.commands.structural_analysis import _annotate_ban_wave_labels

        _annotate_ban_wave_labels(None, self._comm())  # no crash
        payload = {"ban_waves": {"unknown": [{"community": "x"}]}}
        _annotate_ban_wave_labels(payload, self._comm())
        self.assertNotIn("community_label", payload["ban_waves"]["unknown"][0])


class WriteRobustnessJsonTests(TestCase):
    def test_round_trip_under_tempdir(self) -> None:
        from network.exporter import write_robustness_json

        payload = {"strategies": {"pagerank": {"r_wcc": 0.42}}, "config": {"seed": 1}}
        with tempfile.TemporaryDirectory() as tmp:
            write_robustness_json(payload, tmp)
            path = os.path.join(tmp, "data", "robustness.json")
            self.assertTrue(os.path.isfile(path))
            with open(path) as f:
                decoded = json.load(f)
            self.assertEqual(decoded, payload)


class WriteRobustnessTableXlsxTests(TestCase):
    def _payload(self) -> dict[str, Any]:
        return {
            "config": {"seed": 0},
            "graph": {"n": 3, "m": 3, "alpha": 0.05, "backbone_n": 3, "backbone_m": 3, "filtered": True},
            "efficiency": {
                "baseline": 1.0,
                "fractions": [0.0, 0.5, 1.0],
                "curves": {"pagerank": [1.0, 0.4, 0.0]},
            },
            "strategies": {
                "pagerank": {
                    "curve_wcc": [1.0, 0.5, 0.0, 0.0],
                    "curve_scc": [1.0, 0.5, 0.0, 0.0],
                    "curve_reach": [0.5, 0.25, 0.0, 0.0],
                    "curve_strength": [1.0, 0.4, 0.0, 0.0],
                    "r_wcc": 0.125,
                    "r_scc": 0.125,
                    "r_reach": 0.0625,
                    "r_strength": 0.1,
                    "fc_wcc": 0.5,
                    "fc_scc": 0.5,
                    "fc_reach": 0.5,
                    "fc_strength": 0.5,
                    "null": {
                        "r_wcc": {"mean": 0.12, "std": 0.01, "z": 0.5, "p": 0.7},
                        "r_scc": {"mean": 0.12, "std": 0.01, "z": 0.5, "p": 0.7},
                        "r_reach": {"mean": 0.06, "std": 0.005, "z": 1.25, "p": 0.095},
                        "r_strength": {"mean": 0.1, "std": 0.01, "z": 0.0, "p": 1.0},
                        "curve_wcc_mean": [1.0, 0.45, 0.0, 0.0],
                        "curve_wcc_std": [0.0, 0.05, 0.0, 0.0],
                        "curve_scc_mean": [1.0, 0.45, 0.0, 0.0],
                        "curve_scc_std": [0.0, 0.05, 0.0, 0.0],
                        "curve_reach_mean": [0.5, 0.2, 0.0, 0.0],
                        "curve_reach_std": [0.0, 0.05, 0.0, 0.0],
                        "curve_strength_mean": [1.0, 0.35, 0.0, 0.0],
                        "curve_strength_std": [0.0, 0.05, 0.0, 0.0],
                    },
                },
            },
            "modular": {
                "leiden": {
                    "pagerank": {
                        "intra": [1.0, 0.5, 0.0, 0.0],
                        "inter": [1.0, 0.0, 0.0, 0.0],
                        "ratio": [1.0, None, None, None],
                    }
                }
            },
            "ban_waves": {
                "leiden": [
                    {
                        "community": "0",
                        "n": 2,
                        "fraction": 0.667,
                        "s_wcc": 0.333,
                        "random_wcc": 0.5,
                        "s_scc": 0.0,
                        "random_scc": 0.1,
                        "s_reach": 0.0,
                        "random_reach": 0.05,
                        "s_strength": 0.2,
                        "random_strength": 0.3,
                    }
                ]
            },
        }

    def test_workbook_has_summary_curve_and_modular_sheets(self) -> None:
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(self._payload(), output_filename=out, project_title="Test")
            self.assertTrue(os.path.isfile(out))
            wb = openpyxl.load_workbook(out)
            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("Efficiency", wb.sheetnames)
            self.assertTrue(any(name.startswith("Curve") for name in wb.sheetnames))
            self.assertTrue(any(name.startswith("Modular") for name in wb.sheetnames))
            self.assertTrue(any(name.startswith("Ban wave") for name in wb.sheetnames))

    def test_summary_sheet_has_one_row_per_strategy_metric_combo(self) -> None:
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(self._payload(), output_filename=out)
            wb = openpyxl.load_workbook(out)
            ws = wb["Summary"]
            self.assertEqual(ws.max_row, 5)  # header + 4 metrics × 1 strategy
            r_values = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
            self.assertIn(0.125, r_values)
            self.assertIn(0.0625, r_values)
            self.assertIn(0.1, r_values)
            self.assertEqual(ws.cell(row=1, column=7).value, "p")
            p_values = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
            self.assertIn(0.095, p_values)

    def test_ban_wave_sheet_rows_match_payload(self) -> None:
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(self._payload(), output_filename=out)
            wb = openpyxl.load_workbook(out)
            ws = wb["Ban wave leiden"]
            self.assertEqual(ws.max_row, 2)  # header + 1 block
            self.assertEqual(ws.cell(row=2, column=1).value, "0")
            self.assertEqual(ws.cell(row=2, column=2).value, 2)

    def test_ban_wave_sheet_prefers_community_label(self) -> None:
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        payload = self._payload()
        payload["ban_waves"]["leiden"][0]["community_label"] = "Germany"
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(payload, output_filename=out)
            wb = openpyxl.load_workbook(out)
            ws = wb["Ban wave leiden"]
            self.assertEqual(ws.cell(row=2, column=1).value, "Germany")

    def test_handles_payload_without_null(self) -> None:
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        payload = self._payload()
        payload["strategies"]["pagerank"]["null"] = None
        payload.pop("modular")
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(payload, output_filename=out)
            wb = openpyxl.load_workbook(out)
            self.assertIn("Summary", wb.sheetnames)

    def test_year_data_produces_year_suffixed_sheets(self) -> None:
        # With year_data the workbook becomes one contiguous block of sheets per
        # scope: All first, then each year.  Sheet names get a space-separated
        # suffix; the legacy "Summary" sheet (no suffix) must not appear.
        from network.tables import write_robustness_table_xlsx

        import openpyxl

        global_payload = self._payload()
        year_payload = self._payload()
        year_data = [(2019, year_payload), (2020, year_payload)]
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "robustness_table.xlsx")
            write_robustness_table_xlsx(global_payload, output_filename=out, year_data=year_data)
            wb = openpyxl.load_workbook(out)
            for expected in (
                "Summary All",
                "Summary 2019",
                "Summary 2020",
                "Curve pagerank All",
                "Curve pagerank 2019",
                "Curve pagerank 2020",
                "Modular leiden All",
                "Modular leiden 2019",
                "Modular leiden 2020",
            ):
                self.assertIn(expected, wb.sheetnames, msg=f"missing sheet {expected!r}")
            # Legacy un-suffixed sheets must not coexist with the year-grouped layout.
            self.assertNotIn("Summary", wb.sheetnames)
            self.assertNotIn("Curve pagerank", wb.sheetnames)

    def test_sheet_name_helper_respects_31_char_cap(self) -> None:
        # Long partition name + year suffix must stay within Excel's 31-char limit.
        from network.tables import _robustness_sheet_name

        self.assertEqual(_robustness_sheet_name("Summary", ""), "Summary")
        self.assertEqual(_robustness_sheet_name("Summary", "All"), "Summary All")
        long_partition = "Modular leiden_directed_super_long"
        out = _robustness_sheet_name(long_partition, "2020")
        self.assertLessEqual(len(out), 31)
        self.assertTrue(out.endswith(" 2020"))


# ---------------------------------------------------------------------------
# network.layout.resolve_iterations
# ---------------------------------------------------------------------------


class ResolveIterationsTests(TestCase):
    """fa2_iterations accepts an integer or an Nx multiplier; floored at 100."""

    def test_integer_returned_as_is(self) -> None:
        from network.layout import resolve_iterations

        self.assertEqual(resolve_iterations(5000, num_nodes=100), 5000)
        self.assertEqual(resolve_iterations("5000", num_nodes=100), 5000)

    def test_multiplier_scaled_by_node_count(self) -> None:
        from network.layout import resolve_iterations

        self.assertEqual(resolve_iterations("7x", num_nodes=1000), 7000)
        self.assertEqual(resolve_iterations("2.5x", num_nodes=400), 1000)

    def test_floored_at_100(self) -> None:
        from network.layout import resolve_iterations

        self.assertEqual(resolve_iterations(50, num_nodes=1000), 100)
        self.assertEqual(resolve_iterations("0.01x", num_nodes=10), 100)
        self.assertEqual(resolve_iterations("7x", num_nodes=0), 100)

    def test_default_used_when_blank_or_none(self) -> None:
        from network.layout import FA2_ITERATIONS_DEFAULT, resolve_iterations

        # FA2_ITERATIONS_DEFAULT is "7x", so 7 × 50 = 350.
        self.assertEqual(FA2_ITERATIONS_DEFAULT, "7x")
        self.assertEqual(resolve_iterations(None, num_nodes=50), 350)
        self.assertEqual(resolve_iterations("", num_nodes=50), 350)

    def test_case_insensitive_x(self) -> None:
        from network.layout import resolve_iterations

        self.assertEqual(resolve_iterations("3X", num_nodes=200), 600)


class ComputeInterestStructuralWindowTests(TestCase):
    """Window-scoping the structural layer.

    Builds a tiny network of three in-target channels (A→B and a recent fwd
    from C→B), then asserts that:
      * a window_filter on the forwarder date restricts which forwards are
        counted in C and D;
      * an interest_score_override map replaces the persisted
        Message.interest_score in the payload;
      * the payload emits hot_layer_scope / structural_scope / policy.
    """

    def _build(self) -> tuple[Any, Any, Any, Any]:
        label = make_label("In target")
        a = make_channel(telegram_id=9001, title="A", label=label)
        b = make_channel(telegram_id=9002, title="B", label=label)
        c = make_channel(telegram_id=9003, title="C", label=label)
        return label, a, b, c

    def _origin_and_forwards(self, a: Any, b: Any, c: Any) -> Any:
        # Origin post on A.
        origin = Message.objects.create(
            telegram_id=500_000,
            channel=a,
            date=datetime.datetime(2024, 1, 15, tzinfo=datetime.UTC),
            views=100,
            forwards=10,
            total_reactions=5,
        )
        # Old in-window forward by B (Jan 2024).
        Message.objects.create(
            telegram_id=600_001,
            channel=b,
            date=datetime.datetime(2024, 1, 16, tzinfo=datetime.UTC),
            forwarded_from=a,
            fwd_from_channel_post=origin.telegram_id,
        )
        # Out-of-window forward by C (Jan 2025).
        Message.objects.create(
            telegram_id=700_001,
            channel=c,
            date=datetime.datetime(2025, 1, 16, tzinfo=datetime.UTC),
            forwarded_from=a,
            fwd_from_channel_post=origin.telegram_id,
        )
        return origin

    def _graph_data(self, a: Any, b: Any, c: Any) -> tuple[dict, dict]:
        # Minimal GraphData + channel_dict shape used by compute_interest_structural.
        graph_data = {
            "nodes": [
                {"id": str(a.pk), "pagerank": 0.3},
                {"id": str(b.pk), "pagerank": 0.5},
                {"id": str(c.pk), "pagerank": 0.2},
            ],
            "edges": [],
        }
        channel_dict = {
            str(a.pk): {"channel": a, "data": {"communities": {"leiden_directed": "alpha"}}},
            str(b.pk): {"channel": b, "data": {"communities": {"leiden_directed": "beta"}}},
            str(c.pk): {"channel": c, "data": {"communities": {"leiden_directed": "gamma"}}},
        }
        return graph_data, channel_dict

    def test_window_filter_excludes_out_of_window_forwarders(self) -> None:
        from network.interest_structural import compute_interest_structural

        _org, a, b, c = self._build()
        origin = self._origin_and_forwards(a, b, c)
        graph_data, channel_dict = self._graph_data(a, b, c)

        # No window → both forwards counted (C reach = 2 communities).
        all_time = compute_interest_structural(
            graph_data,
            channel_dict,
            community_strategy="LEIDEN_DIRECTED",
            window_days=0,
        )
        self.assertEqual(all_time["structural_scope"], "all-time")
        self.assertEqual(all_time["hot_layer_scope"], "all-time")
        all_time_rec = next(
            r for r in all_time["by_message"] if r["telegram_id"] == origin.telegram_id and r["channel_pk"] == a.pk
        )
        self.assertEqual(all_time_rec["forwarder_count_in_target"], 2)
        self.assertEqual(all_time_rec["c_cross_community"], 2)

        # 2024-only window → only B's forward counts (C reach = 1).
        windowed = compute_interest_structural(
            graph_data,
            channel_dict,
            community_strategy="LEIDEN_DIRECTED",
            window_days=0,
            window_filter={
                # Production shape: the ``__date`` transform (see
                # _date_window_filter). Compares calendar dates, so a bare date
                # is warning-free under USE_TZ — unlike a plain ``date__gte``.
                "date__date__gte": datetime.date(2024, 1, 1),
                "date__date__lte": datetime.date(2024, 12, 31),
            },
        )
        self.assertIn("window 2024-01-01..2024-12-31", windowed["structural_scope"])
        windowed_rec = next(
            r for r in windowed["by_message"] if r["telegram_id"] == origin.telegram_id and r["channel_pk"] == a.pk
        )
        self.assertEqual(windowed_rec["forwarder_count_in_target"], 1)
        self.assertEqual(windowed_rec["c_cross_community"], 1)

    def test_override_replaces_persisted_interest_score(self) -> None:
        from network.interest_structural import compute_interest_structural

        _org, a, b, c = self._build()
        origin = self._origin_and_forwards(a, b, c)
        # Persist a global score on the origin.
        Message.objects.filter(pk=origin.pk).update(interest_score=0.99)

        graph_data, channel_dict = self._graph_data(a, b, c)
        payload = compute_interest_structural(
            graph_data,
            channel_dict,
            community_strategy="LEIDEN_DIRECTED",
            window_days=0,
            interest_score_override={(a.pk, origin.telegram_id): -0.42},
        )
        rec = next(
            r for r in payload["by_message"] if r["telegram_id"] == origin.telegram_id and r["channel_pk"] == a.pk
        )
        # Override beat the persisted 0.99.
        self.assertAlmostEqual(rec["interest_score"], -0.42, places=6)
        # Override without window labels the hot layer as a custom override.
        self.assertEqual(payload["hot_layer_scope"], "overridden")
        self.assertEqual(payload["forwarder_window_policy"], "forwarder-date")


class ScopeLabelTests(TestCase):
    """_scope_label renders the export window from either ORM filter-key shape."""

    def _label(self, window_filter: Any) -> str:
        from network.interest_structural import _scope_label

        return _scope_label(window_filter)

    def test_none_and_empty_are_all_time(self) -> None:
        self.assertEqual(self._label(None), "all-time")
        self.assertEqual(self._label({}), "all-time")

    def test_production_date_transform_keys(self) -> None:
        # The shape _date_window_filter actually builds.
        label = self._label(
            {"date__date__gte": datetime.date(2024, 1, 1), "date__date__lte": datetime.date(2024, 12, 31)}
        )
        self.assertEqual(label, "window 2024-01-01..2024-12-31")

    def test_plain_date_keys_fallback(self) -> None:
        label = self._label(
            {"date__gte": datetime.datetime(2024, 3, 1, tzinfo=datetime.UTC), "date__lt": datetime.date(2024, 4, 1)}
        )
        self.assertEqual(label, "window 2024-03-01..2024-04-01")

    def test_open_ended_bounds(self) -> None:
        self.assertEqual(self._label({"date__date__gte": datetime.date(2024, 1, 1)}), "window 2024-01-01..…")
        self.assertEqual(self._label({"date__date__lte": datetime.date(2024, 12, 31)}), "window …..2024-12-31")

    def test_filter_without_date_keys_is_generic_windowed(self) -> None:
        self.assertEqual(self._label({"channel_id__in": [1, 2]}), "windowed")


class ResolveWindowLabelTests(TestCase):
    """Representative label = longest in-window duration; tiebreak = earliest start; None bounds clamp."""

    @staticmethod
    def _resolve(periods, window_start, window_end, created=None, data_min=None, data_max=None):
        return resolve_window_label(periods, window_start, window_end, created, data_min, data_max)

    def test_longest_duration_wins(self) -> None:
        d = datetime.date
        periods = [
            (1, "A", "#aaaaaa", d(2024, 1, 1), d(2024, 1, 31)),  # 31 days
            (2, "B", "#bbbbbb", d(2024, 2, 1), d(2024, 12, 31)),  # ~334 days
        ]
        self.assertEqual(self._resolve(periods, d(2024, 1, 1), d(2024, 12, 31))[0], 2)

    def test_tiebreak_earliest_start(self) -> None:
        d = datetime.date
        periods = [  # equal 10-day in-window spans → earliest start (label 1) wins
            (1, "A", "#aaaaaa", d(2024, 1, 1), d(2024, 1, 10)),
            (2, "B", "#bbbbbb", d(2024, 2, 1), d(2024, 2, 10)),
        ]
        self.assertEqual(self._resolve(periods, d(2024, 1, 1), d(2024, 12, 31))[0], 1)

    def test_none_bounds_clamped_to_data_range(self) -> None:
        d = datetime.date
        periods = [(7, "Open", "#cccccc", None, None)]
        result = self._resolve(
            periods, None, None, created=d(2024, 1, 1), data_min=d(2024, 1, 1), data_max=d(2024, 6, 1)
        )
        self.assertEqual(result[0], 7)

    def test_no_periods_returns_none(self) -> None:
        self.assertIsNone(self._resolve([], datetime.date(2024, 1, 1), datetime.date(2024, 12, 31)))


@override_settings(TIME_ZONE="UTC")
class ChannelCutoffQBoundaryTests(TestCase):
    """channel_cutoff_q inclusive day boundaries via the Exists subquery (SQLite-safe).

    TIME_ZONE is pinned: the ``__date`` lookups bucket message datetimes into
    *active-timezone* calendar days, so these UTC-midnight boundary fixtures only
    express "inclusive bounds" under a UTC clock — under the deployed Europe/Rome
    zone the same instants belong to the neighbouring local day, and the test
    would depend on the machine's configuration.
    """

    def setUp(self) -> None:
        self.label = make_label("O")
        self.ch = make_channel(
            telegram_id=1,
            label=self.label,
            attribution_start=datetime.date(2024, 1, 1),
            attribution_end=datetime.date(2024, 3, 31),
        )

    def _msg(self, tid, dt) -> None:
        Message.objects.create(telegram_id=tid, channel=self.ch, date=dt)

    def test_boundaries_inclusive(self) -> None:
        utc = datetime.timezone.utc
        self._msg(1, datetime.datetime(2023, 12, 31, 23, 59, tzinfo=utc))  # day before start → out
        self._msg(2, datetime.datetime(2024, 1, 1, 0, 0, tzinfo=utc))  # on start → in
        self._msg(3, datetime.datetime(2024, 3, 31, 23, 59, tzinfo=utc))  # on end (late) → in
        self._msg(4, datetime.datetime(2024, 4, 1, 0, 0, tzinfo=utc))  # day after end → out
        kept = set(Message.objects.filter(channel_cutoff_q()).values_list("telegram_id", flat=True))
        self.assertEqual(kept, {2, 3})

    def test_open_period_includes_everything(self) -> None:
        ch2 = make_channel(telegram_id=2, label=self.label)  # open (None, None) period
        Message.objects.create(
            telegram_id=99, channel=ch2, date=datetime.datetime(2010, 1, 1, tzinfo=datetime.timezone.utc)
        )
        self.assertTrue(Message.objects.filter(channel=ch2).filter(channel_cutoff_q()).exists())


# ---------------------------------------------------------------------------
# measures/_registry.py — parse_measures / MeasureInstance / canonical / companions
# ---------------------------------------------------------------------------


class MeasureParserTests(TestCase):
    def test_plain_and_parameterised_tokens(self) -> None:
        insts = parse_measures(["PAGERANK", "DIFFUSIONLAG(window=60)", "MODULEROLE(basis=leiden)"])
        self.assertEqual(
            [i.token() for i in insts], ["PAGERANK", "DIFFUSIONLAG(window=60)", "MODULEROLE(basis=LEIDEN)"]
        )
        self.assertEqual([i.suffix() for i in insts], ["", "_window_60", "_basis_leiden"])

    def test_all_expands_with_defaults(self) -> None:
        tokens = {i.measure for i in parse_measures(["ALL"])}
        self.assertIn("PAGERANK", tokens)
        self.assertIn("MODULEROLE", tokens)
        self.assertIn("DIFFUSIONLAG", tokens)

    def test_same_measure_twice_with_different_params(self) -> None:
        insts = parse_measures(["DIFFUSIONLAG(window=30)", "DIFFUSIONLAG(window=60)"])
        self.assertEqual([i.suffix() for i in insts], ["_window_30", "_window_60"])

    def test_default_override_fills_bare_token(self) -> None:
        (inst,) = parse_measures(["DIFFUSIONLAG"], defaults={"DIFFUSIONLAG": {"window": 45}})
        self.assertEqual(inst.token(), "DIFFUSIONLAG(window=45)")

    def test_rejects_duplicate_drop_once(self) -> None:
        with self.assertRaisesRegex(ValueError, "more than once"):
            parse_measures(["PAGERANK", "PAGERANK"])

    def test_rejects_identical_parameterised(self) -> None:
        with self.assertRaisesRegex(ValueError, "identical parameters"):
            parse_measures(["DIFFUSIONLAG(window=30)", "DIFFUSIONLAG(window=30)"])

    def test_rejects_params_on_drop_once(self) -> None:
        with self.assertRaisesRegex(ValueError, "takes no parameters"):
            parse_measures(["PAGERANK(window=5)"])

    def test_rejects_out_of_range_and_unknown_param(self) -> None:
        with self.assertRaisesRegex(ValueError, "below the minimum"):
            parse_measures(["DIFFUSIONLAG(window=-1)"])
        with self.assertRaisesRegex(ValueError, "has no parameter"):
            parse_measures(["DIFFUSIONLAG(foo=1)"])

    def test_rejects_unknown_measure(self) -> None:
        with self.assertRaisesRegex(ValueError, "Unknown measure"):
            parse_measures(["NOTAMEASURE"])

    def test_module_role_basis_is_free_form(self) -> None:
        # MODULEROLE's basis is a free-form str (a strategy name *or* a LABELGROUP<id> token),
        # not an enum, so the parser accepts any value; an unknown basis is caught later by the
        # command's _validate_settings cross-check against --community-strategies.
        (inst,) = parse_measures(["MODULEROLE(basis=NOPE)"])
        self.assertEqual(inst.params_dict["basis"], "NOPE")

    def test_canonical_measure_key(self) -> None:
        self.assertEqual(canonical_measure_key("within_module_z_basis_leiden_directed"), "within_module_z")
        self.assertEqual(canonical_measure_key("participation_basis_leiden_directed"), "participation")
        self.assertEqual(canonical_measure_key("diffusion_lag_window_60"), "diffusion_lag")
        self.assertEqual(canonical_measure_key("pagerank"), "pagerank")  # non-parameterised unchanged

    def test_role_companions(self) -> None:
        mod = role_companions("within_module_z_basis_leiden")
        self.assertEqual(mod["role_key"], "module_role_basis_leiden")
        self.assertEqual(mod["count_keys"], [])
        self.assertIsNone(role_companions("pagerank"))


class MeasureComputeHelpersTests(TestCase):
    """Command-layer helpers that suffix node keys per instance and resolve a community basis."""

    def test_rebind_suffixes_numeric_and_aux_keys(self) -> None:
        from network.management.commands.structural_analysis import _rebind_measure_keys

        gd = {
            "nodes": [
                {"id": "1", "within_module_z": 1.2, "module_role": "Connector"},
                {"id": "2", "within_module_z": None, "module_role": None},
            ]
        }
        (inst,) = parse_measures(["MODULEROLE(basis=LEIDEN_DIRECTED)"])
        labels = _rebind_measure_keys(gd, inst, [("within_module_z", "Within-module z")])
        # Numeric column suffixed and label annotated.
        self.assertEqual(labels, [("within_module_z_basis_leiden_directed", "Within-module z (basis=leiden_directed)")])
        # Both the numeric key and the categorical companion are renamed on every node.
        self.assertEqual(gd["nodes"][0]["within_module_z_basis_leiden_directed"], 1.2)
        self.assertEqual(gd["nodes"][0]["module_role_basis_leiden_directed"], "Connector")
        self.assertNotIn("within_module_z", gd["nodes"][0])
        self.assertNotIn("module_role", gd["nodes"][0])

    def test_rebind_noop_for_unparameterised(self) -> None:
        from network.management.commands.structural_analysis import _rebind_measure_keys

        gd = {"nodes": [{"id": "1", "pagerank": 0.5}]}
        (inst,) = parse_measures(["PAGERANK"])
        labels = _rebind_measure_keys(gd, inst, [("pagerank", "PageRank")])
        self.assertEqual(labels, [("pagerank", "PageRank")])
        self.assertEqual(gd["nodes"][0]["pagerank"], 0.5)

    def test_resolve_basis_explicit_and_auto(self) -> None:
        from network.management.commands.structural_analysis import _resolve_community_basis

        (explicit,) = parse_measures(["MODULEROLE(basis=LEIDEN)"])
        self.assertEqual(_resolve_community_basis(explicit, ["leiden", "labelgroup_1"]), "leiden")
        # Explicit basis that was not computed → None (skip).
        self.assertIsNone(_resolve_community_basis(explicit, ["labelgroup_1"]))
        # Auto (blank) module role prefers leiden_directed.
        (auto_mod,) = parse_measures(["MODULEROLE"])
        self.assertEqual(_resolve_community_basis(auto_mod, ["leiden_directed", "leiden"]), "leiden_directed")
